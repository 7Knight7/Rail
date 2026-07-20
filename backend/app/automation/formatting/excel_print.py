"""Excel print setup and column formatting for report output workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.automation.formatting.pdf_table import COMPACT_HEADERS, WIDE_HEADERS

COMPACT_COL_WIDTH = 10.0
MODERATE_COL_WIDTH = 14.0
WIDE_COL_WIDTH = 32.0
NARROW_MARGIN_INCHES = 0.25


def apply_report_print_setup(
    worksheet: Worksheet,
    *,
    col_count: int,
    last_row: int | None = None,
    use_a3: bool = True,
) -> None:
    """Configure landscape print: fit one page wide, rows paginate vertically."""
    if last_row is None:
        last_row = worksheet.max_row
    last_col = get_column_letter(col_count)
    worksheet.print_area = f"A1:{last_col}{last_row}"
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A3 if use_a3 else worksheet.PAPERSIZE_A4
    )
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = NARROW_MARGIN_INCHES
    worksheet.page_margins.right = NARROW_MARGIN_INCHES
    worksheet.page_margins.top = NARROW_MARGIN_INCHES
    worksheet.page_margins.bottom = NARROW_MARGIN_INCHES
    worksheet.print_options.horizontalCentered = True


def apply_column_formatting(
    worksheet: Worksheet,
    headers: list[str],
    *,
    header_row: int = 2,
    data_start_row: int = 3,
    wrap_all_data: bool = False,
) -> None:
    """Set column widths and wrap alignment for compact/moderate/wide fields."""
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(vertical="top", horizontal="center")

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        if header in WIDE_HEADERS:
            worksheet.column_dimensions[letter].width = WIDE_COL_WIDTH
            cell_align = wrap_align
        elif header in COMPACT_HEADERS or header == "S.No.":
            worksheet.column_dimensions[letter].width = COMPACT_COL_WIDTH
            cell_align = center_align if not wrap_all_data else wrap_align
        else:
            worksheet.column_dimensions[letter].width = MODERATE_COL_WIDTH
            cell_align = wrap_align if wrap_all_data or len(header) > 12 else center_align

        for row_idx in range(header_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                if wrap_all_data and row_idx >= data_start_row:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = cell_align
