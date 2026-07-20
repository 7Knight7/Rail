"""South Central Railway row highlighting (shared across reports)."""

from __future__ import annotations

import logging
import re

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.automation.utils import log_automation_event

logger = logging.getLogger(__name__)

SCR_PATTERN = re.compile(r"south\s*central\s*railway", re.IGNORECASE)
SCR_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
SCR_FONT = Font(color="000000")
NO_FILL = PatternFill(fill_type=None)
YELLOW_FILL_RGB = frozenset({"00FFFF00", "FFFF00", "FFFFFF00"})


def cell_has_yellow_fill(cell) -> bool:
    """Return True when the cell uses SCR-style yellow fill."""
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    return rgb in YELLOW_FILL_RGB


def clear_complaint_data_row_formatting(
    worksheet: Worksheet,
    *,
    start_row: int,
    end_row: int | None = None,
    start_col: int = 1,
    end_col: int | None = None,
) -> None:
    """Reports 5/6: reset data rows to no fill (clears template/inherited yellow)."""
    if end_row is None:
        end_row = worksheet.max_row
    if end_col is None:
        end_col = worksheet.max_column

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = NO_FILL


def mode_matches(expected_mode: str, mode_value: str) -> bool:
    """Match portal Mode codes (T/S) or full labels (Train/Station)."""
    raw = (mode_value or "").strip().lower()
    expected = (expected_mode or "").strip().lower()
    if not raw:
        return False
    if expected == "train":
        return raw in {"t", "train"} or "train" in raw
    if expected == "station":
        return raw in {"s", "station"} or "station" in raw
    return expected in raw


def row_contains_scr(values: list[object]) -> bool:
    """Return True if any cell in the row contains South Central Railway."""
    for value in values:
        if value is None:
            continue
        text = re.sub(r"\s+", " ", str(value).strip())
        if SCR_PATTERN.search(text):
            return True
    return False


def highlight_south_central_railway_rows(
    worksheet: Worksheet,
    *,
    start_row: int = 1,
    end_row: int | None = None,
    start_col: int = 1,
    end_col: int | None = None,
) -> int:
    """Apply yellow fill and black text to entire rows containing SCR."""
    if end_row is None:
        end_row = worksheet.max_row
    if end_col is None:
        end_col = worksheet.max_column

    highlighted = 0
    for row_idx in range(start_row, end_row + 1):
        values = [
            worksheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(start_col, end_col + 1)
        ]
        if not row_contains_scr(values):
            continue

        log_automation_event(logger, "scr_row_detected", row=row_idx)
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = SCR_FILL
            cell.font = SCR_FONT
        highlighted += 1
        log_automation_event(logger, "scr_row_highlighted", row=row_idx)

    if highlighted:
        log_automation_event(logger, "scr_highlight_count", count=highlighted)
    return highlighted
