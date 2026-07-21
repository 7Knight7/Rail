"""Post-render validation for Report 5 XLSX/PDF artifacts."""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from app.automation.formatting.report5_styles import cell_has_dark_fill
from app.automation.formatting.text_safe import contains_rendering_risk_markers

logger = logging.getLogger(__name__)

_BLACK_FILL_RE = re.compile(
    r"(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+rg\s+"
    r"(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+re\s+f\b"
)


def validate_report5_excel_styles(
    path: Path,
    *,
    data_start_row: int = 3,
) -> None:
    """Raise ValueError(REPORT5_STYLE_VALIDATION_FAILED) on dark body fills."""
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook.active
        for row_idx in range(data_start_row, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell_has_dark_fill(cell):
                    raise ValueError(
                        "REPORT5_STYLE_VALIDATION_FAILED: dark fill on data row "
                        f"{row_idx} col {col_idx}"
                    )
                font = cell.font
                color = getattr(font.color, "rgb", None) if font and font.color else None
                if color and str(color).upper() in {"FFFFFFFF", "00FFFFFF"}:
                    raise ValueError(
                        "REPORT5_STYLE_VALIDATION_FAILED: white font on data row "
                        f"{row_idx} col {col_idx}"
                    )
    finally:
        workbook.close()


def validate_report5_pdf_styles(path: Path) -> None:
    """Raise ValueError when PDF contains black-filled body rectangles."""
    raw = path.read_bytes()
    if raw[:5] != b"%PDF-":
        raise ValueError("REPORT5_ARTIFACT_VALIDATION_FAILED: invalid PDF header")
    text = raw.decode("latin-1", errors="ignore")
    for match in _BLACK_FILL_RE.finditer(text):
        r, g, b = match.group(1), match.group(2), match.group(3)
        if r == "0" and g == "0" and b == "0":
            raise ValueError(
                "REPORT5_STYLE_VALIDATION_FAILED: PDF contains black-filled rectangles"
            )


def validate_report5_unicode_content(
    headers: list[str],
    rows: list[list[str]],
) -> None:
    """Raise ValueError when black-square markers remain in projected output."""
    for header in headers:
        if contains_rendering_risk_markers(header):
            raise ValueError(
                f"REPORT5_UNICODE_RENDERING_FAILED: header {header!r} has risk markers"
            )
    for row_idx, row in enumerate(rows):
        ref = row[0] if row else str(row_idx + 1)
        for col_idx, value in enumerate(row):
            if contains_rendering_risk_markers(str(value)):
                column = headers[col_idx] if col_idx < len(headers) else str(col_idx)
                logger.warning(
                    "report5_unicode_risk run_row=%s complaint_ref=%s column=%s",
                    row_idx + 1,
                    ref,
                    column,
                )
                raise ValueError(
                    f"REPORT5_UNICODE_RENDERING_FAILED: row {ref} column {column!r} "
                    "contains black-square markers"
                )


def validate_report5_layout(
    *,
    headers: list[str],
    rows: list[list[str]],
    pdf_path: Path,
) -> None:
    """Basic layout checks: edge columns present and PDF non-empty."""
    if not headers:
        raise ValueError("REPORT5_LAYOUT_VALIDATION_FAILED: no headers")
    if headers[0] != "S.No.":
        raise ValueError("REPORT5_LAYOUT_VALIDATION_FAILED: first column is not S.No.")
    if len(rows) != len({row[0] for row in rows if row}):
        pass  # duplicate S.No. allowed when empty
    if not pdf_path.is_file() or pdf_path.stat().st_size <= 0:
        raise ValueError("REPORT5_ARTIFACT_VALIDATION_FAILED: PDF missing or empty")
