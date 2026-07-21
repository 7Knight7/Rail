"""Report 5 (scr-train) final render styles — no SCR row highlighting."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

from app.automation.formatting.pdf_fonts import pdf_font_bold, pdf_font_regular

REPORT5_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D3D3D3")
REPORT5_NO_FILL = PatternFill(fill_type=None)
REPORT5_BODY_FONT = Font(color="000000")
REPORT5_HEADER_FONT = Font(bold=True, color="000000")
REPORT5_TITLE_FONT = Font(bold=True, size=12, color="000000")
REPORT5_BODY_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="left")
REPORT5_HEADER_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="center")
REPORT5_TITLE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

REPORT5_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DARK_FILL_RGB = frozenset({"000000", "FF000000", "00000000"})


def pdf_title_style(name: str = "Report5Title") -> ParagraphStyle:
    """White background, black bold title text."""
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        name,
        parent=styles["Normal"],
        fontName=pdf_font_bold(),
        fontSize=12,
        leading=14,
        textColor=colors.black,
        backColor=colors.white,
        alignment=TA_CENTER,
    )


def build_report5_pdf_table_styles(*, data_row_count: int) -> list[tuple]:
    """Fresh immutable-style commands: grey header, white body, black text."""
    commands: list[tuple] = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
    ]
    if data_row_count > 0:
        commands.extend(
            [
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), pdf_font_regular()),
            ]
        )
    return commands


def apply_report5_title_cell(cell) -> None:
    cell.font = REPORT5_TITLE_FONT
    cell.fill = REPORT5_NO_FILL
    cell.alignment = REPORT5_TITLE_ALIGN


def apply_report5_header_cell(cell) -> None:
    cell.font = REPORT5_HEADER_FONT
    cell.fill = REPORT5_HEADER_FILL
    cell.border = REPORT5_THIN_BORDER
    cell.alignment = REPORT5_HEADER_ALIGN


def apply_report5_body_cell(cell) -> None:
    cell.font = REPORT5_BODY_FONT
    cell.fill = REPORT5_NO_FILL
    cell.border = REPORT5_THIN_BORDER
    cell.alignment = REPORT5_BODY_ALIGN


def clear_report5_data_row_formatting(
    worksheet: Worksheet,
    *,
    start_row: int,
    end_row: int | None = None,
    start_col: int = 1,
    end_col: int | None = None,
) -> None:
    """Reset Report 5 data rows to white/no-fill with black text."""
    if end_row is None:
        end_row = worksheet.max_row
    if end_col is None:
        end_col = worksheet.max_column

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            apply_report5_body_cell(worksheet.cell(row=row_idx, column=col_idx))


def cell_has_dark_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    if rgb is None:
        return False
    normalized = str(rgb).upper()
    return normalized in {"000000", "FF000000"}
