"""Report 10-13 post-ingestion processor (Comprehensive Reports).

Processes four sections extracted from the Comprehensive (with drill down) page:
- Report 10: C&W complaints division wise
- Report 11: Security complaints
- Report 12: Punctuality complaints
- Report 13: Electrical Equipment complaints division wise

Each section produces a stacked table with:
- Section heading
- Date range
- Selected columns
- Data rows sorted by Received descending
- Total row

Output: One combined XLSX and one combined PDF.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.automation.config import config
from app.automation.comprehensive1013_filters import (
    COMPREHENSIVE_1013_SECTION_IDS,
    SectionConfig,
    get_section_config_by_id,
)
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT, build_fitted_table
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.processing.comprehensive_output_columns import (
    ADDITIVE_COLUMNS,
    COMPREHENSIVE_COLUMN_IDS,
    COMPREHENSIVE_COLUMN_LABELS,
    NON_ADDITIVE_COLUMNS,
    column_labels,
    default_column_ids,
    normalize_header_to_column_id,
)
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_report_dir,
)

# Report 10-13 PDF only: compact margins/spacing so four sections fit one page.
_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 14.0)
_PDF_SECTION_GAP_PT = 8.0
_PDF_AFTER_HEADING_PT = 4.0
_PDF_TITLE_AFTER_PT = 6.0
_PDF_HEIGHT_BUFFER_PT = 12.0


def _escape_paragraph_xml(text: str) -> str:
    """Escape XML special characters for reportlab Paragraph markup."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _heading_table_centered_over_width(
    heading: str,
    table_width: float,
    style: ParagraphStyle,
) -> Table:
    """Render a section heading centred above a table of the given width."""
    centered_style = ParagraphStyle(
        name=f"{style.name}Centered",
        parent=style,
        alignment=TA_CENTER,
    )
    heading_table = Table(
        [[Paragraph(_escape_paragraph_xml(heading), centered_style)]],
        colWidths=[table_width],
        hAlign="LEFT",
    )
    heading_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return heading_table

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "comprehensive1013_processor"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


@dataclass
class SectionDataset:
    """Data for a single comprehensive report section."""

    section_config: SectionConfig
    headers: list[str]
    rows: list[list[str]]
    total_row: list[str]
    raw_headers: list[str]
    column_ids: list[str]


class Comprehensive1013Processor:
    """Process Comprehensive Reports 10-13 and emit combined Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        if column_selection:
            log_automation_event(
                logger,
                "comprehensive1013_column_selection_received",
                column_selection=column_selection,
            )

        sections, total_input_rows = self._load_sections(source_a_path, column_selection)

        if not sections:
            return ProcessingResult(
                success=False,
                error="No section data found in combined index",
                source_a_path=str(source_a_path),
            )

        date_range = date_range_for_processing(column_selection)
        report_date = date_range.title_suffix()
        filename_suffix = date_range.filename_suffix()

        excel_dir = ensure_directory(resolve_report_dir(config.output_excel_dir, report_slug))
        pdf_dir = ensure_directory(resolve_report_dir(config.output_pdf_dir, report_slug))
        base_name = f"Rail_Madad_Report_10_13_Comprehensive_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(excel_path, sections, report_date=report_date)
            log_automation_event(logger, "comprehensive1013_excel_generated", excel_path=str(excel_path))
            self._write_pdf(pdf_path, sections, report_date=report_date)
            log_automation_event(logger, "comprehensive1013_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                input_row_count=total_input_rows,
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_a_rows=total_input_rows,
            )

        total_output_rows = sum(len(s.rows) for s in sections)

        section_row_counts = {s.section_config.section_id: len(s.rows) for s in sections}
        selected_columns_per_section = {
            s.section_config.section_id: s.column_ids for s in sections
        }

        log_automation_event(
            logger,
            "comprehensive1013_processing_completed",
            source_a=str(source_a_path),
            input_row_count=total_input_rows,
            section_count=len(sections),
            total_output_rows=total_output_rows,
            section_row_counts=section_row_counts,
        )

        return ProcessingResult(
            success=True,
            input_row_count=total_input_rows,
            processed_row_count=total_output_rows,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input_rows,
            output_columns=list(COMPREHENSIVE_COLUMN_IDS),
            visible_columns=[COMPREHENSIVE_COLUMN_LABELS[c] for c in COMPREHENSIVE_COLUMN_IDS],
            selected_column_ids=COMPREHENSIVE_COLUMN_IDS,
            column_order=list(COMPREHENSIVE_COLUMN_IDS),
            configuration_source="default",
        )

    def _load_sections(
        self,
        source_a_path: Path,
        column_selection: dict[str, Any] | None,
    ) -> tuple[list[SectionDataset], int]:
        """Load all sections from the combined index and their CSVs."""
        sections: list[SectionDataset] = []
        total_input_rows = 0

        index_entries = self._read_combined_index(source_a_path)
        if not index_entries:
            base_dir = source_a_path.parent
            for section_id in COMPREHENSIVE_1013_SECTION_IDS:
                section_config = get_section_config_by_id(section_id)
                if section_config is None:
                    continue
                csv_path = base_dir / f"{section_id}.csv"
                if csv_path.is_file():
                    index_entries[section_id] = {
                        "section_id": section_id,
                        "csv_path": str(csv_path),
                        "status": "success",
                    }

        for section_id in COMPREHENSIVE_1013_SECTION_IDS:
            section_config = get_section_config_by_id(section_id)
            if section_config is None:
                continue

            entry = index_entries.get(section_id)
            if entry is None or str(entry.get("status", "")).lower() != "success":
                log_automation_event(
                    logger,
                    "comprehensive1013_section_skipped",
                    section_id=section_id,
                    reason="not in index or failed",
                )
                continue

            csv_path = Path(str(entry.get("csv_path") or ""))
            if not csv_path.is_file():
                log_automation_event(
                    logger,
                    "comprehensive1013_section_csv_not_found",
                    section_id=section_id,
                    expected_path=str(csv_path),
                )
                continue

            raw_rows, raw_headers = self._read_csv(csv_path)
            if not raw_rows:
                continue

            data_rows, portal_total_row = self._split_total_row(raw_rows)
            total_input_rows += len(data_rows)

            selected_ids = self._resolve_column_ids(section_id, column_selection)
            projected_headers, projected_rows = self._project_columns(
                raw_headers, data_rows, selected_ids
            )
            total_row = self._compute_total_row(
                projected_headers, projected_rows, portal_total_row, raw_headers, selected_ids
            )

            sections.append(
                SectionDataset(
                    section_config=section_config,
                    headers=projected_headers,
                    rows=projected_rows,
                    total_row=total_row,
                    raw_headers=raw_headers,
                    column_ids=selected_ids,
                )
            )
            log_automation_event(
                logger,
                "comprehensive1013_section_loaded",
                section_id=section_id,
                row_count=len(projected_rows),
                selected_columns=selected_ids,
            )

        return sections, total_input_rows

    def _read_combined_index(self, source_a_path: Path) -> dict[str, dict[str, str]]:
        """Read the combined index CSV if it exists."""
        if source_a_path.name != "comprehensive_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                section_id = (row.get("section_id") or "").strip()
                if not section_id:
                    continue
                entries[section_id] = {
                    "section_id": section_id,
                    "section_name": (row.get("section_name") or "").strip(),
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "row_count": (row.get("row_count") or "0").strip(),
                    "status": (row.get("status") or "").strip(),
                    "error": (row.get("error") or "").strip(),
                }
        return entries

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        """Read a CSV file and return rows as dicts plus header list."""
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _split_total_row(
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        """Split data rows from total row (last row with 'Total' in Division column)."""
        if not rows:
            return [], None
        last_row = rows[-1]
        division_val = str(last_row.get("Division") or last_row.get("Organisation") or "").strip().lower()
        if division_val == "total":
            return rows[:-1], last_row
        return rows, None

    def _resolve_column_ids(
        self,
        section_id: str,
        column_selection: dict[str, Any] | None,
    ) -> list[str]:
        """Resolve which columns to include for a section."""
        if column_selection:
            section_columns = column_selection.get("sections", {}).get(section_id)
            if section_columns:
                selected = section_columns.get("selected_column_ids")
                if selected:
                    return list(selected)
            all_selected = column_selection.get("selected_column_ids")
            if all_selected:
                return list(all_selected)
        return default_column_ids()

    def _project_columns(
        self,
        raw_headers: list[str],
        data_rows: list[dict[str, str]],
        selected_ids: list[str],
    ) -> tuple[list[str], list[list[str]]]:
        """Project rows to only selected columns, returning headers and row lists."""
        header_to_id: dict[str, str] = {}
        for h in raw_headers:
            col_id = normalize_header_to_column_id(h)
            if col_id:
                header_to_id[h] = col_id

        output_headers = column_labels(selected_ids)
        output_rows: list[list[str]] = []

        for row_idx, row in enumerate(data_rows):
            output_row: list[str] = []
            for col_id in selected_ids:
                if col_id == "sno":
                    output_row.append(str(row_idx + 1))
                else:
                    value = ""
                    for h, cid in header_to_id.items():
                        if cid == col_id:
                            value = row.get(h, "")
                            break
                    output_row.append(value)
            output_rows.append(output_row)

        return output_headers, output_rows

    def _compute_total_row(
        self,
        projected_headers: list[str],
        projected_rows: list[list[str]],
        portal_total_row: dict[str, str] | None,
        raw_headers: list[str],
        selected_ids: list[str],
    ) -> list[str]:
        """Compute total row for a section."""
        header_to_id: dict[str, str] = {}
        for h in raw_headers:
            col_id = normalize_header_to_column_id(h)
            if col_id:
                header_to_id[h] = col_id

        total_row: list[str] = []
        for col_id in selected_ids:
            if col_id == "sno":
                total_row.append("")
            elif col_id == "division":
                total_row.append("Total")
            elif col_id == "share_percent":
                total_row.append("100.00")
            elif col_id in NON_ADDITIVE_COLUMNS:
                if portal_total_row:
                    value = ""
                    for h, cid in header_to_id.items():
                        if cid == col_id:
                            value = portal_total_row.get(h, "")
                            break
                    total_row.append(value)
                else:
                    total_row.append("")
            elif col_id in ADDITIVE_COLUMNS:
                col_idx = selected_ids.index(col_id)
                col_sum = 0
                for row in projected_rows:
                    try:
                        val = row[col_idx].replace(",", "").strip()
                        if val:
                            col_sum += int(float(val))
                    except (ValueError, IndexError):
                        pass
                total_row.append(str(col_sum))
            else:
                total_row.append("")

        return total_row

    @staticmethod
    def _section_width(sections: list[SectionDataset]) -> int:
        """Return maximum column count across sections."""
        if not sections:
            return 1
        return max(len(section.headers) for section in sections)

    def _write_excel(
        self,
        target_path: Path,
        sections: list[SectionDataset],
        *,
        report_date: str,
    ) -> None:
        """Write combined Excel with all sections stacked vertically."""
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Comprehensive Reports 10-13"

        col_count = self._section_width(sections)
        main_title = normalize_report_title(
            f"Report 10-13 (Comprehensive Reports) {report_date}",
            report_slug="comprehensive-10-13",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(col_count, 1))
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 3

        for section in sections:
            section_cols = max(len(section.headers), 1)
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=section_cols,
            )
            section_title_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=section.section_config.section_title,
            )
            section_title_cell.font = Font(bold=True, size=11)
            section_title_cell.alignment = Alignment(horizontal="left")
            current_row += 1

            for col_idx, header in enumerate(section.headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
            current_row += 1

            for row_values in section.rows:
                for col_idx, value in enumerate(row_values, start=1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                current_row += 1

            for col_idx, value in enumerate(section.total_row, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.fill = TOTAL_FILL
            current_row += 1

            current_row += 1

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        sections: list[SectionDataset],
        *,
        report_date: str,
    ) -> None:
        """Write all four sections stacked continuously on a single PDF page."""
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        base_pagesize = landscape(A3)
        page_width = base_pagesize[0]

        styles = getSampleStyleSheet()
        section_style = ParagraphStyle(
            "ComprehensiveSection",
            parent=styles["Heading2"],
            fontSize=10,
            leading=12,
            spaceBefore=0,
            spaceAfter=0,
            textColor=colors.black,
        )
        empty_style = ParagraphStyle(
            "ComprehensiveEmpty",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            spaceBefore=0,
            spaceAfter=0,
        )

        main_title = normalize_report_title(
            f"Report 10-13 (Comprehensive Reports) {report_date}",
            report_slug="comprehensive-10-13",
        )
        title_style = ParagraphStyle(
            "ComprehensiveTitle",
            parent=styles["Heading1"],
            fontSize=12,
            leading=14,
            alignment=1,
            spaceBefore=0,
            spaceAfter=0,
        )

        # One title only; no PageBreak / repeated titles between sections.
        story: list = [
            Paragraph(_escape_paragraph_xml(main_title), title_style),
            Spacer(1, _PDF_TITLE_AFTER_PT),
        ]

        for section_idx, section in enumerate(sections):
            if section_idx > 0:
                story.append(Spacer(1, _PDF_SECTION_GAP_PT))

            if section.rows:
                table_data: list[list[object]] = [list(section.headers)]
                for row_values in section.rows:
                    table_data.append(list(row_values))
                table_data.append(list(section.total_row))

                style_commands: list[tuple] = [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.85, 0.85, 0.85)),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]

                table, section_pagesize, _section_margin = build_fitted_table(
                    table_data,
                    style_commands,
                    margin=margin,
                )
                if section_pagesize[0] > page_width:
                    page_width = section_pagesize[0]

                table_width = float(sum(table._colWidths))
                story.append(
                    _heading_table_centered_over_width(
                        section.section_config.section_title,
                        table_width,
                        section_style,
                    )
                )
                story.append(Spacer(1, _PDF_AFTER_HEADING_PT))
                story.append(table)
            else:
                story.append(Paragraph("No data available for this section.", empty_style))

        usable_width = page_width - (2 * margin)
        content_height = 0.0
        for flowable in story:
            _width, height = flowable.wrap(usable_width, 100000)
            content_height += float(height)

        page_height = max(
            base_pagesize[1],
            content_height + (2 * margin) + _PDF_HEIGHT_BUFFER_PT,
        )
        pagesize = (page_width, page_height)

        log_automation_event(
            logger,
            "comprehensive1013_pdf_single_page_layout",
            page_width=page_width,
            page_height=page_height,
            content_height=content_height,
            margin=margin,
            section_count=len(sections),
        )

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        doc.build(story)
        temp_path.replace(target_path)
