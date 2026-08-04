"""Report Vande Bharat post-ingestion processor.

Loads the extracted CSV and emits Report Vande Bharat.xlsx + Report Vande Bharat.pdf
using the shared formatting helpers.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, TableStyle

from app.automation.config import config
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.artifact_titles import build_artifact_main_title
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT, fit_column_widths, preferred_column_widths
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.report18_filters import REPORT18_FILE_STEM, REPORT18_LOG_PREFIX
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report18_processor"
REPORT18_SHEET_TITLE = "Vande Bharat"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 14.0)


def _escape_paragraph_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


class Report18Processor:
    """Process Report Vande Bharat extract into Excel/PDF artifacts."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        _ = source_b_path
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        headers, rows = self._load_table(source_a_path)
        if not headers:
            return ProcessingResult(
                success=False,
                error="REPORT18_TABLE_MISSING: extracted CSV has no headers",
                source_a_path=str(source_a_path),
                input_row_count=0,
            )
        if not rows:
            return ProcessingResult(
                success=False,
                error="REPORT18_TABLE_MISSING: extracted CSV has no data rows",
                source_a_path=str(source_a_path),
                input_row_count=0,
            )

        date_range = date_range_for_processing(column_selection)
        main_title = build_artifact_main_title("report18", date_range)

        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, parent.name)
            )

        # Exact artifact names required by product: Report Vande Bharat.xlsx / .pdf
        excel_path = excel_dir / f"{REPORT18_FILE_STEM}.xlsx"
        pdf_path = pdf_dir / f"{REPORT18_FILE_STEM}.pdf"

        try:
            self._write_excel(excel_path, headers, rows, main_title=main_title)
            logger.info("%s Excel generated", REPORT18_LOG_PREFIX)
            log_automation_event(logger, "report18_excel_generated", excel_path=str(excel_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT18_XLSX_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=len(rows),
            )

        try:
            self._write_pdf(pdf_path, headers, rows, main_title=main_title)
            logger.info("%s PDF generated", REPORT18_LOG_PREFIX)
            log_automation_event(logger, "report18_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT18_PDF_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=len(rows),
                excel_path=str(excel_path),
            )

        log_automation_event(
            logger,
            "report18_processing_completed",
            source_a=str(source_a_path),
            input_row_count=len(rows),
            total_output_rows=len(rows),
        )

        return ProcessingResult(
            success=True,
            attempted=True,
            input_row_count=len(rows),
            processed_row_count=len(rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=len(rows),
            output_columns=list(headers),
            visible_columns=list(headers),
            selected_column_ids=list(headers),
            column_order=list(headers),
            configuration_source="default",
        )

    def _load_table(self, source_a_path: Path) -> tuple[list[str], list[list[str]]]:
        if not source_a_path.is_file():
            return [], []
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = [list(row) for row in reader]
        if not rows:
            return [], []
        headers = [str(c).strip() for c in rows[0]]
        data = [[str(c).strip() for c in row] for row in rows[1:] if any(str(c).strip() for c in row)]
        # Pad/truncate rows to header width
        width = len(headers)
        normalized: list[list[str]] = []
        for row in data:
            if len(row) < width:
                row = row + [""] * (width - len(row))
            elif len(row) > width:
                row = row[:width]
            normalized.append(row)
        return headers, normalized

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT18_SHEET_TITLE

        col_count = max(len(headers), 1)
        title = normalize_report_title(main_title, report_slug="report18")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, row_values in enumerate(rows, start=3):
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)

        for col_idx in range(1, col_count + 1):
            letter = worksheet.cell(row=2, column=col_idx).column_letter
            worksheet.column_dimensions[letter].width = 14

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
    ) -> None:
        ensure_pdf_unicode_fonts()
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Report18Title",
            parent=styles["Heading1"],
            fontSize=14,
            leading=16,
            alignment=TA_CENTER,
            spaceAfter=8,
            fontName=pdf_font_bold(),
        )

        normalized_main_title = normalize_report_title(main_title, report_slug="report18")
        story: list[Any] = [
            Paragraph(_escape_paragraph_xml(normalized_main_title), title_style),
        ]

        page_width, _page_height = landscape(A4)
        usable_width = page_width - (2 * margin)

        table_data: list[list[object]] = [list(headers)]
        table_data.extend([list(r) for r in rows])

        preferred = preferred_column_widths(
            table_data,
            font_size=8,
            headers=list(headers),
        )
        col_widths = fit_column_widths(preferred, usable_width)
        table = LongTable(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
                    ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
                    ("FONTNAME", (0, 1), (-1, -1), pdf_font_regular()),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(table)

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
            title=normalized_main_title,
        )
        doc.build(story)
        temp_path.replace(target_path)
