"""Report 2 post-ingestion processor (Division Wise Top 25)."""

from __future__ import annotations

import csv
import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from app.automation.formatting.pdf_fonts import pdf_font_bold, pdf_title_style
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.automation.config import config
from app.automation.formatting.excel_print import apply_column_formatting, apply_report_print_setup
from app.automation.formatting.pdf_table import build_fitted_table
from app.automation.formatting.text_pipeline import (
    normalize_report_title,
    prepare_output_for_rendering,
    verify_text_rendering,
)
from app.automation.formatting.pdf_verify import verify_report_output
from app.automation.formatting.scr import SCR_FILL, SCR_FONT, row_contains_scr
from app.automation.formatting.serial import apply_serial_number, is_serial_header, renumber_data_rows
from app.automation.processing.base import ProcessingResult
from app.automation.processing.column_config import project_for_output, resolve_projection_column_keys
from app.automation.processing.output_columns import (
    REPORT2_HIDDEN_COLUMNS,
    SOURCE_B_DATA_COLUMNS,
    build_report2_display_total_row,
)

HIDDEN_COLUMNS = REPORT2_HIDDEN_COLUMNS
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    previous_day_report_date,
    resolve_report_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report2_division_wise_processor"

TOP_N = 25

FEEDBACK_FILENAME = "report2_division_feedback_raw.csv"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class Report2Processor:
    """Process Division Wise Top 25 dataset and emit formatted Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict | None = None,
    ) -> ProcessingResult:
        log_automation_event(
            logger,
            "report2_processor_start",
            source_a_path=str(source_a_path),
            source_a_exists=source_a_path.exists(),
            source_b_path=str(source_b_path) if source_b_path else None,
            source_b_provided=source_b_path is not None,
            source_b_exists=source_b_path.exists() if source_b_path else False,
        )

        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        # Report 2 REQUIRES explicit source_b_path - no filesystem fallback
        # This ensures we never use stale feedback data from a previous run
        if source_b_path is None:
            log_automation_event(
                logger,
                "report2_processor_source_b_required",
                error="source_b_path is required for Report 2 - no fallback allowed",
            )
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                error="Report 2 requires explicit source_b_path (Feedback Division Wise CSV). "
                      "Fallback search disabled to prevent stale data usage.",
            )

        if not source_b_path.exists():
            log_automation_event(
                logger,
                "report2_processor_source_b_missing",
                source_b_path=str(source_b_path),
                error="Source B file does not exist",
            )
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                source_b_path=str(source_b_path),
                error=f"Source B file missing: {source_b_path}",
            )

        feedback_path = source_b_path

        # Capture source modification times for verification
        source_a_mtime = source_a_path.stat().st_mtime
        source_b_mtime = feedback_path.stat().st_mtime

        source_a_rows, source_a_headers = self._read_csv(source_a_path)
        data_a, _total_a = self._split_total_row(source_a_rows)

        source_b_rows, source_b_headers = self._read_csv(feedback_path)

        # Verify Source B has the required feedback columns
        missing_feedback_cols = [col for col in SOURCE_B_DATA_COLUMNS if col not in source_b_headers]
        if missing_feedback_cols:
            log_automation_event(
                logger,
                "report2_source_b_missing_columns",
                missing=missing_feedback_cols,
                available=source_b_headers,
                error="Source B is missing required feedback columns",
            )
            data_b, _ = self._split_total_row(source_b_rows)
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                source_b_path=str(feedback_path),
                source_a_rows=len(data_a),
                source_b_rows=len(data_b),
                source_a_mtime=source_a_mtime,
                source_b_mtime=source_b_mtime,
                error=f"Source B missing required columns: {missing_feedback_cols}",
            )

        log_automation_event(
            logger,
            "report2_source_b_columns_verified",
            required=SOURCE_B_DATA_COLUMNS,
            available=source_b_headers,
        )

        merge_result = self.build_merged_table(
            source_a_rows,
            source_a_headers,
            source_b_rows,
            source_b_headers,
        )
        if merge_result is None:
            data_b, _ = self._split_total_row(source_b_rows)
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                source_b_path=str(feedback_path),
                source_a_rows=len(data_a),
                source_b_rows=len(data_b),
                source_a_mtime=source_a_mtime,
                source_b_mtime=source_b_mtime,
                error="REPORT2_MERGE_FAILED: Could not merge division and feedback datasets.",
            )

        merged_headers, merged_rows, scr_row_flags = merge_result
        data_b, _ = self._split_total_row(source_b_rows)
        matched_count = max(len(merged_rows) - 1, 0)

        log_automation_event(
            logger,
            "report2_merge_completed",
            source_a_rows=len(data_a),
            source_b_rows=len(data_b),
            merged_row_count=len(merged_rows),
            matched_count=matched_count,
        )

        selected_keys, config_source = resolve_projection_column_keys(
            report_slug,
            column_selection=column_selection,
        )
        output_headers, output_rows, visible_columns, selected_column_ids, config_source = (
            project_for_output(
                report_slug,
                full_headers=merged_headers,
                rows=merged_rows,
                selected_keys=selected_keys,
                config_source=config_source,
            )
        )
        log_automation_event(
            logger,
            "report2_columns_projected",
            selected_column_count=len(selected_column_ids),
            selected_column_ids=selected_column_ids,
            configuration_source=config_source,
            final_headers=output_headers,
            final_row_count=len(output_rows),
        )

        if any(is_serial_header(header) for header in output_headers):
            output_rows = renumber_data_rows(output_headers, output_rows)
            log_automation_event(
                logger,
                "report2_serial_numbers_regenerated",
                final_row_count=len(output_rows),
            )

        scr_row_indices = {
            idx
            for idx, is_scr in enumerate(scr_row_flags)
            if is_scr and idx < len(output_rows)
        }
        if scr_row_indices:
            log_automation_event(
                logger,
                "report2_scr_highlight_applied",
                scr_row_indices=sorted(scr_row_indices),
            )

        source_b_path_str = str(feedback_path)
        source_b_row_count = len(data_b)

        # Verify merged headers include all feedback columns
        feedback_cols_in_merged = [col for col in SOURCE_B_DATA_COLUMNS if col in merged_headers]
        log_automation_event(
            logger,
            "report2_merged_headers",
            total_columns=len(merged_headers),
            feedback_columns_present=len(feedback_cols_in_merged),
            feedback_columns=feedback_cols_in_merged,
            matched_count=matched_count,
        )

        output_headers, output_rows = prepare_output_for_rendering(
            report_slug,
            output_headers,
            output_rows,
        )

        report_date = previous_day_report_date()
        run_timestamp = datetime.now().strftime("%H%M%S")
        excel_dir = ensure_directory(resolve_report_dir(config.output_excel_dir, report_slug))
        pdf_dir = ensure_directory(resolve_report_dir(config.output_pdf_dir, report_slug))
        base_name = f"Rail_Madad_Report_2_Division_Wise_Bottom_25_{report_date}_{run_timestamp}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        log_automation_event(
            logger,
            "report2_output_paths",
            report_date=report_date,
            run_timestamp=run_timestamp,
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
        )

        try:
            self._write_excel(
                excel_path,
                output_headers,
                output_rows,
                report_date=report_date,
                scr_row_indices=scr_row_indices,
            )
            self._write_pdf(
                pdf_path,
                output_headers,
                output_rows,
                report_date=report_date,
                scr_row_indices=scr_row_indices,
            )
            verify_report_output(
                report_slug=report_slug,
                headers=output_headers,
                rows=output_rows,
                pdf_path=pdf_path,
                excel_path=excel_path,
            )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(data_a),
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_b_path=source_b_path_str,
                source_a_rows=len(data_a),
                source_b_rows=source_b_row_count,
                source_a_mtime=source_a_mtime,
                source_b_mtime=source_b_mtime,
                run_timestamp=run_timestamp,
            )

        # Capture output modification time for verification
        output_mtime = pdf_path.stat().st_mtime if pdf_path.exists() else None

        log_automation_event(
            logger,
            "phase8_dataset_loaded",
            source_a=str(source_a_path),
            source_b=source_b_path_str,
            input_row_count=len(data_a),
            top_n_selected=max(len(merged_rows) - 1, 0),
            source_a_mtime=source_a_mtime,
            source_b_mtime=source_b_mtime,
            output_mtime=output_mtime,
            run_timestamp=run_timestamp,
        )

        return ProcessingResult(
            success=True,
            input_row_count=len(data_a),
            processed_row_count=len(merged_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_b_path=source_b_path_str,
            source_a_rows=len(data_a),
            source_b_rows=source_b_row_count,
            source_a_mtime=source_a_mtime,
            source_b_mtime=source_b_mtime,
            output_mtime=output_mtime,
            run_timestamp=run_timestamp,
            output_columns=output_headers,
            visible_columns=visible_columns,
            selected_column_ids=selected_column_ids,
            column_order=list(selected_column_ids),
            configuration_source=config_source,
        )

    def build_merged_table(
        self,
        source_a_rows: list[dict[str, str]],
        source_a_headers: list[str],
        source_b_rows: list[dict[str, str]],
        source_b_headers: list[str],
    ) -> tuple[list[str], list[list[str]], list[bool]] | None:
        """Build full merged table including totals; None when merge validation fails."""
        data_a, _total_a = self._split_total_row(source_a_rows)
        top_n_rows = data_a[:TOP_N]
        data_b, _total_b = self._split_total_row(source_b_rows)

        merged_headers = source_a_headers + ["S.No.", "Organisation"] + SOURCE_B_DATA_COLUMNS
        merged_rows, matched_count, _, _, _matched_pairs, scr_flags = self._merge_rows(
            top_n_rows,
            source_a_headers,
            data_b,
            source_b_headers,
        )
        if matched_count == 0:
            return None

        feedback_col_start = len(source_a_headers) + 2
        all_feedback_blank = all(
            all(cell == "" for cell in row[feedback_col_start:])
            for row in merged_rows
        )
        if all_feedback_blank:
            return None

        merged_rows.append(
            build_report2_display_total_row(
                merged_headers=merged_headers,
                data_rows=merged_rows,
                source_a_headers=source_a_headers,
                source_b_headers=source_b_headers,
                source_b_columns=SOURCE_B_DATA_COLUMNS,
            )
        )
        scr_flags.append(False)
        return merged_headers, merged_rows, scr_flags

    def _find_feedback_csv(self, report_slug: str) -> Path | None:
        """DEPRECATED: Filesystem fallback search for feedback CSV.

        This method is no longer used by Report 2 processing. The processor now
        requires explicit source_b_path to prevent stale data usage. This method
        is retained only for backward compatibility with tests.
        """
        import warnings
        warnings.warn(
            "_find_feedback_csv is deprecated - Report 2 requires explicit source_b_path",
            DeprecationWarning,
            stacklevel=2,
        )
        log_automation_event(
            logger,
            "report2_deprecated_fallback_called",
            report_slug=report_slug,
            warning="This fallback should not be used in production",
        )
        extracted_dir = resolve_report_dir(config.extracted_data_dir, report_slug)
        preferred = extracted_dir / FEEDBACK_FILENAME
        if preferred.exists():
            return preferred
        matches = sorted(extracted_dir.glob("*feedback*.csv"))
        return matches[0] if matches else None

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _normalize_org(name: str) -> str:
        """Simple normalization: lowercase and collapse whitespace."""
        return re.sub(r"\s+", " ", name.strip()).lower()

    @staticmethod
    def _extract_base_division(name: str) -> str:
        """Extract base division name before parenthetical suffix.

        Handles the mismatch between Source A railway zone suffixes and
        Source B station code suffixes:
        - 'DELHI DIVISION (Northern Railway)' -> 'delhi division'
        - 'DELHI DIVISION (DLI)' -> 'delhi division'
        - 'IRC (IRCTC)' -> 'irc'
        - 'LUCKNOW DIVN (NER)' -> 'lucknow division'

        Normalization includes:
        - Remove trailing parenthetical suffix
        - Collapse whitespace
        - Lowercase
        - Normalize DIVN -> DIVISION
        - Normalize & -> AND
        """
        # Remove trailing parenthetical suffix (railway zone or station code)
        base = re.sub(r"\s*\([^)]*\)\s*$", "", name)
        # Drop trailing railway labels when present outside parentheses
        base = re.sub(r"\s+railway\s*$", "", base, flags=re.I)
        # Normalize & to AND before stripping punctuation
        base = re.sub(r"\s*&\s*", " and ", base)
        # Normalize remaining punctuation/hyphens to spaces
        base = re.sub(r"[^\w\s]", " ", base)
        # Collapse whitespace and lowercase
        base = re.sub(r"\s+", " ", base.strip()).lower()
        # Normalize DIVN to DIVISION
        base = re.sub(r"\bdivn\b", "division", base)
        return base

    @staticmethod
    def _is_total_row(row: dict[str, str]) -> bool:
        org = row.get("Organisation", "") or row.get("Division", "")
        return "total" in org.strip().lower()

    def _split_total_row(
        self,
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        if self._is_total_row(rows[-1]):
            return rows[:-1], rows[-1]
        return rows, None

    @staticmethod
    def _load_portal_total_sidecar(
        source_a_path: Path,
        source_a_headers: list[str],
    ) -> dict[str, str] | None:
        """Load portal total row saved before top-25 trim (Report 2 only)."""
        import json

        sidecar = source_a_path.parent / "report2_division_comprehensive_portal_total.json"
        if not sidecar.is_file():
            return None
        try:
            payload = json.loads(sidecar.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None
        if not isinstance(payload, dict):
            return None
        return {header: str(payload.get(header, "")) for header in source_a_headers}

    def _build_feedback_lookup(
        self,
        source_b_rows: list[dict[str, str]],
    ) -> tuple[dict[str, dict[str, str]], list[str], dict[str, list[str]]]:
        """Build feedback lookup by base division name.

        Ambiguous bases (multiple Source B orgs → same base) are omitted from
        the lookup so merges leave blank feedback rather than guessing.
        """
        lookup: dict[str, dict[str, str]] = {}
        duplicates: list[str] = []
        base_to_orgs: dict[str, list[str]] = {}
        ambiguous_bases: set[str] = set()

        for row in source_b_rows:
            org = row.get("Organisation", "") or row.get("Division", "")
            base = self._extract_base_division(org)

            if base not in base_to_orgs:
                base_to_orgs[base] = []
            base_to_orgs[base].append(org)

            if base in lookup or base in ambiguous_bases:
                ambiguous_bases.add(base)
                duplicates.append(org)
                lookup.pop(base, None)
                continue
            lookup[base] = row

        if duplicates:
            log_automation_event(
                logger,
                "report2_feedback_duplicate_orgs",
                duplicates=duplicates,
                ambiguous_bases=sorted(ambiguous_bases),
                count=len(duplicates),
                warning="Ambiguous Source B bases excluded from merge (blank feedback)",
            )

        log_automation_event(
            logger,
            "report2_feedback_lookup_built",
            total_source_b_rows=len(source_b_rows),
            unique_base_names=len(lookup),
            ambiguous_base_count=len(ambiguous_bases),
            sample_base_names=list(lookup.keys())[:5],
        )

        return lookup, duplicates, base_to_orgs

    def _feedback_values_for_row(
        self,
        org: str,
        lookup: dict[str, dict[str, str]],
    ) -> tuple[list[str], dict[str, str] | None]:
        """Get feedback values for a division using base-name lookup.

        Returns (feedback_values, matched_row).
        """
        base = self._extract_base_division(org)
        matched_row = lookup.get(base)
        if matched_row:
            return [matched_row.get(column, "") for column in SOURCE_B_DATA_COLUMNS], matched_row
        return [""] * len(SOURCE_B_DATA_COLUMNS), None

    @staticmethod
    def _is_scr_division(division: str) -> bool:
        from app.automation.formatting.scr import SCR_PATTERN

        text = re.sub(r"\s+", " ", division.strip())
        return bool(SCR_PATTERN.search(text))

    def _merge_rows(
        self,
        source_a_rows: list[dict[str, str]],
        source_a_headers: list[str],
        source_b_rows: list[dict[str, str]],
        source_b_headers: list[str],
    ) -> tuple[list[list[str]], int, list[str], list[str], list[dict], list[bool]]:
        """Merge Source A rows with Source B feedback using base-name matching.

        Returns:
            tuple of (merged_rows, matched_count, unmatched_source_a, unmatched_source_b,
            matched_pairs, scr_row_flags)
        """
        lookup, duplicates, base_to_orgs = self._build_feedback_lookup(source_b_rows)
        merged: list[list[str]] = []
        scr_flags: list[bool] = []
        unmatched_source_a: list[str] = []
        matched_count = 0
        matched_pairs: list[dict] = []

        # Track which Source B divisions were matched
        matched_source_b_bases: set[str] = set()

        for index, row in enumerate(source_a_rows, start=1):
            org_a = row.get("Division", "") or row.get("Organisation", "")
            base_a = self._extract_base_division(org_a)

            source_a_values = apply_serial_number(
                source_a_headers,
                [row.get(header, "") for header in source_a_headers],
                index,
            )
            b_sno = str(index)
            b_org = org_a  # Use Source A name for display

            b_values, matched_row = self._feedback_values_for_row(org_a, lookup)

            if matched_row:
                matched_count += 1
                matched_source_b_bases.add(base_a)
                org_b = matched_row.get("Organisation", "") or matched_row.get("Division", "")
                matched_pairs.append({
                    "source_a": org_a,
                    "source_b": org_b,
                    "base_name": base_a,
                    "feedback_received": matched_row.get("Feedback Received", ""),
                })
            else:
                unmatched_source_a.append(org_a)

            merged.append(source_a_values + [b_sno, b_org] + b_values)
            scr_flags.append(self._is_scr_division(org_a))

        # Find Source B divisions not matched to any Source A
        source_a_bases = {self._extract_base_division(r.get("Division", "") or r.get("Organisation", ""))
                         for r in source_a_rows}
        unmatched_source_b = [
            row.get("Organisation", "") or row.get("Division", "")
            for row in source_b_rows
            if self._extract_base_division(row.get("Organisation", "") or row.get("Division", "")) not in source_a_bases
        ]

        # Comprehensive logging
        log_automation_event(
            logger,
            "report2_merge_summary",
            source_a_count=len(source_a_rows),
            source_b_count=len(source_b_rows),
            matched_count=matched_count,
            unmatched_source_a_count=len(unmatched_source_a),
            unmatched_source_b_count=len(unmatched_source_b),
            unmatched_source_a=unmatched_source_a,
            unmatched_source_b=unmatched_source_b[:10],  # Limit for log size
            duplicate_feedback_orgs=duplicates,
            sample_matched=matched_pairs[:3],
        )

        if unmatched_source_a:
            log_automation_event(
                logger,
                "report2_unmatched_source_a",
                divisions=unmatched_source_a,
                warning="These Source A divisions had no matching feedback data",
            )

        if unmatched_source_b:
            log_automation_event(
                logger,
                "report2_unmatched_source_b",
                divisions=unmatched_source_b[:10],
                total_unmatched=len(unmatched_source_b),
                info="These Source B divisions were not in Source A Top 25",
            )

        return merged, matched_count, unmatched_source_a, unmatched_source_b, matched_pairs, scr_flags

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
        scr_row_indices: set[int] | None = None,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report 2"

        title = normalize_report_title(
            f"Rail Madad Report No 2 - Division Wise Complaints & Feedback Report "
            f"- Bottom 25 Divisions on date {report_date}",
            report_slug="division",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        header_row = 2
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        data_start = 3
        for row_offset, row_values in enumerate(rows):
            row_idx = data_start + row_offset
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].hidden = False

        if scr_row_indices:
            for row_offset in scr_row_indices:
                if row_offset >= len(rows):
                    continue
                row_idx = data_start + row_offset
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = SCR_FILL
                    cell.font = SCR_FONT
        else:
            from app.automation.formatting.scr import highlight_south_central_railway_rows

            highlight_south_central_railway_rows(
                worksheet,
                start_row=data_start,
                end_row=worksheet.max_row,
                start_col=1,
                end_col=len(headers),
            )

        if rows:
            totals_row = worksheet.max_row
            totals_fill = PatternFill(fill_type="solid", fgColor="E8E8E8")
            for col_idx in range(1, len(headers) + 1):
                cell = worksheet.cell(row=totals_row, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = totals_fill

        apply_column_formatting(
            worksheet,
            headers,
            header_row=header_row,
            data_start_row=data_start,
        )
        apply_report_print_setup(worksheet, col_count=len(headers))

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
        scr_row_indices: set[int] | None = None,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        visible_indices = list(range(1, len(headers) + 1))
        visible_headers = headers

        table_data: list[list[str]] = [visible_headers]
        pdf_scr_indices: set[int] = set()
        for row_offset, row_values in enumerate(rows):
            visible_row = [row_values[idx - 1] if idx <= len(row_values) else "" for idx in visible_indices]
            table_data.append(visible_row)
            if scr_row_indices is not None:
                if row_offset in scr_row_indices:
                    pdf_scr_indices.add(len(table_data) - 1)
            elif row_contains_scr(visible_row):
                pdf_scr_indices.add(len(table_data) - 1)

        style_commands: list[tuple] = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
        ]
        for row_idx in pdf_scr_indices:
            style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.yellow))
            style_commands.append(("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.black))
        if rows:
            style_commands.append(
                ("FONTNAME", (0, len(table_data) - 1), (-1, len(table_data) - 1), pdf_font_bold())
            )
            style_commands.append(
                ("BACKGROUND", (0, len(table_data) - 1), (-1, len(table_data) - 1), colors.HexColor("#E8E8E8"))
            )

        table, pagesize, margin = build_fitted_table(table_data, style_commands)
        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        story = [
            Paragraph(
                normalize_report_title(
                    f"Rail Madad Report No 2 - Division Wise Complaints & Feedback Report "
                    f"- Bottom 25 Divisions on date {report_date}",
                    report_slug="division",
                ),
                pdf_title_style("Report2Title"),
            ),
            Spacer(1, 12),
            table,
        ]
        doc.build(story)
        temp_path.replace(target_path)
