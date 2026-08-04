"""Report 14 post-ingestion processor — Watering Complaints Previous + Upcoming.

Loads previous_watering.csv and upcoming_watering.csv via report14_combined_index.csv
and emits one combined XLSX + one PDF with side-by-side metric columns.
"""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from app.automation.config import config
from app.automation.date_range import date_range_for_processing
from app.automation.formatting.artifact_titles import build_artifact_main_title
from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_bold, pdf_font_regular
from app.automation.formatting.pdf_table import SAFE_MARGIN_PT, fit_column_widths
from app.automation.formatting.text_pipeline import normalize_report_title
from app.automation.processing.base import ProcessingResult
from app.automation.processing.report14_output_columns import (
    REPORT14_LABEL_BY_ID,
    report14_default_ids,
    report14_labels,
    validate_selected_report14_fields,
)
from app.automation.report14_filters import (
    METRIC_COLUMNS,
    OUTPUT_HEADERS,
    SOURCE_PREVIOUS,
    SOURCE_UPCOMING,
)
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report14_processor"
REPORT14_SHEET_TITLE = "Watering Complaints"
REPORT14_FILE_STEM = "Rail_Madad_Report_14_Watering_Complaints"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)

_PDF_MARGIN_PT = min(SAFE_MARGIN_PT, 12.0)


def _escape_paragraph_xml(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _parse_num(value: str) -> float:
    cleaned = re.sub(r"[^\d.\-]", "", str(value or "").replace(",", ""))
    if not cleaned or cleaned in {".", "-", "-."}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _format_num(value: float) -> str:
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}"


class Report14Processor:
    """Merge Previous + Upcoming watering extracts into combined Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict[str, Any] | None = None,
    ) -> ProcessingResult:
        _ = source_b_path
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        prev_rows, prev_headers, up_rows, up_headers, total_input = self._load_sources(
            source_a_path
        )
        if prev_rows is None or up_rows is None:
            return ProcessingResult(
                success=False,
                error="REPORT14_TABLE_MISSING: need both previous and upcoming sources",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
            )

        selected_ids = report14_default_ids()
        if column_selection:
            raw = (
                column_selection.get("selected_column_ids")
                or column_selection.get("column_order")
                or []
            )
            if raw:
                selected_ids = validate_selected_report14_fields(raw)

        merged_headers, merged_rows = self._merge_side_by_side(
            prev_rows, prev_headers or [], up_rows, up_headers or []
        )
        visible_headers = report14_labels(selected_ids)
        id_to_label = REPORT14_LABEL_BY_ID
        col_indexes = [
            merged_headers.index(id_to_label[cid])
            for cid in selected_ids
            if cid in id_to_label and id_to_label[cid] in merged_headers
        ]
        if not col_indexes:
            col_indexes = list(range(len(merged_headers)))
            visible_headers = list(merged_headers)
            selected_ids = report14_default_ids()

        projected_rows = [
            [row[i] if i < len(row) else "" for i in col_indexes] for row in merged_rows
        ]

        date_range = date_range_for_processing(column_selection)
        main_title = build_artifact_main_title("report14", date_range)
        title_suffix = date_range.title_suffix()
        filename_suffix = date_range.filename_suffix()
        subtitle = (
            f"Previous & Upcoming Watering Point — SCR Division Wise {title_suffix}"
        )

        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            parent = source_a_path.parent
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, parent.name)
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, parent.name)
            )

        base_name = f"{REPORT14_FILE_STEM}_{filename_suffix}"
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(
                excel_path,
                visible_headers,
                projected_rows,
                main_title=main_title,
                subtitle=subtitle,
            )
            log_automation_event(logger, "report14_excel_generated", excel_path=str(excel_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT14_XLSX_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
            )

        try:
            self._write_pdf(
                pdf_path,
                visible_headers,
                projected_rows,
                main_title=main_title,
                subtitle=subtitle,
            )
            log_automation_event(logger, "report14_pdf_generated", pdf_path=str(pdf_path))
        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"REPORT14_PDF_FAILED: {exc}",
                source_a_path=str(source_a_path),
                input_row_count=total_input,
                excel_path=str(excel_path),
            )

        data_rows = [r for r in projected_rows if not self._is_total_row(r)]
        log_automation_event(
            logger,
            "report14_processing_completed",
            source_a=str(source_a_path),
            input_row_count=total_input,
            total_output_rows=len(data_rows),
        )

        return ProcessingResult(
            success=True,
            attempted=True,
            input_row_count=total_input,
            processed_row_count=len(data_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=total_input,
            output_columns=visible_headers,
            visible_columns=visible_headers,
            selected_column_ids=list(selected_ids),
            column_order=list(selected_ids),
            configuration_source="manual_snapshot"
            if column_selection and column_selection.get("selected_column_ids")
            else "default",
        )

    def _load_sources(
        self, source_a_path: Path
    ) -> tuple[
        list[dict[str, str]] | None,
        list[str] | None,
        list[dict[str, str]] | None,
        list[str] | None,
        int,
    ]:
        index_entries = self._read_combined_index(source_a_path)
        base_dir = source_a_path.parent

        def resolve_csv(source_id: str, filename: str) -> Path | None:
            entry = index_entries.get(source_id)
            if entry and str(entry.get("status", "")).lower() == "success":
                candidate = Path(str(entry.get("csv_path") or ""))
                if candidate.is_file():
                    return candidate
            fallback = base_dir / filename
            return fallback if fallback.is_file() else None

        prev_path = resolve_csv(SOURCE_PREVIOUS.source_id, SOURCE_PREVIOUS.filename)
        up_path = resolve_csv(SOURCE_UPCOMING.source_id, SOURCE_UPCOMING.filename)
        if prev_path is None or up_path is None:
            return None, None, None, None, 0

        prev_rows, prev_headers = self._read_csv(prev_path)
        up_rows, up_headers = self._read_csv(up_path)
        prev_data, _ = self._split_total_row(prev_rows)
        up_data, _ = self._split_total_row(up_rows)
        total_input = len(prev_data) + len(up_data)
        return prev_data, prev_headers, up_data, up_headers, total_input

    def _read_combined_index(self, source_a_path: Path) -> dict[str, dict[str, str]]:
        if source_a_path.name != "report14_combined_index.csv":
            return {}
        if not source_a_path.is_file():
            return {}
        entries: dict[str, dict[str, str]] = {}
        with source_a_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                source_id = (row.get("source_id") or "").strip()
                if not source_id:
                    continue
                entries[source_id] = {
                    "source_id": source_id,
                    "csv_path": (row.get("csv_path") or "").strip(),
                    "status": (row.get("status") or "").strip(),
                }
        return entries

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _split_total_row(
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        last = rows[-1]
        if any(str(v).strip().lower() == "total" for v in last.values()):
            return rows[:-1], last
        return rows, None

    @staticmethod
    def _normalize_key(name: str) -> str:
        return re.sub(r"\s+", " ", name.strip()).lower()

    def _station_key_for_row(self, row: dict[str, str], headers: list[str]) -> str:
        for candidate in (
            "Station",
            "Division",
            "Organisation",
            "Organization",
            "Watering Point",
            "Name",
        ):
            for h in headers:
                if h.strip().lower() == candidate.lower():
                    val = str(row.get(h, "")).strip()
                    if val and val.lower() != "total":
                        return self._normalize_key(val)
        # First non-metric, non-serial column with content.
        for h in headers:
            hl = h.strip().lower()
            if hl in {"s.no.", "s.no", "sno", "sr.no.", "sr no"}:
                continue
            if any(hl == m.lower() for m in METRIC_COLUMNS):
                continue
            val = str(row.get(h, "")).strip()
            if val and val.lower() != "total":
                return self._normalize_key(val)
        return ""

    def _metric_map(self, row: dict[str, str], headers: list[str]) -> dict[str, str]:
        header_map = {h.strip().lower(): h for h in headers}
        out: dict[str, str] = {}
        for metric in METRIC_COLUMNS:
            key = header_map.get(metric.lower())
            out[metric] = str(row.get(key, "")).strip() if key else ""
        return out

    def _display_station(self, row: dict[str, str], headers: list[str]) -> str:
        for candidate in (
            "Station",
            "Division",
            "Organisation",
            "Organization",
            "Watering Point",
            "Name",
        ):
            for h in headers:
                if h.strip().lower() == candidate.lower():
                    val = str(row.get(h, "")).strip()
                    if val:
                        return val
        for h in headers:
            hl = h.strip().lower()
            if hl in {"s.no.", "s.no", "sno"} or any(hl == m.lower() for m in METRIC_COLUMNS):
                continue
            val = str(row.get(h, "")).strip()
            if val:
                return val
        return ""

    def _merge_side_by_side(
        self,
        prev_rows: list[dict[str, str]],
        prev_headers: list[str],
        up_rows: list[dict[str, str]],
        up_headers: list[str],
    ) -> tuple[list[str], list[list[str]]]:
        prev_by_key: dict[str, dict[str, str]] = {}
        prev_display: dict[str, str] = {}
        for row in prev_rows:
            key = self._station_key_for_row(row, prev_headers)
            if not key:
                continue
            prev_by_key[key] = self._metric_map(row, prev_headers)
            prev_display[key] = self._display_station(row, prev_headers)

        up_by_key: dict[str, dict[str, str]] = {}
        up_display: dict[str, str] = {}
        for row in up_rows:
            key = self._station_key_for_row(row, up_headers)
            if not key:
                continue
            up_by_key[key] = self._metric_map(row, up_headers)
            up_display[key] = self._display_station(row, up_headers)

        all_keys = list(dict.fromkeys([*prev_by_key.keys(), *up_by_key.keys()]))

        def sort_key(k: str) -> tuple[float, str]:
            recv = _parse_num(prev_by_key.get(k, {}).get("Received", "0"))
            recv_up = _parse_num(up_by_key.get(k, {}).get("Received", "0"))
            return (-max(recv, recv_up), k)

        all_keys.sort(key=sort_key)

        blank_metrics = {m: "" for m in METRIC_COLUMNS}
        data_rows: list[list[str]] = []
        for idx, key in enumerate(all_keys, start=1):
            station = prev_display.get(key) or up_display.get(key) or key
            prev = prev_by_key.get(key, blank_metrics)
            up = up_by_key.get(key, blank_metrics)
            data_rows.append(
                [
                    str(idx),
                    station,
                    prev.get("Opening Balance", ""),
                    prev.get("Received", ""),
                    prev.get("% Share", ""),
                    prev.get("Closed", ""),
                    prev.get("Closing Balance", ""),
                    prev.get("% Disposal", ""),
                    up.get("Opening Balance", ""),
                    up.get("Received", ""),
                    up.get("% Share", ""),
                    up.get("Closed", ""),
                    up.get("Closing Balance", ""),
                    up.get("% Disposal", ""),
                ]
            )

        # Total row — sum numeric metric columns.
        total = [""] * len(OUTPUT_HEADERS)
        total[0] = ""
        total[1] = "Total"
        for col_idx in (2, 3, 5, 6, 8, 9, 11, 12):
            s = sum(_parse_num(r[col_idx]) for r in data_rows)
            total[col_idx] = _format_num(s)
        # % columns recompute from totals when possible.
        for share_idx, recv_idx in ((4, 3), (10, 9)):
            total_recv = _parse_num(total[recv_idx])
            if total_recv > 0:
                total[share_idx] = "100.00"
            else:
                total[share_idx] = ""
        for disp_idx, closed_idx, recv_idx, open_idx in (
            (7, 5, 3, 2),
            (13, 11, 9, 8),
        ):
            closed = _parse_num(total[closed_idx])
            received = _parse_num(total[recv_idx])
            opening = _parse_num(total[open_idx])
            denom = opening + received
            if denom > 0:
                total[disp_idx] = f"{(closed / denom) * 100.0:.2f}"
            else:
                total[disp_idx] = ""

        data_rows.append(total)
        return list(OUTPUT_HEADERS), data_rows

    @staticmethod
    def _is_total_row(row: list[str]) -> bool:
        return any(str(c).strip().lower() == "total" for c in row)

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
        subtitle: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT14_SHEET_TITLE

        col_count = max(len(headers), 1)
        main_title = normalize_report_title(main_title, report_slug="report14")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = worksheet.cell(row=1, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center")

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        sub_cell = worksheet.cell(
            row=2,
            column=1,
            value=normalize_report_title(subtitle, report_slug="report14"),
        )
        sub_cell.font = Font(bold=True, size=10)
        sub_cell.alignment = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, row_values in enumerate(rows, start=4):
            is_total = self._is_total_row(row_values)
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 2 else "center",
                    wrap_text=True,
                )
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = TOTAL_FILL

        worksheet.column_dimensions["A"].width = 6
        worksheet.column_dimensions["B"].width = 22
        for col_idx in range(3, col_count + 1):
            letter = worksheet.cell(row=3, column=col_idx).column_letter
            worksheet.column_dimensions[letter].width = 11

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        main_title: str,
        subtitle: str,
    ) -> None:
        ensure_pdf_unicode_fonts()
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        margin = _PDF_MARGIN_PT
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Report14Title",
            parent=styles["Heading1"],
            fontSize=12,
            leading=14,
            alignment=TA_CENTER,
            spaceAfter=3,
            fontName=pdf_font_bold(),
        )
        subtitle_style = ParagraphStyle(
            "Report14Subtitle",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            spaceAfter=6,
            fontName=pdf_font_regular(),
        )
        cell_style = ParagraphStyle(
            "Report14Cell",
            parent=styles["Normal"],
            fontSize=7,
            leading=8,
            fontName=pdf_font_regular(),
        )
        header_style = ParagraphStyle(
            "Report14Header",
            parent=styles["Normal"],
            fontSize=7,
            leading=8,
            fontName=pdf_font_bold(),
            textColor=colors.white,
            alignment=TA_CENTER,
        )

        page_width, _ = landscape(A4)
        usable = page_width - 2 * margin
        n_cols = max(len(headers), 1)
        raw_widths = [22.0] + [48.0] + [36.0] * max(n_cols - 2, 0)
        col_widths = fit_column_widths(raw_widths[:n_cols], usable)

        table_data: list[list[Any]] = [
            [Paragraph(_escape_paragraph_xml(h), header_style) for h in headers]
        ]
        for row in rows:
            cells: list[Any] = []
            for i, val in enumerate(row):
                cells.append(Paragraph(_escape_paragraph_xml(val), cell_style))
            # Pad short rows
            while len(cells) < n_cols:
                cells.append(Paragraph("", cell_style))
            table_data.append(cells[:n_cols])

        table = LongTable(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds: list[tuple] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), pdf_font_bold()),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#666666")),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        for r_idx, row in enumerate(rows, start=1):
            if self._is_total_row(row):
                style_cmds.append(
                    ("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#D9D9D9"))
                )
                style_cmds.append(("FONTNAME", (0, r_idx), (-1, r_idx), pdf_font_bold()))
        table.setStyle(TableStyle(style_cmds))

        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        story = [
            Paragraph(_escape_paragraph_xml(main_title), title_style),
            Paragraph(_escape_paragraph_xml(subtitle), subtitle_style),
            Spacer(1, 4),
            table,
        ]
        doc.build(story)
        temp_path.replace(target_path)
