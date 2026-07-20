"""Report 6 SCR post-ingestion processor (Station Mode Unsatisfactory Feedback)."""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from app.automation.formatting.pdf_fonts import pdf_title_style
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.automation.config import config
from app.automation.formatting.excel_print import apply_column_formatting, apply_report_print_setup
from app.automation.formatting.pdf_table import build_wrapped_fitted_table
from app.automation.formatting.text_pipeline import (
    normalize_report_title,
    prepare_output_for_rendering,
    verify_text_rendering,
)
from app.automation.formatting.pdf_verify import verify_report_output
from app.automation.formatting.scr import (
    NO_FILL,
    clear_complaint_data_row_formatting,
    mode_matches,
)
from app.automation.processing.base import ProcessingResult
from app.automation.processing.column_config import project_scr_for_output
from app.automation.processing.output_columns import REMOVED_OUTPUT_LABELS, trim_worksheet_columns
from app.automation.scr_field_map import REPORT6_REQUIRED_CANONICAL, canonicalize_scr_rows
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    previous_day_report_date,
    resolve_report_dir,
    resolve_run_scoped_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report6_scr_station_processor"

TEMPLATE_FILENAME = "scr_station_original.xlsx"
TEMPLATE_MISSING_ERROR = "REPORT6_TEMPLATE_OR_FORMULA_MISSING"

INTERNAL_COLUMNS = REPORT6_REQUIRED_CANONICAL

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class Report6Processor:
    """Process SCR Station Mode Unsatisfactory Feedback dataset."""

    processor_name = PROCESSOR_NAME
    expected_mode = "Station"

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        source_rows, source_headers = self._read_csv(source_a_path)
        source_rows = canonicalize_scr_rows(source_rows)
        if source_rows:
            source_headers = sorted({k for row in source_rows for k in row})

        station_rows = [
            row for row in source_rows
            if mode_matches(
                self.expected_mode,
                row.get("complaintMode", "") or row.get("mode", "") or row.get("Mode", ""),
            )
        ]

        dup_refs = self._duplicate_complaint_refs(station_rows or source_rows)
        if dup_refs:
            return ProcessingResult(
                success=False,
                error=(
                    "REPORT6_ROW_MULTIPLICATION_DETECTED: duplicate complaintRefNo values: "
                    f"{dup_refs[:5]}"
                ),
                input_row_count=len(source_rows),
                source_a_path=str(source_a_path),
            )

        using_manual_selection = bool(
            column_selection
            and (
                column_selection.get("configuration_source") == "manual_snapshot"
                or column_selection.get("selected_column_ids")
                or column_selection.get("column_order")
            )
        )

        if not station_rows and source_rows and not using_manual_selection:
            return ProcessingResult(
                success=False,
                error=f"No {self.expected_mode} mode rows found in data",
                input_row_count=len(source_rows),
                source_a_path=str(source_a_path),
            )

        if station_rows:
            missing_required = sorted(REPORT6_REQUIRED_CANONICAL - set(source_headers))
            if missing_required:
                return ProcessingResult(
                    success=False,
                    error=f"Report 6 missing required source columns: {', '.join(missing_required)}",
                    input_row_count=len(source_rows),
                    source_a_path=str(source_a_path),
                )

        try:
            output_headers, output_rows, visible_columns, keys, config_source = (
                project_scr_for_output(
                    report_slug,
                    station_rows,
                    column_selection=column_selection,
                )
            )
        except ValueError as exc:
            return ProcessingResult(
                success=False,
                error=str(exc),
                input_row_count=len(source_rows),
                source_a_path=str(source_a_path),
            )

        if not using_manual_selection:
            removed_in_output = REMOVED_OUTPUT_LABELS & set(output_headers)
            if removed_in_output:
                return ProcessingResult(
                    success=False,
                    error=f"COLUMN_PROJECTION_FAILED: removed labels in output: {sorted(removed_in_output)}",
                    input_row_count=len(source_rows),
                    source_a_path=str(source_a_path),
                )

        input_station_count = len(station_rows) if station_rows else 0
        if output_rows and input_station_count and len(output_rows) != input_station_count:
            return ProcessingResult(
                success=False,
                error=(
                    "REPORT6_ROW_MULTIPLICATION_DETECTED: "
                    f"processed {len(output_rows)} rows from {input_station_count} source rows"
                ),
                input_row_count=len(source_rows),
                source_a_path=str(source_a_path),
            )

        log_automation_event(
            logger,
            "report6_processed_headers",
            output_headers=output_headers,
            output_row_count=len(output_rows),
            selected_column_count=len(keys),
        )

        output_headers, output_rows = prepare_output_for_rendering(
            report_slug,
            output_headers,
            output_rows,
        )

        report_date = previous_day_report_date()
        run_timestamp = datetime.now().strftime("%H%M%S")
        run_id = (column_selection or {}).get("run_id") if column_selection else None
        if run_id:
            excel_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_excel_dir, report_slug, str(run_id))
            )
            pdf_dir = ensure_directory(
                resolve_run_scoped_dir(config.output_pdf_dir, report_slug, str(run_id))
            )
        else:
            excel_dir = ensure_directory(resolve_report_dir(config.output_excel_dir, report_slug))
            pdf_dir = ensure_directory(resolve_report_dir(config.output_pdf_dir, report_slug))
        base_name = (
            f"Rail_Madad_Report_6_SCR_Station_Unsatisfactory_{report_date}_{run_timestamp}"
        )
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        template_path = self._find_template()
        use_template = (
            template_path
            and template_path.exists()
            and not using_manual_selection
        )

        try:
            if use_template:
                self._write_excel_from_template(
                    template_path,
                    excel_path,
                    output_headers,
                    output_rows,
                    report_date=report_date,
                )
            else:
                if template_path and template_path.exists() and using_manual_selection:
                    log_automation_event(
                        logger,
                        "template_skipped",
                        reason="manual_column_selection",
                        selected_column_count=len(keys),
                    )
                else:
                    log_automation_event(
                        logger,
                        "template_not_found",
                        expected_path=str(template_path) if template_path else "N/A",
                        fallback="programmatic_generation",
                    )
                self._write_excel(excel_path, output_headers, output_rows, report_date=report_date)
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(source_rows),
                success=False,
                error=f"REPORT6_EXCEL_GENERATION_FAILED: {exc}",
                source_a_path=str(source_a_path),
                source_a_rows=len(source_rows),
            )

        try:
            self._verify_excel_output(
                excel_path,
                expected_headers=output_headers,
                expected_row_count=len(output_rows),
            )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(source_rows),
                success=False,
                error=f"REPORT6_EXCEL_VALIDATION_FAILED: {exc}",
                source_a_path=str(source_a_path),
                source_a_rows=len(source_rows),
                excel_path=str(excel_path),
            )

        try:
            self._write_pdf(pdf_path, output_headers, output_rows, report_date=report_date)
            verify_report_output(
                report_slug=report_slug,
                headers=output_headers,
                rows=output_rows,
                pdf_path=pdf_path,
                excel_path=excel_path,
            )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(source_rows),
                success=False,
                error=f"REPORT6_PDF_VALIDATION_FAILED: {exc}",
                source_a_path=str(source_a_path),
                source_a_rows=len(source_rows),
                excel_path=str(excel_path),
            )

        log_automation_event(
            logger,
            "phase8_dataset_loaded",
            source_a=str(source_a_path),
            source_headers=source_headers,
            output_headers=output_headers,
            input_row_count=len(source_rows),
            output_row_count=len(output_rows),
            mode=self.expected_mode,
        )

        return ProcessingResult(
            success=True,
            input_row_count=len(source_rows),
            processed_row_count=len(output_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_a_rows=len(source_rows),
            output_columns=output_headers,
            visible_columns=visible_columns,
            selected_column_ids=keys,
            column_order=list(keys),
            configuration_source=config_source,
        )

    def _find_template(self) -> Path | None:
        """Find the template workbook."""
        possible_paths = [
            Path(config.output_excel_dir).parent / "templates" / TEMPLATE_FILENAME,
            Path(__file__).parent.parent.parent / "infrastructure" / "seed" / "sample_workbooks" / TEMPLATE_FILENAME,
            Path(config.output_excel_dir) / TEMPLATE_FILENAME,
        ]
        for path in possible_paths:
            if path.exists():
                return path
        return None

    @staticmethod
    def _duplicate_complaint_refs(rows: list[dict[str, str]]) -> list[str]:
        seen: dict[str, int] = {}
        for row in rows:
            ref = str(
                row.get("complaintRefNo")
                or row.get("Ref. No.")
                or row.get("complaint_ref_no")
                or ""
            ).strip()
            if not ref:
                continue
            seen[ref] = seen.get(ref, 0) + 1
        return sorted(ref for ref, count in seen.items() if count > 1)

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    def _write_excel_from_template(
        self,
        template_path: Path,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        """Write Excel using template column widths; layout matches programmatic output."""
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = load_workbook(str(template_path))
        worksheet = workbook.active

        title = normalize_report_title(
            f"Rail Madad Report No 6 - SCR {self.expected_mode} Mode Unsatisfactory Feedback "
            f"on date {report_date}",
            report_slug="scr-station",
        )
        for merged in list(worksheet.merged_cells.ranges):
            if merged.min_row <= 2:
                worksheet.unmerge_cells(str(merged))
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(len(headers), 1),
        )
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        header_row = 2
        data_start_row = 3
        if worksheet.max_row >= data_start_row:
            for row_idx in range(data_start_row, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=None)

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        trim_worksheet_columns(worksheet, keep_columns=len(headers))

        for row_offset, row_values in enumerate(rows):
            row_idx = data_start_row + row_offset
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.fill = NO_FILL

        clear_complaint_data_row_formatting(
            worksheet,
            start_row=data_start_row,
            end_row=worksheet.max_row,
            start_col=1,
            end_col=len(headers),
        )

        apply_column_formatting(
            worksheet,
            headers,
            header_row=header_row,
            data_start_row=data_start_row,
            wrap_all_data=True,
        )
        apply_report_print_setup(worksheet, col_count=len(headers))

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report 6"

        title = normalize_report_title(
            f"Rail Madad Report No 6 - SCR {self.expected_mode} Mode Unsatisfactory Feedback "
            f"on date {report_date}",
            report_slug="scr-station",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        header_row = 2
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        data_start = 3
        for row_offset, row_values in enumerate(rows):
            row_idx = data_start + row_offset
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.fill = NO_FILL

        clear_complaint_data_row_formatting(
            worksheet,
            start_row=data_start,
            end_row=worksheet.max_row,
            start_col=1,
            end_col=len(headers),
        )

        apply_column_formatting(
            worksheet,
            headers,
            header_row=header_row,
            data_start_row=data_start,
            wrap_all_data=True,
        )
        apply_report_print_setup(worksheet, col_count=len(headers))

        workbook.save(temp_path)
        temp_path.replace(target_path)

    @staticmethod
    def _verify_excel_output(
        path: Path,
        *,
        expected_headers: list[str],
        expected_row_count: int,
    ) -> None:
        if not path.is_file() or path.stat().st_size <= 0:
            raise ValueError("workbook missing or empty")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            if not first_row:
                raise ValueError("workbook has no rows")
            first_cells = [str(cell).strip() if cell is not None else "" for cell in first_row]
            if len(first_cells) == 1 or any("Rail Madad Report" in cell for cell in first_cells):
                header_row = next(rows_iter, None)
            else:
                header_row = first_row
            if not header_row:
                raise ValueError("workbook has no header row")
            headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
            if headers != expected_headers:
                raise ValueError(
                    f"header mismatch expected={expected_headers} actual={headers}"
                )
            data_rows = list(rows_iter)
            if len(data_rows) != expected_row_count:
                raise ValueError(
                    f"row count mismatch expected={expected_row_count} actual={len(data_rows)}"
                )
            if expected_headers:
                if headers[0] != expected_headers[0] or headers[-1] != expected_headers[-1]:
                    raise ValueError("first or last selected column missing from workbook")
        finally:
            workbook.close()

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")

        table_data: list[list[str]] = [headers, *rows]

        style_commands: list[tuple] = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]

        table, pagesize, margin = build_wrapped_fitted_table(
            table_data,
            style_commands,
        )
        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        story = [
            Paragraph(
                normalize_report_title(
                    f"Rail Madad Report No 6 - SCR {self.expected_mode} Mode Unsatisfactory Feedback on date {report_date}",
                    report_slug="scr-station",
                ),
                pdf_title_style("Report6Title"),
            ),
            Spacer(1, 12),
            table,
        ]
        doc.build(story)
        temp_path.replace(target_path)
