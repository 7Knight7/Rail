"""Report 1 post-ingestion processor (Comprehensive + Feedback merge)."""

from __future__ import annotations

import csv
import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from app.automation.formatting.pdf_fonts import pdf_font_bold, pdf_title_style
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.automation.config import config
from app.automation.formatting.excel_print import apply_column_formatting, apply_report_print_setup
from app.automation.formatting.pdf_table import build_fitted_table
from app.automation.formatting.text_pipeline import (
    normalize_report_title,
    prepare_output_for_rendering,
    verify_text_rendering,
)
from app.automation.formatting.pdf_verify import verify_report_output
from app.automation.formatting.scr import SCR_FILL, SCR_FONT, row_contains_scr
from app.automation.formatting.serial import apply_serial_number, is_serial_header, renumber_data_rows
from app.automation.processing.base import ProcessingResult
from app.automation.processing.column_config import project_for_output, resolve_projection_column_keys
from app.automation.processing.output_columns import (
    REPORT1_HIDDEN_COLUMNS,
    SOURCE_B_DATA_COLUMNS,
    build_merged_total_row,
    fill_report1_avg_disposal_time_total,
)

HIDDEN_COLUMNS = REPORT1_HIDDEN_COLUMNS
from app.automation.utils import (
    ensure_directory,
    log_automation_event,
    previous_day_report_date,
    resolve_report_dir,
)

logger = logging.getLogger(__name__)

PROCESSOR_NAME = "report1_zone_wise_processor"

FEEDBACK_FILENAME = "report1_feedback_zone_raw.csv"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class Report1Processor:
    """Merge Comprehensive + Feedback datasets and emit formatted Excel/PDF."""

    processor_name = PROCESSOR_NAME

    def process(
        self,
        *,
        source_a_path: Path,
        report_slug: str,
        source_b_path: Path | None = None,
        column_selection: dict | None = None,
    ) -> ProcessingResult:
        if source_a_path.suffix.lower() == ".pdf":
            return ProcessingResult(success=False, error="PDF cannot be used as processing input")

        # Require explicit current-run Source B path — no filesystem fallback to stale CSVs.
        if source_b_path is None:
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                error="Feedback dataset missing: source_b_path required (no stale fallback)",
            )
        feedback_path = Path(source_b_path)
        if not feedback_path.exists() or feedback_path.stat().st_size <= 0:
            return ProcessingResult(
                success=False,
                source_a_path=str(source_a_path),
                source_b_path=str(feedback_path),
                error=f"Feedback dataset missing or empty: {feedback_path}",
            )

        source_a_rows, source_a_headers = self._read_csv(source_a_path)
        source_b_rows, source_b_headers = self._read_csv(feedback_path)
        data_a, _ = self._split_total_row(source_a_rows)
        data_b, _ = self._split_total_row(source_b_rows)

        merged_headers, merged_rows, scr_row_flags = self.build_merged_table(
            source_a_rows,
            source_a_headers,
            source_b_rows,
            source_b_headers,
        )
        log_automation_event(
            logger,
            "report1_merge_completed",
            source_a_rows=len(data_a),
            source_b_rows=len(data_b),
            merged_row_count=len(merged_rows),
        )
        selected_keys, config_source = resolve_projection_column_keys(
            report_slug,
            column_selection=column_selection,
        )
        output_headers, output_rows, visible_columns, selected_column_ids, config_source = (
            project_for_output(
                report_slug,
                full_headers=merged_headers,
                rows=merged_rows,
                selected_keys=selected_keys,
                config_source=config_source,
            )
        )
        log_automation_event(
            logger,
            "report1_columns_projected",
            selected_column_count=len(selected_column_ids),
            selected_column_ids=selected_column_ids,
            configuration_source=config_source,
            final_headers=output_headers,
            final_row_count=len(output_rows),
        )

        if any(is_serial_header(header) for header in output_headers):
            output_rows = renumber_data_rows(output_headers, output_rows)
            log_automation_event(
                logger,
                "report1_serial_numbers_regenerated",
                final_row_count=len(output_rows),
            )

        scr_row_indices = {
            idx
            for idx, is_scr in enumerate(scr_row_flags)
            if is_scr and idx < len(output_rows)
        }
        if scr_row_indices:
            log_automation_event(
                logger,
                "report1_scr_highlight_applied",
                scr_row_indices=sorted(scr_row_indices),
            )

        output_headers, output_rows = prepare_output_for_rendering(
            report_slug,
            output_headers,
            output_rows,
        )

        report_date = previous_day_report_date()
        run_timestamp = datetime.now().strftime("%H%M%S")
        excel_dir = ensure_directory(
            resolve_report_dir(config.output_excel_dir, report_slug)
        )
        pdf_dir = ensure_directory(
            resolve_report_dir(config.output_pdf_dir, report_slug)
        )
        base_name = (
            f"Rail_Madad_Report_1_Zone_Wise_Complaints_Feedback_{report_date}_{run_timestamp}"
        )
        excel_path = excel_dir / f"{base_name}.xlsx"
        pdf_path = pdf_dir / f"{base_name}.pdf"

        try:
            self._write_excel(
                excel_path,
                output_headers,
                output_rows,
                report_date=report_date,
                scr_row_indices=scr_row_indices,
            )
            self._write_pdf(
                pdf_path,
                output_headers,
                output_rows,
                report_date=report_date,
                scr_row_indices=scr_row_indices,
            )
            verify_report_output(
                report_slug=report_slug,
                headers=output_headers,
                rows=output_rows,
                pdf_path=pdf_path,
                excel_path=excel_path,
            )
        except Exception as exc:
            return ProcessingResult(
                input_row_count=len(data_a) + len(data_b),
                success=False,
                error=str(exc),
                source_a_path=str(source_a_path),
                source_b_path=str(feedback_path),
                source_a_rows=len(data_a),
                source_b_rows=len(data_b),
            )

        log_automation_event(
            logger,
            "phase8_dataset_loaded",
            source_a=str(source_a_path),
            source_b=str(feedback_path),
            input_row_count=len(data_a) + len(data_b),
        )

        return ProcessingResult(
            success=True,
            input_row_count=len(data_a) + len(data_b),
            processed_row_count=len(merged_rows),
            excel_path=str(excel_path),
            pdf_path=str(pdf_path),
            source_a_path=str(source_a_path),
            source_b_path=str(feedback_path),
            source_a_rows=len(data_a),
            source_b_rows=len(data_b),
            output_columns=output_headers,
            visible_columns=visible_columns,
            selected_column_ids=selected_column_ids,
            column_order=list(selected_column_ids),
            configuration_source=config_source,
        )

    def build_merged_table(
        self,
        source_a_rows: list[dict[str, str]],
        source_a_headers: list[str],
        source_b_rows: list[dict[str, str]],
        source_b_headers: list[str],
    ) -> tuple[list[str], list[list[str]], list[bool]]:
        """Build full merged table including totals (pre-output projection)."""
        data_a, total_a = self._split_total_row(source_a_rows)
        data_b, total_b = self._split_total_row(source_b_rows)
        merged_headers = source_a_headers + ["S.No.", "Organisation"] + SOURCE_B_DATA_COLUMNS
        merged_rows, scr_flags = self._merge_rows(data_a, source_a_headers, data_b, source_b_headers)
        total_row = build_merged_total_row(
            merged_headers=merged_headers,
            data_rows=merged_rows,
            source_a_headers=source_a_headers,
            source_b_headers=source_b_headers,
            total_a=total_a,
            total_b=total_b,
            source_b_columns=SOURCE_B_DATA_COLUMNS,
            org_label_a="Total",
            org_label_b="Total",
        )
        fill_report1_avg_disposal_time_total(
            total_row,
            source_a_headers=source_a_headers,
            data_rows=merged_rows,
            total_a=total_a,
        )
        merged_rows.append(total_row)
        scr_flags.append(False)
        return merged_headers, merged_rows, scr_flags

    def _find_feedback_csv(self, report_slug: str) -> Path | None:
        extracted_dir = resolve_report_dir(config.extracted_data_dir, report_slug)
        preferred = extracted_dir / FEEDBACK_FILENAME
        if preferred.exists():
            return preferred
        matches = sorted(extracted_dir.glob("*feedback*.csv"))
        return matches[0] if matches else None

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            rows = [{header: row.get(header, "") for header in headers} for row in reader]
        return rows, headers

    @staticmethod
    def _normalize_org(name: str) -> str:
        return re.sub(r"\s+", " ", name.strip()).lower()

    @staticmethod
    def _is_total_row(row: dict[str, str]) -> bool:
        org = row.get("Organisation", "")
        return "total" in org.strip().lower()

    def _split_total_row(
        self,
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], dict[str, str] | None]:
        if not rows:
            return [], None
        if self._is_total_row(rows[-1]):
            return rows[:-1], rows[-1]
        return rows, None

    @staticmethod
    def _is_irctc_catering(org: str) -> bool:
        normalized = Report1Processor._normalize_org(org)
        return "irctc-catering" in normalized or normalized == "irctc catering"

    @staticmethod
    def _is_irctc_online(org: str) -> bool:
        normalized = Report1Processor._normalize_org(org)
        return "irctc-online" in normalized or normalized == "irctc online"

    @staticmethod
    def _is_irctc_combined(org: str) -> bool:
        return Report1Processor._normalize_org(org) == "irctc"

    def _build_feedback_lookup(
        self,
        source_b_rows: list[dict[str, str]],
    ) -> dict[str, dict[str, str]]:
        lookup: dict[str, dict[str, str]] = {}
        irctc_row: dict[str, str] | None = None
        for row in source_b_rows:
            org = row.get("Organisation", "")
            if self._is_irctc_combined(org):
                irctc_row = row
                continue
            lookup[self._normalize_org(org)] = row
        if irctc_row is not None:
            lookup["__irctc__"] = irctc_row
        return lookup

    def _feedback_values_for_row(
        self,
        org: str,
        lookup: dict[str, dict[str, str]],
        *,
        is_second_irctc: bool,
    ) -> list[str]:
        if is_second_irctc:
            return [""] * len(SOURCE_B_DATA_COLUMNS)

        if self._is_irctc_catering(org) and "__irctc__" in lookup:
            row = lookup["__irctc__"]
        else:
            row = lookup.get(self._normalize_org(org), {})

        return [row.get(column, "") for column in SOURCE_B_DATA_COLUMNS]

    @staticmethod
    def _is_scr_organisation(org: str) -> bool:
        from app.automation.formatting.scr import SCR_PATTERN

        text = re.sub(r"\s+", " ", org.strip())
        return bool(SCR_PATTERN.search(text))

    def _merge_rows(
        self,
        source_a_rows: list[dict[str, str]],
        source_a_headers: list[str],
        source_b_rows: list[dict[str, str]],
        source_b_headers: list[str],
    ) -> tuple[list[list[str]], list[bool]]:
        lookup = self._build_feedback_lookup(source_b_rows)
        merged: list[list[str]] = []
        scr_flags: list[bool] = []

        for index, row in enumerate(source_a_rows, start=1):
            org = row.get("Organisation", "")
            is_second_irctc = self._is_irctc_online(org)

            source_a_values = apply_serial_number(
                source_a_headers,
                [row.get(header, "") for header in source_a_headers],
                index,
            )
            b_sno = str(index)
            b_org = org
            b_values = self._feedback_values_for_row(
                org,
                lookup,
                is_second_irctc=is_second_irctc,
            )
            merged.append(source_a_values + [b_sno, b_org] + b_values)
            scr_flags.append(self._is_scr_organisation(org))

        return merged, scr_flags

    def _write_excel(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
        scr_row_indices: set[int] | None = None,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report 1"

        title = normalize_report_title(
            f"Rail Madad Report No 1 - Zone Wise Complaints & Feedback Report "
            f"on date {report_date}",
            report_slug="report1",
        )
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        header_row = 2
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        data_start = 3
        for row_offset, row_values in enumerate(rows):
            row_idx = data_start + row_offset
            for col_idx, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].hidden = False

        if scr_row_indices:
            for row_offset in scr_row_indices:
                if row_offset >= len(rows):
                    continue
                row_idx = data_start + row_offset
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = SCR_FILL
                    cell.font = SCR_FONT
        else:
            from app.automation.formatting.scr import highlight_south_central_railway_rows

            highlight_south_central_railway_rows(
                worksheet,
                start_row=data_start,
                end_row=worksheet.max_row,
                start_col=1,
                end_col=len(headers),
            )

        if rows:
            totals_row = worksheet.max_row
            totals_fill = PatternFill(fill_type="solid", fgColor="E8E8E8")
            for col_idx in range(1, len(headers) + 1):
                cell = worksheet.cell(row=totals_row, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = totals_fill

        apply_column_formatting(
            worksheet,
            headers,
            header_row=header_row,
            data_start_row=data_start,
        )
        apply_report_print_setup(worksheet, col_count=len(headers))

        workbook.save(temp_path)
        temp_path.replace(target_path)

    def _write_pdf(
        self,
        target_path: Path,
        headers: list[str],
        rows: list[list[str]],
        *,
        report_date: str,
        scr_row_indices: set[int] | None = None,
    ) -> None:
        temp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        visible_headers = headers
        visible_indices = list(range(1, len(headers) + 1))

        table_data: list[list[str]] = [visible_headers]
        pdf_scr_indices: set[int] = set()
        for row_offset, row_values in enumerate(rows):
            visible_row = [row_values[idx - 1] for idx in visible_indices]
            table_data.append(visible_row)
            if scr_row_indices is not None:
                if row_offset in scr_row_indices:
                    pdf_scr_indices.add(len(table_data) - 1)
            elif row_contains_scr(visible_row):
                pdf_scr_indices.add(len(table_data) - 1)

        style_commands: list[tuple] = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
        for row_idx in pdf_scr_indices:
            style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.yellow))
            style_commands.append(("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.black))
        if rows:
            style_commands.append(
                ("FONTNAME", (0, len(table_data) - 1), (-1, len(table_data) - 1), pdf_font_bold())
            )
            style_commands.append(
                ("BACKGROUND", (0, len(table_data) - 1), (-1, len(table_data) - 1), colors.HexColor("#E8E8E8"))
            )

        table, pagesize, margin = build_fitted_table(table_data, style_commands)
        doc = SimpleDocTemplate(
            str(temp_path),
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
        )
        story = [
            Paragraph(
                normalize_report_title(
                    f"Rail Madad Report No 1 - Zone Wise Complaints & Feedback Report "
                    f"on date {report_date}",
                    report_slug="report1",
                ),
                pdf_title_style("Report1Title"),
            ),
            Spacer(1, 12),
            table,
        ]
        doc.build(story)
        temp_path.replace(target_path)
