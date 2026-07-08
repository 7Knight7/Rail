"""Parse RailMadad spreadsheet headers and infer column metadata."""

from __future__ import annotations

import csv
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import ValidationError

ColumnDataType = str  # text | number | date | status | boolean

STATUS_KEYWORDS = frozenset(
    {"status", "state", "current status", "complaint status", "feedback status"}
)
DATE_KEYWORDS = frozenset(
    {"date", "registered", "closed", "created", "updated", "registration", "closure"}
)
BOOLEAN_KEYWORDS = frozenset({"is ", "has ", "flag", "active", "enabled", "resolved"})
NUMBER_KEYWORDS = frozenset(
    {"count", "no", "number", "amount", "score", "total", "qty", "quantity", "age"}
)


@dataclass(frozen=True)
class ParsedColumn:
    id: str
    field_name: str
    display_name: str
    data_type: ColumnDataType
    filterable: bool
    sortable: bool


def _slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.strip().lower())
    return slug.strip("_") or "column"


def _infer_type(header: str, samples: list[Any]) -> ColumnDataType:
    lower = header.strip().lower()

    if any(keyword in lower for keyword in BOOLEAN_KEYWORDS):
        return "boolean"

    if lower.endswith(" status") or lower in STATUS_KEYWORDS or "status" in lower:
        return "status"

    if any(keyword in lower for keyword in DATE_KEYWORDS):
        return "date"

    non_empty = [value for value in samples if value is not None and str(value).strip() != ""]
    if not non_empty:
        if any(keyword in lower for keyword in NUMBER_KEYWORDS):
            return "number"
        return "text"

    bool_values = {str(value).strip().lower() for value in non_empty}
    if bool_values.issubset({"true", "false", "yes", "no", "1", "0", "y", "n"}):
        return "boolean"

    numeric_count = 0
    date_count = 0
    for value in non_empty[:20]:
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric_count += 1
            continue
        if isinstance(value, (datetime, date)):
            date_count += 1
            continue
        text = str(value).strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", text):
            numeric_count += 1
            continue
        if _looks_like_date(text):
            date_count += 1

    if date_count >= max(1, len(non_empty) // 2):
        return "date"
    if numeric_count >= max(1, len(non_empty) // 2):
        return "number"
    if any(keyword in lower for keyword in NUMBER_KEYWORDS):
        return "number"
    return "text"


def _looks_like_date(value: str) -> bool:
    patterns = (
        r"^\d{4}-\d{2}-\d{2}",
        r"^\d{2}[/-]\d{2}[/-]\d{4}",
        r"^\d{2}-\d{2}-\d{4}",
    )
    return any(re.match(pattern, value) for pattern in patterns)


def _build_columns(headers: list[str], sample_rows: list[list[Any]]) -> list[ParsedColumn]:
    columns: list[ParsedColumn] = []
    seen_fields: dict[str, int] = {}

    for index, header in enumerate(headers):
        display_name = str(header).strip()
        if not display_name:
            continue

        field_name = _slugify(display_name)
        if field_name in seen_fields:
            seen_fields[field_name] += 1
            field_name = f"{field_name}_{seen_fields[field_name]}"
        else:
            seen_fields[field_name] = 1

        samples = [row[index] if index < len(row) else None for row in sample_rows]
        data_type = _infer_type(display_name, samples)

        columns.append(
            ParsedColumn(
                id=str(uuid.uuid4()),
                field_name=field_name,
                display_name=display_name,
                data_type=data_type,
                filterable=True,
                sortable=True,
            )
        )

    return columns


class SpreadsheetParser:
    def parse_file(self, file_path: Path, *, header_row: int = 1, sheet_name: str | None = None) -> list[ParsedColumn]:
        suffix = file_path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self._parse_xlsx(file_path, header_row=header_row, sheet_name=sheet_name)
        if suffix == ".csv":
            return self._parse_csv(file_path, header_row=header_row)
        raise ValidationError(f"Unsupported spreadsheet format: {suffix}")

    def parse_bytes(
        self,
        content: bytes,
        filename: str,
        *,
        header_row: int = 1,
        sheet_name: str | None = None,
    ) -> list[ParsedColumn]:
        suffix = Path(filename).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self._parse_xlsx_bytes(content, header_row=header_row, sheet_name=sheet_name)
        if suffix == ".csv":
            return self._parse_csv_bytes(content, header_row=header_row)
        raise ValidationError(f"Unsupported spreadsheet format: {suffix}")

    def _parse_xlsx(self, file_path: Path, *, header_row: int, sheet_name: str | None) -> list[ParsedColumn]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            return self._parse_rows(rows, header_row=header_row)
        finally:
            workbook.close()

    def _parse_xlsx_bytes(self, content: bytes, *, header_row: int, sheet_name: str | None) -> list[ParsedColumn]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            return self._parse_rows(rows, header_row=header_row)
        finally:
            workbook.close()

    def _parse_csv(self, file_path: Path, *, header_row: int) -> list[ParsedColumn]:
        text = file_path.read_text(encoding="utf-8-sig")
        return self._parse_csv_bytes(text.encode("utf-8"), header_row=header_row)

    def _parse_csv_bytes(self, content: bytes, *, header_row: int) -> list[ParsedColumn]:
        text = content.decode("utf-8-sig")
        reader = csv.reader(StringIO(text))
        rows = list(reader)
        return self._parse_rows(rows, header_row=header_row)

    def _parse_rows(self, rows: list[Any], *, header_row: int) -> list[ParsedColumn]:
        if not rows:
            raise ValidationError("Spreadsheet is empty")

        header_index = max(header_row - 1, 0)
        if header_index >= len(rows):
            raise ValidationError(f"Header row {header_row} is out of range")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
        if not any(headers):
            raise ValidationError("No column headers found in spreadsheet")

        sample_rows = rows[header_index + 1 : header_index + 51]
        return _build_columns(headers, sample_rows)


def parse_spreadsheet_headers(
    source: Path | bytes,
    *,
    filename: str | None = None,
    header_row: int = 1,
    sheet_name: str | None = None,
) -> list[ParsedColumn]:
    parser = SpreadsheetParser()
    if isinstance(source, Path):
        return parser.parse_file(source, header_row=header_row, sheet_name=sheet_name)
    if not filename:
        raise ValidationError("Filename is required when parsing spreadsheet bytes")
    return parser.parse_bytes(source, filename, header_row=header_row, sheet_name=sheet_name)
