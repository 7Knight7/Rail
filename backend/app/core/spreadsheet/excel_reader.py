"""Read RailMadad Excel files and return the original dataset unchanged."""

from __future__ import annotations

import uuid
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import ValidationError
from app.core.spreadsheet.parser import ColumnDataType, _infer_type

SUPPORTED_EXCEL_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


@dataclass(frozen=True)
class ExcelColumn:
    """Column metadata with the exact header name from the source file."""

    index: int
    name: str
    data_type: ColumnDataType
    filterable: bool = True
    sortable: bool = True
    id: str = field(default_factory=lambda: str(uuid.uuid4()))

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["display_name"] = self.name
        payload["field_name"] = self.name
        return payload


@dataclass(frozen=True)
class ExcelDatasetMetadata:
    file_path: str
    filename: str
    sheet_name: str
    header_row: int
    row_count: int
    column_count: int
    parsed_at: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ExcelDataset:
    columns: list[ExcelColumn]
    rows: list[list[Any]]
    metadata: ExcelDatasetMetadata

    def to_dict(self) -> dict[str, Any]:
        return {
            "columns": [column.to_dict() for column in self.columns],
            "rows": self.rows,
            "metadata": self.metadata.to_dict(),
        }


class ExcelReader:
    """Load an Excel workbook and return the original dataset without transformation."""

    def read(
        self,
        file_path: str | Path,
        *,
        header_row: int = 1,
        sheet_name: str | None = None,
    ) -> ExcelDataset:
        path = Path(file_path)
        if not path.exists():
            raise ValidationError(f"Excel file not found: {path}")

        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_EXCEL_SUFFIXES:
            raise ValidationError(f"Unsupported Excel format: {suffix}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            if worksheet is None:
                raise ValidationError("Workbook does not contain any worksheets")

            raw_rows = list(worksheet.iter_rows(values_only=True))
            if not raw_rows:
                raise ValidationError("Excel file is empty")

            header_index = max(header_row - 1, 0)
            if header_index >= len(raw_rows):
                raise ValidationError(f"Header row {header_row} is out of range")

            header_cells = raw_rows[header_index]
            columns = self._build_columns(header_cells, raw_rows, header_index)
            if not columns:
                raise ValidationError("No columns found in Excel header row")

            data_rows = raw_rows[header_index + 1 :]
            data_rows = self._trim_trailing_empty_rows(data_rows)
            rows = [
                self._extract_row_values(row, len(columns))
                for row in data_rows
            ]

            metadata = ExcelDatasetMetadata(
                file_path=str(path.resolve()),
                filename=path.name,
                sheet_name=worksheet.title,
                header_row=header_row,
                row_count=len(rows),
                column_count=len(columns),
                parsed_at=datetime.now(UTC).isoformat(),
            )
            return ExcelDataset(columns=columns, rows=rows, metadata=metadata)
        finally:
            workbook.close()

    def _build_columns(
        self,
        header_cells: tuple[Any, ...],
        raw_rows: list[tuple[Any, ...]],
        header_index: int,
    ) -> list[ExcelColumn]:
        columns: list[ExcelColumn] = []
        sample_rows = raw_rows[header_index + 1 : header_index + 51]

        for index, cell in enumerate(header_cells):
            name = "" if cell is None else str(cell)
            samples = [
                row[index] if index < len(row) else None
                for row in sample_rows
            ]
            columns.append(
                ExcelColumn(
                    index=index,
                    name=name,
                    data_type=_infer_type(name, samples),
                )
            )

        return columns

    def _extract_row_values(self, row: tuple[Any, ...], column_count: int) -> list[Any]:
        values: list[Any] = []
        for index in range(column_count):
            if index < len(row):
                values.append(row[index])
            else:
                values.append(None)
        return values

    def _trim_trailing_empty_rows(self, rows: list[tuple[Any, ...]]) -> list[tuple[Any, ...]]:
        trimmed = list(rows)
        while trimmed and self._is_empty_row(trimmed[-1]):
            trimmed.pop()
        return trimmed

    @staticmethod
    def _is_empty_row(row: tuple[Any, ...]) -> bool:
        return all(cell is None or str(cell).strip() == "" for cell in row)


def read_excel_file(
    file_path: str | Path,
    *,
    header_row: int = 1,
    sheet_name: str | None = None,
) -> ExcelDataset:
    """Convenience wrapper around :class:`ExcelReader`."""
    return ExcelReader().read(file_path, header_row=header_row, sheet_name=sheet_name)
