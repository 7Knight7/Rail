import hashlib
import json
import logging
from datetime import UTC, datetime
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ValidationError
from app.core.spreadsheet.parser import SpreadsheetParser
from app.features.datasets.schemas import ColumnMetadata, DatasetMetadataResponse
from app.features.uploads.service import UploadService
from app.infrastructure.database.models import ReportDatasetModel

logger = logging.getLogger(__name__)

SUPPORTED_REPORT_IDS = frozenset(
    {
        "merging",
        "division",
        "division_feedback",
        "train-no",
        "types",
        "scr-train",
        "scr-station",
        "report1",
        "report1_feedback",
    }
)


def file_content_checksum(file_path: Path) -> str:
    """SHA256 of file contents for idempotent ingestion."""
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class DatasetRepository:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    async def get_by_report_id(self, report_id: str) -> ReportDatasetModel | None:
        result = await self._session.execute(
            select(ReportDatasetModel).where(ReportDatasetModel.report_id == report_id).limit(1)
        )
        return result.scalar_one_or_none()

    async def upsert(
        self,
        *,
        report_id: str,
        source_filename: str,
        source_file_path: str | None,
        header_row: int,
        row_count: int,
        columns: list[ColumnMetadata],
        content_checksum: str | None = None,
    ) -> ReportDatasetModel:
        existing = await self.get_by_report_id(report_id)
        payload = json.dumps([column.model_dump(by_alias=True) for column in columns])

        if existing:
            existing.source_filename = source_filename
            existing.source_file_path = source_file_path
            existing.header_row = header_row
            existing.row_count = row_count
            existing.columns_json = payload
            existing.parsed_at = datetime.now(UTC)
            if content_checksum is not None:
                existing.content_checksum = content_checksum
            await self._session.commit()
            await self._session.refresh(existing)
            return existing

        model = ReportDatasetModel(
            report_id=report_id,
            source_filename=source_filename,
            source_file_path=source_file_path,
            header_row=header_row,
            row_count=row_count,
            columns_json=payload,
            content_checksum=content_checksum,
        )
        self._session.add(model)
        await self._session.commit()
        await self._session.refresh(model)
        return model


class DatasetService:
    def __init__(self, session: AsyncSession, upload_service: UploadService | None = None) -> None:
        self._repository = DatasetRepository(session)
        self._upload_service = upload_service or UploadService()
        self._parser = SpreadsheetParser()

    def _canonical_report_id(self, report_id: str) -> str:
        from app.automation.report_keys import canonicalize_report_key

        return canonicalize_report_key(report_id)

    def _validate_report_id(self, report_id: str) -> str:
        canonical = self._canonical_report_id(report_id)
        if canonical not in SUPPORTED_REPORT_IDS:
            raise NotFoundError("Report dataset", report_id)
        return canonical

    async def ensure_dataset_exists(self, report_id: str) -> ReportDatasetModel:
        """Idempotently ensure a dataset row exists (seed from template if missing)."""
        from app.automation.report_keys import DATASET_TEMPLATES
        from app.infrastructure.seed.seed_report_datasets import SAMPLE_DIR, _ensure_workbook

        canonical = self._validate_report_id(report_id)
        existing = await self._repository.get_by_report_id(canonical)
        if existing is not None:
            return existing

        template_name = DATASET_TEMPLATES.get(canonical)
        if template_name is None:
            return await self._repository.upsert(
                report_id=canonical,
                source_filename=f"{canonical}_placeholder.csv",
                source_file_path=None,
                header_row=1,
                row_count=0,
                columns=[],
            )

        file_path = SAMPLE_DIR / template_name
        _ensure_workbook(file_path)
        await self.ingest_file(
            canonical,
            file_path=file_path,
            source_filename=template_name,
            header_row=1,
        )
        model = await self._repository.get_by_report_id(canonical)
        if model is None:
            raise NotFoundError("Report dataset", canonical)
        logger.info("Ensured dataset exists for report %s via template %s", canonical, template_name)
        return model

    def _to_response(self, model: ReportDatasetModel) -> DatasetMetadataResponse:
        columns_data = json.loads(model.columns_json)
        return DatasetMetadataResponse(
            reportId=model.report_id,
            sourceFilename=model.source_filename,
            headerRow=model.header_row,
            rowCount=model.row_count,
            columns=[ColumnMetadata.model_validate(item) for item in columns_data],
            parsedAt=model.parsed_at.isoformat(),
        )

    async def get_metadata(self, report_id: str) -> DatasetMetadataResponse:
        canonical = self._validate_report_id(report_id)
        model = await self._repository.get_by_report_id(canonical)
        if not model:
            raise NotFoundError("Report dataset", report_id)
        return self._to_response(model)

    async def ingest_upload(
        self,
        report_id: str,
        *,
        upload_id: str,
        header_row: int = 1,
        sheet_name: str | None = None,
    ) -> DatasetMetadataResponse:
        canonical = self._validate_report_id(report_id)

        for extension in (".xlsx", ".xls", ".csv"):
            try:
                file_path = await self._upload_service.get_file_path(upload_id, extension)
                break
            except ValidationError:
                file_path = None
        else:
            raise NotFoundError("Upload", upload_id)

        return await self.ingest_file(
            canonical,
            file_path=file_path,
            source_filename=file_path.name,
            header_row=header_row,
            sheet_name=sheet_name,
        )

    async def ingest_file(
        self,
        report_id: str,
        *,
        file_path: Path,
        source_filename: str,
        header_row: int = 1,
        sheet_name: str | None = None,
    ) -> DatasetMetadataResponse:
        canonical = self._validate_report_id(report_id)
        path = Path(file_path)
        if not path.is_file():
            raise ValidationError(f"Ingest file not found: {path}")
        if path.suffix.lower() == ".pdf":
            raise ValidationError("PDF cannot be ingested into datasets")

        checksum = file_content_checksum(path)
        existing = await self._repository.get_by_report_id(canonical)
        if existing is not None:
            stored = existing.content_checksum
            same_path = (existing.source_file_path or "") == str(path.resolve())
            if stored and stored == checksum and same_path:
                logger.info("Skipping unchanged ingest for %s (checksum match)", canonical)
                return self._to_response(existing)

        columns, row_count = self._inspect_file(
            path, header_row=header_row, sheet_name=sheet_name
        )

        model = await self._repository.upsert(
            report_id=canonical,
            source_filename=source_filename,
            source_file_path=str(path.resolve()),
            header_row=header_row,
            row_count=row_count,
            columns=columns,
            content_checksum=checksum,
        )
        if model.row_count != row_count:
            raise ValidationError(
                f"Ingest row_count mismatch for {canonical}: db={model.row_count} csv={row_count}"
            )
        logger.info(
            "Ingested dataset metadata for report %s (%s columns, %s rows)",
            canonical,
            len(columns),
            row_count,
        )
        return self._to_response(model)

    def _inspect_file(
        self,
        file_path: Path,
        *,
        header_row: int,
        sheet_name: str | None = None,
    ) -> tuple[list[ColumnMetadata], int]:
        """Single-pass column discovery + row count for CSV; one parse for workbooks."""
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._inspect_csv(file_path, header_row=header_row)

        parsed_columns = self._parser.parse_file(
            file_path,
            header_row=header_row,
            sheet_name=sheet_name,
        )
        columns = [
            ColumnMetadata(
                id=column.id,
                fieldName=column.field_name,
                displayName=column.display_name,
                dataType=column.data_type,
                filterable=column.filterable,
                sortable=column.sortable,
            )
            for column in parsed_columns
        ]
        row_count = self._count_rows(file_path, header_row=header_row)
        return columns, row_count

    def _inspect_csv(
        self, file_path: Path, *, header_row: int
    ) -> tuple[list[ColumnMetadata], int]:
        import csv as csv_mod

        with file_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv_mod.reader(handle)
            rows = list(reader)

        if not rows:
            return [], 0

        header_idx = max(header_row - 1, 0)
        if header_idx >= len(rows):
            return [], 0

        headers = [str(h).strip() or f"col_{i}" for i, h in enumerate(rows[header_idx])]
        data_rows = rows[header_idx + 1 :]
        data_rows = [r for r in data_rows if any(str(c).strip() for c in r)]

        columns: list[ColumnMetadata] = []
        for idx, name in enumerate(headers):
            field = name or f"col_{idx}"
            columns.append(
                ColumnMetadata(
                    id=f"c{idx}",
                    fieldName=field,
                    displayName=field,
                    dataType="string",
                    filterable=True,
                    sortable=True,
                )
            )
        return columns, len(data_rows)

    def _count_rows(self, file_path: Path, *, header_row: int) -> int:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            with file_path.open(encoding="utf-8-sig") as handle:
                return max(sum(1 for _ in handle) - header_row, 0)

        from openpyxl import load_workbook

        book = load_workbook(file_path, read_only=True, data_only=True)
        try:
            rows = book.active.max_row or 0
            return max(rows - header_row, 0)
        finally:
            book.close()
