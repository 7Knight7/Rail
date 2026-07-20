"""Build preview rows from processed manual-run artifacts."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


def read_preview_rows(
    *,
    excel_path: str | None,
    csv_path: str | None,
    visible_columns: list[str] | None,
    limit: int = 10,
    allow_csv_fallback: bool = True,
) -> list[dict[str, str | int | float]]:
    rows: list[dict[str, str | int | float]] = []
    if excel_path and Path(excel_path).is_file():
        rows = _read_excel_preview(excel_path, limit=limit)
    elif allow_csv_fallback and csv_path and Path(csv_path).is_file():
        rows = _read_csv_preview(csv_path, limit=limit)

    if not rows or not visible_columns:
        return rows

    filtered: list[dict[str, str | int | float]] = []
    for row in rows:
        projected = {col: row.get(col, "") for col in visible_columns if col in row}
        filtered.append(projected if projected else row)
    return filtered or rows


def _read_excel_preview(path: str, *, limit: int) -> list[dict[str, str | int | float]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if not first_row:
            return []

        first_cells = [str(cell).strip() if cell is not None else "" for cell in first_row]
        if len(first_cells) == 1 or any("Rail Madad Report" in cell for cell in first_cells):
            header_row = next(rows_iter, None)
        else:
            header_row = first_row

        if not header_row:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        headers = [header for header in headers if header]
        if not headers:
            return []

        preview: list[dict[str, str | int | float]] = []
        for values in rows_iter:
            if len(preview) >= limit:
                break
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            row: dict[str, str | int | float] = {}
            for idx, header in enumerate(headers):
                if idx >= len(values):
                    break
                value = values[idx]
                if value is None:
                    row[header] = ""
                elif isinstance(value, (int, float)):
                    row[header] = value
                else:
                    row[header] = str(value).strip()
            if row:
                preview.append(row)
        return preview
    finally:
        workbook.close()


def _read_csv_preview(path: str, *, limit: int) -> list[dict[str, str | int | float]]:
    preview: list[dict[str, str | int | float]] = []
    with Path(path).open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if len(preview) >= limit:
                break
            if not any(str(v).strip() for v in row.values()):
                continue
            preview.append({k: (v or "") for k, v in row.items()})
    return preview
