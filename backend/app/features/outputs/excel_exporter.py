"""Write processed datasets to Excel workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.features.processing.schemas import ProcessDatasetResponse


class ExcelExporter:
    """Export processed dataset rows to a final Excel file."""

    def write(self, processed: ProcessDatasetResponse, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"

        column_names = [column.name for column in processed.columns]
        sheet.append(column_names)

        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font

        highlight_by_row = self._highlight_index(processed.highlights)
        for row_index, row in enumerate(processed.rows, start=2):
            sheet.append([row.get(column, "") for column in column_names])
            highlight = highlight_by_row.get(row_index - 2)
            if highlight:
                fill = PatternFill(
                    start_color=self._normalize_color(highlight.get("backgroundColor", "#FEF3C7")),
                    end_color=self._normalize_color(highlight.get("backgroundColor", "#FEF3C7")),
                    fill_type="solid",
                )
                for cell in sheet[row_index]:
                    cell.fill = fill

        for column_index, column_name in enumerate(column_names, start=1):
            lengths = [len(str(column_name))]
            lengths.extend(
                len(str(row.get(column_name, "")))
                for row in processed.rows[:200]
            )
            max_length = max(lengths)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)

        workbook.save(output_path)
        workbook.close()
        return output_path

    @staticmethod
    def _highlight_index(highlights: list[dict]) -> dict[int, dict]:
        indexed: dict[int, dict] = {}
        for highlight in highlights:
            row_index = highlight.get("rowIndex")
            if row_index is None:
                continue
            indexed[int(row_index)] = highlight
        return indexed

    @staticmethod
    def _normalize_color(color: str) -> str:
        return color.lstrip("#").upper()[:6] or "FEF3C7"
