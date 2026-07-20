"""Report 5 (scr-train) process-only manual generation tests."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import AsyncMock, patch
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.automation.processing.column_config import projection_labels_for_slug
from app.automation.processing.report5_processor import Report5Processor
from app.domain.entities.user import User, UserRole
from app.features.auth.dependencies import require_officer_or_admin, validate_csrf_token
from app.features.reports.preview_projection import build_scr_preview_rows
from app.features.reports.scr_fresh import validate_manual_scr_column_snapshot
from app.features.reports.service import ManualReportService, build_config_snapshot
from app.features.reports.schemas import ManualGenerateRequest
from app.infrastructure.database.models import AutomationArtifactModel, AutomationRunModel, ReportDatasetModel
from app.main import app
from datetime import UTC, datetime

ACCEPTANCE_FILTER = [
    "scr-train.complaint_ref_no",
    "scr-train.created_on",
    "scr-train.user_id",
    "scr-train.comp_type_name",
    "scr-train.complaint_desc",
]

FIXTURES_R5 = Path(__file__).resolve().parents[2] / "fixtures" / "report5"


@pytest.fixture
def officer_user() -> User:
    now = datetime.now(UTC)
    return User(
        id="test-officer",
        username="officer",
        email="officer@test.local",
        password_hash="hash",
        role=UserRole.OFFICER,
        is_active=True,
        created_at=now,
        updated_at=now,
    )


@pytest.fixture
async def api_client(officer_user: User):
    async def override_user() -> User:
        return officer_user

    def override_csrf() -> None:
        return None

    app.dependency_overrides[require_officer_or_admin] = override_user
    app.dependency_overrides[validate_csrf_token] = override_csrf

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client

    app.dependency_overrides.clear()


async def _seed_scr_train_dataset(session: AsyncSession, csv_path: Path) -> None:
    session.add(
        ReportDatasetModel(
            id=str(uuid4()),
            report_id="scr-train",
            name="SCR Train",
            source_file_path=str(csv_path),
            row_count=10,
            checksum="test",
        )
    )
    await session.commit()


@pytest.mark.asyncio
async def test_other_slugs_still_use_cdp_manual_async(api_client: AsyncClient):
    with patch("app.features.reports.service.AutomationService") as mock_cls:
        mock_service = AsyncMock()
        mock_service.start_manual_async.return_value = ("run-r6", "running")
        mock_cls.return_value = mock_service

        response = await api_client.post(
            "/api/v1/reports/merging/generate",
            json={
                "selected_column_ids": ["report1.source_a.sno", "report1.source_a.organisation"],
                "column_order": ["report1.source_a.sno", "report1.source_a.organisation"],
                "export_format": "xlsx",
                "configuration_source": "manual_snapshot",
            },
        )

    assert response.status_code == 200
    assert response.json()["status"] == "Extracting"
    mock_service.start_manual_async.assert_awaited_once()


def test_build_config_snapshot_preserves_selected_ids_and_requested_formats():
    body = ManualGenerateRequest(
        selected_column_ids=ACCEPTANCE_FILTER,
        column_order=ACCEPTANCE_FILTER,
        requested_formats=["xlsx", "pdf"],
    )
    snapshot = build_config_snapshot(body, report_slug="scr-train")
    assert snapshot["selected_column_ids"] == ACCEPTANCE_FILTER
    assert snapshot["column_order"] == ACCEPTANCE_FILTER
    assert snapshot["requested_formats"] == ["xlsx", "pdf"]
    assert snapshot["configuration_source"] == "manual_snapshot"


@pytest.mark.asyncio
async def test_invalid_column_snapshot_raises():
    snapshot = {
        "selected_column_ids": ACCEPTANCE_FILTER + ["scr-train.not_a_real_column"],
        "column_order": ACCEPTANCE_FILTER + ["scr-train.not_a_real_column"],
        "configuration_source": "manual_snapshot",
    }
    with pytest.raises(ValueError, match="invalid="):
        await validate_manual_scr_column_snapshot("scr-train", snapshot)


def test_preview_and_processor_headers_match_for_acceptance_filter(tmp_path: Path):
    csv_path = tmp_path / "train.csv"
    csv_path.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )

    preview = build_scr_preview_rows(
        "scr-train",
        csv_path,
        selected_column_ids=ACCEPTANCE_FILTER,
        column_order=ACCEPTANCE_FILTER,
    )
    import app.automation.processing.report5_processor as r5_mod

    r5_mod.Report5Processor._find_template = lambda self: None  # type: ignore[method-assign]
    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": ACCEPTANCE_FILTER,
            "column_order": ACCEPTANCE_FILTER,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    assert preview["visible_columns"] == result.visible_columns
    assert preview["visible_columns"] == projection_labels_for_slug("scr-train", ACCEPTANCE_FILTER)


def test_filtered_processor_allows_complaint_mode_column(tmp_path: Path):
    """Regression: user-selected Complaint Mode must not trip REMOVED_OUTPUT_LABELS assert."""
    csv_path = tmp_path / "train.csv"
    csv_path.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    selected = ACCEPTANCE_FILTER + ["scr-train.complaint_mode"]
    import app.automation.processing.report5_processor as r5_mod

    r5_mod.Report5Processor._find_template = lambda self: None  # type: ignore[method-assign]
    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": selected,
            "column_order": selected,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    assert "Complaint Mode" in (result.visible_columns or [])


def test_filtered_excel_pdf_contain_exact_selected_columns(tmp_path: Path):
    csv_path = tmp_path / "train.csv"
    csv_path.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    import app.automation.processing.report5_processor as r5_mod

    r5_mod.Report5Processor._find_template = lambda self: None  # type: ignore[method-assign]
    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": ACCEPTANCE_FILTER,
            "column_order": ACCEPTANCE_FILTER,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    assert result.excel_path and result.pdf_path

    workbook = load_workbook(result.excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(c).strip() if c else "" for c in rows[1]]
        assert headers == result.visible_columns
        assert len(rows) - 2 == result.processed_row_count
    finally:
        workbook.close()

    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"


def test_sno_sequential_when_selected(tmp_path: Path):
    csv_path = tmp_path / "train.csv"
    csv_path.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    selected = ["scr-train.sno", "scr-train.complaint_ref_no"]
    import app.automation.processing.report5_processor as r5_mod

    r5_mod.Report5Processor._find_template = lambda self: None  # type: ignore[method-assign]
    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": selected,
            "column_order": selected,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    workbook = load_workbook(result.excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        data_rows = list(sheet.iter_rows(min_row=3, values_only=True))
        serials = [int(row[0]) for row in data_rows if row[0] is not None]
        assert serials == list(range(1, len(serials) + 1))
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_get_run_status_scr_train_dual_artifacts_expose_pdf_urls(
    test_session: AsyncSession,
    tmp_path: Path,
):
    """Status poll must return separate PDF preview/download URLs for scr-train."""
    from openpyxl import Workbook

    run_id = str(uuid4())
    slug = "scr-train"
    excel_path = tmp_path / "out.xlsx"
    pdf_path = tmp_path / "out.pdf"
    wb = Workbook()
    wb.active["A1"] = "Header"
    wb.save(excel_path)
    pdf_path.write_bytes(b"%PDF-1.4\n")

    metadata = json.dumps(
        {
            "selected_column_ids": ACCEPTANCE_FILTER,
            "column_order": ACCEPTANCE_FILTER,
            "configuration_source": "manual_snapshot",
        }
    )
    manual_config = {
        "report_slug": slug,
        "generation_mode": "process_only",
        "selected_column_ids": ACCEPTANCE_FILTER,
        "column_order": ACCEPTANCE_FILTER,
        "configuration_source": "manual_snapshot",
    }
    report_payload = {
        "manual_config": manual_config,
        "result": {
            "success": True,
            "connected": True,
            "tab_found": True,
            "reports_successful": 1,
            "reports_failed": 0,
            "reports": [
                {
                    "slug": slug,
                    "status": "success",
                    "processing_success": True,
                    "processing_attempted": True,
                    "ingestion_success": True,
                    "source_row_count": 10,
                    "processed_row_count": 10,
                    "source_csv_path": str(tmp_path / "source.csv"),
                    "excel_path": str(excel_path),
                    "pdf_path": str(pdf_path),
                    "visible_columns": projection_labels_for_slug(slug, ACCEPTANCE_FILTER),
                }
            ],
        },
    }
    run = AutomationRunModel(
        id=run_id,
        profile_id="test-profile",
        status="completed",
        trigger_type="manual_report",
        result_json=json.dumps(report_payload),
        completed_at=datetime.now(UTC),
    )
    test_session.add(run)
    await test_session.commit()

    excel_id = str(uuid4())
    pdf_id = str(uuid4())
    test_session.add_all(
        [
            AutomationArtifactModel(
                id=excel_id,
                run_id=run_id,
                artifact_type="excel",
                file_path=str(excel_path),
                file_size_bytes=excel_path.stat().st_size,
                report_slug=slug,
                report_name=slug,
                status="ready",
                metadata_json=metadata,
            ),
            AutomationArtifactModel(
                id=pdf_id,
                run_id=run_id,
                artifact_type="pdf",
                file_path=str(pdf_path),
                file_size_bytes=pdf_path.stat().st_size,
                report_slug=slug,
                report_name=slug,
                status="ready",
                metadata_json=metadata,
            ),
        ]
    )
    await test_session.commit()

    status = await ManualReportService().get_run_status(test_session, run_id, expected_slug=slug)

    assert status.status == "Completed"
    assert status.excel_artifact_id == excel_id
    assert status.pdf_artifact_id == pdf_id
    assert status.excel_download_url == f"/api/v1/automation/artifacts/{excel_id}/download"
    assert status.pdf_download_url == f"/api/v1/automation/artifacts/{pdf_id}/download"
    assert status.pdf_preview_url == f"/api/v1/automation/artifacts/{pdf_id}/preview"
    assert status.excel_filename.endswith(".xlsx") or status.excel_filename == "out.xlsx"
    assert status.pdf_filename == "out.pdf"


@pytest.mark.asyncio
async def test_fresh_extraction_running_status_maps_to_extracting(test_session: AsyncSession):
    run_id = str(uuid4())
    manual_config = {
        "report_slug": "scr-train",
        "generation_mode": "fresh_extraction",
        "force_fresh_extraction": True,
        "selected_column_ids": ACCEPTANCE_FILTER,
        "column_order": ACCEPTANCE_FILTER,
        "configuration_source": "manual_snapshot",
    }
    run = AutomationRunModel(
        id=run_id,
        profile_id="test-profile",
        status="running",
        trigger_type="manual_report",
        result_json=json.dumps({"manual_config": manual_config}),
    )
    test_session.add(run)
    await test_session.commit()

    status = await ManualReportService().get_run_status(
        test_session, run_id, expected_slug="scr-train"
    )
    assert status.status == "Extracting"
