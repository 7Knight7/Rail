"""Status polling and filtered export for SCR Report 5/6."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.automation.config import config
from app.automation.processing.column_config import project_scr_for_output, projection_labels_for_slug
from app.automation.processing.report5_processor import Report5Processor
from app.automation.processing.report6_processor import Report6Processor
from app.features.reports.preview_projection import build_scr_preview_rows
from app.features.reports.service import ManualReportService
from app.infrastructure.database.models import AutomationArtifactModel, AutomationRunModel

R5_FILTER = [
    "scr-train.complaint_ref_no",
    "scr-train.created_on",
    "scr-train.user_id",
    "scr-train.complaint_desc",
]
R6_FILTER = [
    "scr-station.complaint_ref_no",
    "scr-station.train_station",
    "scr-station.remarks",
    "scr-station.user_id",
]


def _scr_manual_run(
    *,
    slug: str,
    run_id: str,
    selected_ids: list[str],
    processing_success: bool = False,
    error: str = "PDF_GENERATION_FAILED: test",
    visible_columns: list[str] | None = None,
) -> AutomationRunModel:
    manual_config = {
        "report_slug": slug,
        "export_format": "xlsx",
        "selected_column_ids": selected_ids,
        "column_order": selected_ids,
        "configuration_source": "manual_snapshot",
        "report_date": "18-07-2026",
    }
    report = {
        "slug": slug,
        "status": "partial_success" if not processing_success else "success",
        "processing_success": processing_success,
        "processing_attempted": True,
        "ingestion_success": True,
        "source_row_count": 10,
        "source_csv_path": "/tmp/source.csv",
        "processed_row_count": 10 if processing_success else 0,
        "error": error if not processing_success else None,
        "visible_columns": visible_columns or [],
    }
    payload = {
        "manual_config": manual_config,
        "result": {
            "success": processing_success,
            "connected": True,
            "tab_found": True,
            "reports_successful": 1 if processing_success else 0,
            "reports_failed": 0 if processing_success else 1,
            "reports": [report],
        },
    }
    return AutomationRunModel(
        id=run_id,
        profile_id="test-profile",
        status="completed",
        trigger_type="manual_report",
        result_json=json.dumps(payload),
        completed_at=datetime.now(UTC),
    )


@pytest.mark.asyncio
async def test_get_run_status_scr_failed_run_returns_real_error_not_500(
    test_session: AsyncSession,
):
    """Regression: missing visible_columns must not crash status poll with NameError."""
    run_id = str(uuid4())
    run = _scr_manual_run(
        slug="scr-station",
        run_id=run_id,
        selected_ids=R6_FILTER,
        error="PDF_GENERATION_FAILED: UNAPPROVED_OUTPUT_COLUMN example",
    )
    test_session.add(run)
    await test_session.commit()

    status = await ManualReportService().get_run_status(test_session, run_id, expected_slug="scr-station")

    assert status.status == "Failed"
    assert "PDF_GENERATION_FAILED" in (status.error or "")
    assert status.visible_columns == projection_labels_for_slug("scr-station", R6_FILTER)


@pytest.mark.asyncio
async def test_get_run_status_scr_dual_artifacts_when_processing_succeeds(
    test_session: AsyncSession,
    tmp_path: Path,
):
    run_id = str(uuid4())
    slug = "scr-train"
    excel_path = tmp_path / "out.xlsx"
    pdf_path = tmp_path / "out.pdf"
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "Header"
    wb.save(excel_path)
    pdf_path.write_bytes(b"%PDF-1.4\n")

    metadata = json.dumps(
        {
            "selected_column_ids": R5_FILTER,
            "column_order": R5_FILTER,
        }
    )
    run = _scr_manual_run(
        slug=slug,
        run_id=run_id,
        selected_ids=R5_FILTER,
        processing_success=True,
        visible_columns=projection_labels_for_slug(slug, R5_FILTER),
    )
    report_payload = json.loads(run.result_json)
    report_payload["result"]["reports"][0]["excel_path"] = str(excel_path)
    report_payload["result"]["reports"][0]["pdf_path"] = str(pdf_path)
    report_payload["result"]["reports"][0]["processing_success"] = True
    run.result_json = json.dumps(report_payload)
    test_session.add(run)
    await test_session.commit()

    metadata = json.dumps(
        {"selected_column_ids": R5_FILTER, "column_order": R5_FILTER}
    )
    excel_id = str(uuid4())
    pdf_id = str(uuid4())
    test_session.add_all(
        [
            AutomationArtifactModel(
                id=excel_id,
                run_id=run_id,
                artifact_type="excel",
                file_path=str(excel_path),
                file_size_bytes=excel_path.stat().st_size,
                report_slug=slug,
                report_name=slug,
                status="ready",
                metadata_json=metadata,
            ),
            AutomationArtifactModel(
                id=pdf_id,
                run_id=run_id,
                artifact_type="pdf",
                file_path=str(pdf_path),
                file_size_bytes=pdf_path.stat().st_size,
                report_slug=slug,
                report_name=slug,
                status="ready",
                metadata_json=metadata,
            ),
        ]
    )
    await test_session.commit()

    status = await ManualReportService().get_run_status(test_session, run_id, expected_slug=slug)

    assert status.status == "Completed"
    assert status.excel_artifact_id == excel_id
    assert status.pdf_artifact_id == pdf_id
    assert status.excel_download_url
    assert status.pdf_download_url
    assert status.pdf_preview_url
    assert status.visible_columns == projection_labels_for_slug(slug, R5_FILTER)


def test_preview_and_processor_share_projection(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    fixtures = Path(__file__).resolve().parents[2] / "fixtures" / "report5"
    csv_path = tmp_path / "scr-train_complaints_raw.csv"
    csv_path.write_text(
        (fixtures / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )

    preview = build_scr_preview_rows(
        "scr-train",
        csv_path,
        selected_column_ids=R5_FILTER,
        column_order=R5_FILTER,
    )
    processor = Report5Processor()
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    result = processor.process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": R5_FILTER,
            "column_order": R5_FILTER,
            "configuration_source": "manual_snapshot",
            "report_slug": "scr-train",
        },
    )
    assert result.success, result.error
    assert preview["visible_columns"] == result.visible_columns
    assert preview["selected_column_ids"] == R5_FILTER


def test_filtered_report5_excel_pdf_headers(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    fixtures = Path(__file__).resolve().parents[2] / "fixtures" / "report5"
    csv_path = tmp_path / "scr-train_complaints_raw.csv"
    csv_path.write_text(
        (fixtures / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))

    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection={
            "selected_column_ids": R5_FILTER,
            "column_order": R5_FILTER,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    assert result.visible_columns == [
        "Complaint Ref Number",
        "Created On",
        "User ID",
        "Complaint Description",
    ]
    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert headers == result.visible_columns
    assert Path(result.pdf_path).read_bytes()[:5] == b"%PDF-"


def test_filtered_report6_excel_pdf_headers(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    fixtures = Path(__file__).resolve().parents[2] / "fixtures" / "report6"
    csv_path = tmp_path / "scr-station_complaints_raw.csv"
    csv_path.write_text(
        (fixtures / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))

    result = Report6Processor().process(
        source_a_path=csv_path,
        report_slug="scr-station",
        column_selection={
            "selected_column_ids": R6_FILTER,
            "column_order": R6_FILTER,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success, result.error
    assert result.visible_columns == [
        "Complaint Ref Number",
        "Train/Station",
        "Remarks",
        "User ID",
    ]
    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert headers == result.visible_columns
    assert Path(result.pdf_path).read_bytes()[:5] == b"%PDF-"


def test_unavailable_scr_column_raises_select_error():
    rows = [{"complaintRefNo": "1", "userId": "u1"}]
    with pytest.raises(ValueError, match="SELECTED_COLUMN_UNAVAILABLE"):
        project_scr_for_output(
            "scr-train",
            rows,
            selected_keys=["scr-train.complaint_ref_no", "scr-train.created_on"],
        )
