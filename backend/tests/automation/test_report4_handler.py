"""Report 4 handler tests: refresh wait, retry, status, run isolation."""

from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.automation.generator import ReportGenerationError
from app.automation.handlers.report4_handler import Report4Handler, type_slug
from app.automation.processing.report4_processor import Report4Processor
from app.automation.report4_filters import TypeConfig, get_type_configs
from app.automation.reports import REPORT_4_TYPES
from app.automation.schemas import ReportResult


@pytest.fixture
def handler() -> Report4Handler:
    return Report4Handler()


def _configs(*names: str) -> list[TypeConfig]:
    all_cfgs = {c.name: c for c in get_type_configs()}
    return [all_cfgs[n] for n in names]


@pytest.mark.asyncio
async def test_stale_old_table_not_accepted(handler: Report4Handler):
    """Same fingerprint after submit must not count as refresh success."""
    report_root = MagicMock()
    page = MagicMock()

    handler._wait_for_loaders = AsyncMock()
    handler._table_fingerprint = AsyncMock(return_value="OLD##5##row1")

    ok = await handler._wait_for_table_refresh(
        report_root,
        page,
        "OLD##5##row1",
        type_name="Security",
        timeout_ms=400,
    )
    assert ok is False


@pytest.mark.asyncio
async def test_refresh_wait_succeeds_after_delayed_replacement(handler: Report4Handler):
    """Fingerprint change after a brief stale period succeeds."""
    report_root = MagicMock()
    page = MagicMock()
    handler._wait_for_loaders = AsyncMock()

    fingerprints = [
        "OLD##5##row1",  # still stale
        "OLD##5##row1",
        "",  # cleared
        "NEW##5##row2",
        "NEW##5##row2",  # confirm
    ]

    async def fp_side_effect(*_a, **_k):
        if fingerprints:
            return fingerprints.pop(0)
        return "NEW##5##row2"

    handler._table_fingerprint = AsyncMock(side_effect=fp_side_effect)

    ok = await handler._wait_for_table_refresh(
        report_root,
        page,
        "OLD##5##row1",
        type_name="Security",
        timeout_ms=5_000,
    )
    assert ok is True


@pytest.mark.asyncio
async def test_failed_first_attempt_retries_with_full_filters(
    handler: Report4Handler,
    tmp_path: Path,
):
    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._wait_for_received_header = AsyncMock()
    handler._sort_received = AsyncMock()
    handler._save_type_failure_artifacts = AsyncMock()

    csv_path = tmp_path / "report4_security_raw.csv"
    csv_path.write_text("h\n1\n", encoding="utf-8")
    handler._extract_type = AsyncMock(return_value=(csv_path, 1))

    calls = {"n": 0}

    async def submit_side_effect(page, session, report, type_config, *, attempt):
        calls["n"] += 1
        if attempt == 1:
            raise ReportGenerationError("Report types did not display after generate")
        return MagicMock()

    handler._submit_type_once = AsyncMock(side_effect=submit_side_effect)

    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()
    cfg = _configs("Security")[0]

    with patch("app.automation.handlers.report4_handler.asyncio.sleep", new=AsyncMock()):
        outcome = await handler._run_type_with_retry(
            page, session, REPORT_4_TYPES, cfg, tmp_path
        )

    assert outcome["status"] == "success"
    assert handler._submit_type_once.await_count == 2
    assert handler.navigation.navigate_to_report.await_count == 1  # retry reopen


@pytest.mark.asyncio
async def test_one_failed_type_does_not_abort_remaining(
    handler: Report4Handler,
    tmp_path: Path,
):
    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._wait_for_received_header = AsyncMock()
    handler._sort_received = AsyncMock()
    handler._save_type_failure_artifacts = AsyncMock()

    async def extract_side_effect(report_root, report, type_config, extracted_dir):
        path = extracted_dir / f"report4_{type_slug(type_config.name)}_raw.csv"
        path.write_text("Train No.,Received\n1,10\n", encoding="utf-8")
        return path, 1

    handler._extract_type = AsyncMock(side_effect=extract_side_effect)

    async def submit_side_effect(page, session, report, type_config, *, attempt):
        if type_config.name == "Bedroll":
            raise ReportGenerationError("Report types did not display after generate")
        return MagicMock()

    handler._submit_type_once = AsyncMock(side_effect=submit_side_effect)
    handler.finalize_after_extract = AsyncMock(
        return_value=ReportResult(
            slug="types",
            dataset_key="types",
            status="success",
            ingestion_success=True,
            processing_success=True,
            excel_path=str(tmp_path / "out.xlsx"),
            pdf_path=str(tmp_path / "out.pdf"),
            excel_download_url="/excel",
            pdf_preview_url="/preview",
            pdf_download_url="/pdf",
            source_paths=[],
            row_counts={},
        )
    )

    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()
    configs = _configs("Security", "Bedroll", "Water Availability")

    with (
        patch(
            "app.automation.handlers.report4_handler.get_type_configs",
            return_value=configs,
        ),
        patch("app.automation.handlers.report4_handler.asyncio.sleep", new=AsyncMock()),
        patch(
            "app.automation.handlers.report4_handler.get_run_context",
            return_value=None,
        ),
        patch(
            "app.automation.handlers.report4_handler.resolve_report_dir",
            return_value=tmp_path / "types",
        ),
        patch(
            "app.automation.handlers.report4_handler.config"
        ) as cfg,
    ):
        cfg.extracted_data_dir = str(tmp_path / "extracted")
        cfg.screenshots_dir = str(tmp_path / "shots")
        result = await handler.execute(page, session, REPORT_4_TYPES)

    assert result.status == "partial_success"
    assert "Bedroll" in (result.error or "")
    # Security + Water extracted; Bedroll skipped after 2 attempts
    assert handler._extract_type.await_count == 2
    handler.finalize_after_extract.assert_awaited_once()


@pytest.mark.asyncio
async def test_all_seven_success_status_success(
    handler: Report4Handler,
    tmp_path: Path,
):
    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._wait_for_received_header = AsyncMock()
    handler._sort_received = AsyncMock()
    handler._submit_type_once = AsyncMock(return_value=MagicMock())

    async def extract_side_effect(report_root, report, type_config, extracted_dir):
        path = extracted_dir / f"report4_{type_slug(type_config.name)}_raw.csv"
        path.write_text("Train No.,Received\n1,10\n", encoding="utf-8")
        return path, 1

    handler._extract_type = AsyncMock(side_effect=extract_side_effect)
    handler.finalize_after_extract = AsyncMock(
        return_value=ReportResult(
            slug="types",
            dataset_key="types",
            status="success",
            ingestion_success=True,
            processing_success=True,
            excel_path=str(tmp_path / "out.xlsx"),
            pdf_path=str(tmp_path / "out.pdf"),
            excel_download_url="/excel",
            pdf_preview_url="/preview",
            pdf_download_url="/pdf",
        )
    )

    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()

    with (
        patch(
            "app.automation.handlers.report4_handler.get_run_context",
            return_value=None,
        ),
        patch(
            "app.automation.handlers.report4_handler.resolve_report_dir",
            return_value=tmp_path / "types",
        ),
        patch("app.automation.handlers.report4_handler.config") as cfg,
    ):
        cfg.extracted_data_dir = str(tmp_path / "extracted")
        result = await handler.execute(page, session, REPORT_4_TYPES)

    assert result.status == "success"
    assert handler._extract_type.await_count == 7


@pytest.mark.asyncio
async def test_no_success_status_failed(handler: Report4Handler, tmp_path: Path):
    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._save_type_failure_artifacts = AsyncMock()
    handler._submit_type_once = AsyncMock(
        side_effect=ReportGenerationError("Report types did not display after generate")
    )
    handler.finalize_after_extract = AsyncMock()

    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()
    configs = _configs("Security", "Bedroll")

    with (
        patch(
            "app.automation.handlers.report4_handler.get_type_configs",
            return_value=configs,
        ),
        patch("app.automation.handlers.report4_handler.asyncio.sleep", new=AsyncMock()),
        patch(
            "app.automation.handlers.report4_handler.get_run_context",
            return_value=None,
        ),
        patch(
            "app.automation.handlers.report4_handler.resolve_report_dir",
            return_value=tmp_path / "types",
        ),
        patch("app.automation.handlers.report4_handler.config") as cfg,
    ):
        cfg.extracted_data_dir = str(tmp_path / "extracted")
        cfg.screenshots_dir = str(tmp_path / "shots")
        result = await handler.execute(page, session, REPORT_4_TYPES)

    assert result.status == "failed"
    handler.finalize_after_extract.assert_not_awaited()


@pytest.mark.asyncio
async def test_current_run_index_excludes_stale_and_no_nested_path(
    handler: Report4Handler,
    tmp_path: Path,
):
    stale_dir = tmp_path / "types"
    stale_dir.mkdir(parents=True)
    (stale_dir / "report4_security_raw.csv").write_text("stale\n", encoding="utf-8")
    (stale_dir / "types").mkdir()  # nested types/types trap

    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._wait_for_received_header = AsyncMock()
    handler._sort_received = AsyncMock()
    handler._submit_type_once = AsyncMock(return_value=MagicMock())

    async def extract_side_effect(report_root, report, type_config, extracted_dir):
        # Nested path must not be created by handler extract dir
        assert extracted_dir.name != "types" or extracted_dir.parent.name != "types"
        assert "types" in extracted_dir.parts
        path = extracted_dir / f"report4_{type_slug(type_config.name)}_raw.csv"
        path.write_text("Train No.,Received\n1,9\n", encoding="utf-8")
        return path, 1

    handler._extract_type = AsyncMock(side_effect=extract_side_effect)

    captured: dict = {}

    async def finalize_side_effect(**kwargs):
        captured["csv_path"] = kwargs["csv_path"]
        captured["source_paths"] = kwargs["source_paths"]
        return ReportResult(
            slug="types",
            dataset_key="types",
            status="success",
            ingestion_success=True,
            processing_success=True,
            excel_path=str(tmp_path / "a.xlsx"),
            pdf_path=str(tmp_path / "a.pdf"),
            excel_download_url="/e",
            pdf_preview_url="/p",
            pdf_download_url="/d",
        )

    handler.finalize_after_extract = AsyncMock(side_effect=finalize_side_effect)

    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()
    configs = _configs("Security")

    run_ctx = MagicMock()
    run_ctx.run_id = "run-abc"
    run_ctx.defer_processing = True
    run_ctx.merge_result = MagicMock()
    run_ctx.timing = MagicMock()
    run_ctx.timing.spans = {}

    with (
        patch(
            "app.automation.handlers.report4_handler.get_type_configs",
            return_value=configs,
        ),
        patch(
            "app.automation.handlers.report4_handler.get_run_context",
            return_value=run_ctx,
        ),
        patch(
            "app.automation.handlers.report4_handler.resolve_report_dir",
            return_value=tmp_path / "types",
        ),
        patch("app.automation.handlers.report4_handler.config") as cfg,
    ):
        cfg.extracted_data_dir = str(tmp_path / "extracted")
        result = await handler.execute(page, session, REPORT_4_TYPES)

    index_path = Path(captured["csv_path"])
    assert index_path.name == "types_combined_index.csv"
    assert index_path.parent.name == "run-abc"
    assert "types/types" not in str(index_path).replace("\\", "/")
    assert str(stale_dir / "report4_security_raw.csv") not in captured["source_paths"]

    with index_path.open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 1
    assert rows[0]["status"] == "success"
    assert "run-abc" in rows[0]["csv_path"].replace("\\", "/")
    assert result.status == "success"


@pytest.mark.asyncio
async def test_ingestion_processor_and_artifact_urls(
    handler: Report4Handler,
    tmp_path: Path,
):
    handler.navigation = MagicMock()
    handler.navigation.navigate_to_report = AsyncMock()
    handler.ensure_mis_page = AsyncMock(side_effect=lambda page, session, ctx="": page)
    handler._wait_for_received_header = AsyncMock()
    handler._sort_received = AsyncMock()
    handler._submit_type_once = AsyncMock(return_value=MagicMock())

    async def extract_side_effect(report_root, report, type_config, extracted_dir):
        path = extracted_dir / f"report4_{type_slug(type_config.name)}_raw.csv"
        path.write_text("Train No.,Received\n1,10\n", encoding="utf-8")
        return path, 1

    handler._extract_type = AsyncMock(side_effect=extract_side_effect)

    excel = tmp_path / "out.xlsx"
    pdf = tmp_path / "out.pdf"
    excel.write_bytes(b"xlsx")
    pdf.write_bytes(b"pdf")

    async def finalize(**kwargs):
        assert kwargs["csv_path"].name == "types_combined_index.csv"
        return ReportResult(
            slug="types",
            dataset_key="types",
            status="success",
            ingestion_success=True,
            processing_success=True,
            excel_path=str(excel),
            pdf_path=str(pdf),
            excel_download_url="/api/v1/automation/artifacts/ex/download",
            pdf_preview_url="/api/v1/automation/artifacts/pd/preview",
            pdf_download_url="/api/v1/automation/artifacts/pd/download",
            source_paths=kwargs.get("source_paths") or [],
            row_counts=kwargs.get("row_counts") or {},
        )

    handler.finalize_after_extract = AsyncMock(side_effect=finalize)
    page = MagicMock()
    page.wait_for_selector = AsyncMock()
    session = MagicMock()
    configs = _configs("Security", "Coach Cleanliness")

    with (
        patch(
            "app.automation.handlers.report4_handler.get_type_configs",
            return_value=configs,
        ),
        patch(
            "app.automation.handlers.report4_handler.get_run_context",
            return_value=None,
        ),
        patch(
            "app.automation.handlers.report4_handler.resolve_report_dir",
            return_value=tmp_path / "types",
        ),
        patch("app.automation.handlers.report4_handler.config") as cfg,
    ):
        cfg.extracted_data_dir = str(tmp_path / "extracted")
        result = await handler.execute(page, session, REPORT_4_TYPES)

    handler.finalize_after_extract.assert_awaited_once()
    assert result.excel_download_url
    assert result.pdf_preview_url
    assert result.pdf_download_url


def test_processor_reads_index_paths_only(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    run_dir = tmp_path / "types" / "run-1"
    run_dir.mkdir(parents=True)
    stale = tmp_path / "types" / "report4_security_raw.csv"
    stale.write_text(
        "Train No.,Train Name,Owning Zone,Owning Division,Received\n"
        "999,Stale,Zone,Div,1\n",
        encoding="utf-8",
    )
    good = run_dir / "report4_security_raw.csv"
    good.write_text(
        "Train No.,Train Name,Owning Zone,Owning Division,Received\n"
        "123,GoodTrain,South Central Railway,SC,50\n",
        encoding="utf-8",
    )
    index = run_dir / "types_combined_index.csv"
    with index.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["type_name", "csv_path", "row_count", "status", "error"])
        writer.writerow(["Security", str(good), 1, "success", ""])
        writer.writerow(["Bedroll", "", 0, "failed", "boom"])

    result = Report4Processor().process(source_a_path=index, report_slug="types")
    assert result.success is True
    assert result.excel_path and Path(result.excel_path).exists()
    assert result.pdf_path and Path(result.pdf_path).exists()

    from openpyxl import load_workbook

    wb = load_workbook(result.excel_path)
    ws = wb.active
    values = [
        ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    ]
    assert "GoodTrain" in values
    assert "Stale" not in values
