"""Report 5 final rendering: field mapping, Unicode, PDF layout."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Spacer
from reportlab.platypus import Paragraph

from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_title_style
from app.automation.formatting.pdf_table import (
    REPORT5_MIN_FONT_SIZE,
    allocate_tier_column_widths,
    build_wrapped_fitted_table,
)
from app.automation.formatting.text_safe import contains_rendering_risk_markers, normalize_report_text
from app.automation.processing.column_config import project_scr_for_output
from app.automation.processing.report5_processor import Report5Processor
from app.automation.scr_field_map import canonicalize_scr_row, portal_header_to_canonical


def _portal_row(**overrides: str) -> dict[str, str]:
    base = {
        "Ref. No.": "2026071910989",
        "Mode": "T",
        "Registration Date": "19-07-26 21:24",
        "Closing Date": "19-07-26 21:51",
        "Train/Station": "12798",
        "Type": "Coach - Cleanliness",
        "Sub Type": "Cockroach / Rodents",
        "Dept": "CNW",
        "Status": "Closed",
        "Disposal Time": "0:27",
        "Zone": "SC",
        "Div": "HYB",
        "feedbackRemark": "My issue was not addressed",
        "trainNameForReport/Station Name": "CTO-KCG VENKATADRI EXP [SUPERFAST]",
        "complaintDesc": (
            "AI Generated Complaint Description There are rats. "
            "Complaint User Input కోచ్ లో ఎలకలు ఉన్నాయి"
        ),
        "remarks": "Inconvenience caused is deeply regretted.",
        "userId": "cnw_sc_hyb",
    }
    base.update(overrides)
    return base


def test_portal_header_aliases_for_report5_status_fields():
    assert portal_header_to_canonical("Disposal Time") == "diff"
    assert portal_header_to_canonical("Avg. Diff") == "diff"
    assert portal_header_to_canonical("Status") == "status"


def test_canonicalize_scr_row_maps_disposal_time_to_diff():
    canonical = canonicalize_scr_row(_portal_row())
    assert canonical["diff"] == "0:27"
    assert canonical["status"] == "Closed"
    assert canonical.get("finalStatus", "") == ""


def test_report5_projection_maps_final_status_and_avg_diff():
    rows = [canonicalize_scr_row(_portal_row())]
    headers, out_rows, _, keys, _ = project_scr_for_output(
        "scr-train",
        rows,
        selected_keys=[
            "scr-train.complaint_ref_no",
            "scr-train.final_status",
            "scr-train.avg_cliff",
            "scr-train.complaint_desc",
        ],
    )
    assert headers == [
        "Complaint Ref Number",
        "Final Status",
        "Avg. Diff",
        "Complaint Description",
    ]
    assert out_rows[0][1] == "Closed"
    assert out_rows[0][2] == "0:27"
    assert "ఎలకలు" in out_rows[0][3]
    assert keys[-1] == "scr-train.complaint_desc"


def test_missing_selected_avg_diff_source_raises():
    rows = [{"complaintRefNo": "1", "status": "Closed", "userId": "u1"}]
    with pytest.raises(ValueError, match="SELECTED_COLUMN_UNAVAILABLE"):
        project_scr_for_output(
            "scr-train",
            rows,
            selected_keys=["scr-train.complaint_ref_no", "scr-train.avg_cliff"],
        )


def test_unicode_preserved_and_risk_markers_removed():
    raw = "కోచ్\u25a0\u25a0 లో ఎలకలు"
    cleaned = normalize_report_text(raw, field_kind="text", column_name="Complaint Description")
    assert "ఎలకలు" in cleaned
    assert contains_rendering_risk_markers(cleaned) is False


def test_report5_pdf_layout_uses_a3_and_readable_header_widths():
    headers = [
        "S.No.",
        "Complaint Ref Number",
        "Final Status",
        "Avg. Diff",
        "Feedback Remark",
        "Complaint Description",
        "Remarks",
        "Train Name For Report",
    ]
    row = [
        "1",
        "2026071910989",
        "Closed",
        "0:27",
        "Long feedback remark text",
        "Telugu కోచ్ లో ఎలకలు and English details",
        "Official remarks follow up",
        "CTO-KCG VENKATADRI EXP [SUPERFAST]",
    ]
    table_data = [headers, row]
    table, pagesize, margin = build_wrapped_fitted_table(
        table_data,
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
        min_font_size=REPORT5_MIN_FONT_SIZE,
        max_font_size=REPORT5_MIN_FONT_SIZE,
        landscape_start_index=1,
    )
    assert pagesize == landscape(A3)
    usable = pagesize[0] - (2 * margin)
    wrapped_w, _ = table.wrap(usable, pagesize[1])
    assert wrapped_w <= usable + 1.0

    widths = allocate_tier_column_widths(headers, usable, font_size=REPORT5_MIN_FONT_SIZE)
    final_status_idx = headers.index("Final Status")
    assert widths[final_status_idx] >= widths[headers.index("S.No.")]


def test_report5_processor_xlsx_pdf_match_projected_values(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "scr-train"
    extracted.mkdir(parents=True)
    csv_path = extracted / "scr-train_complaints_raw.csv"
    csv_path.write_text(
        "complaintRefNo,createdOn,status,diff,complaintMode,trainStation,complaintTypeName,"
        "subTypeName,deptCode,zoneCode,divCode,feedbackRemark,trainNameForReport,"
        "complaintDesc,remarks,userId\n"
        "2026071910989,19-07-26 21:24,Closed,0:27,T,12798,Coach - Cleanliness,Cockroach / Rodents,"
        "CNW,SC,HYB,My issue,CTO-KCG EXP,"
        "AI text Complaint User Input కోచ్ లో ఎలకలు,Regretted,cnw_sc_hyb\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)

    selection = {
        "selected_column_ids": [
            "scr-train.sno",
            "scr-train.complaint_ref_no",
            "scr-train.final_status",
            "scr-train.avg_cliff",
            "scr-train.complaint_desc",
        ],
        "column_order": [
            "scr-train.sno",
            "scr-train.complaint_ref_no",
            "scr-train.final_status",
            "scr-train.avg_cliff",
            "scr-train.complaint_desc",
        ],
        "configuration_source": "manual_snapshot",
    }

    result = Report5Processor().process(
        source_a_path=csv_path,
        report_slug="scr-train",
        column_selection=selection,
    )
    assert result.success is True, result.error

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    values = [str(ws.cell(row=3, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert values[headers.index("Final Status")] == "Closed"
    assert values[headers.index("Avg. Diff")] == "0:27"
    assert "ఎలకలు" in values[headers.index("Complaint Description")]

    pdf_path = Path(result.pdf_path or "")
    assert pdf_path.is_file()
    pdf_bytes = pdf_path.read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"
    if ensure_pdf_unicode_fonts():
        assert (
            b"/Subtype /Type0" in pdf_bytes
            or b"+0 " in pdf_bytes
            or b"Nirmala" in pdf_bytes
            or b"RailReport" in pdf_bytes
        )
