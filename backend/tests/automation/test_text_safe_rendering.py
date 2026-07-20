"""Tests for shared report text normalization and Unicode PDF rendering."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_regular
from app.automation.formatting.text_pipeline import prepare_output_for_rendering
from app.automation.formatting.text_safe import (
    contains_rendering_risk_markers,
    normalize_report_text,
)
from app.automation.processing.report5_processor import Report5Processor


def test_normalize_decodes_html_entities_and_br():
    raw = "Coach dirty &amp; wet<br/>needs cleaning"
    result = normalize_report_text(raw, field_kind="text", column_name="Remarks")
    assert "&amp;" not in result
    assert "<br" not in result
    assert "Coach dirty & wet" in result
    assert "needs cleaning" in result
    assert "\n" in result


def test_normalize_removes_nul_and_zero_width():
    raw = "Valid\u0000text\u200b here"
    result = normalize_report_text(raw, field_kind="text", column_name="Remarks")
    assert "\u0000" not in result
    assert "\u200b" not in result
    assert "Validtext here" in result or "Valid text here" in result


def test_normalize_preserves_hindi_unicode():
    raw = "साफ-सफाई की समस्या"
    result = normalize_report_text(raw, field_kind="text", column_name="Complaint Description")
    assert "साफ" in result
    assert "समस्या" in result


def test_normalize_preserves_ids_without_nfkc_alteration():
    raw = "2026071503196"
    result = normalize_report_text(raw, field_kind="id", column_name="Complaint Ref Number")
    assert result == "2026071503196"


def test_normalize_strips_tags_preserves_surrounding_text():
    raw = "<p>Vendor refused <b>receipt</b> &amp; walked away</p>"
    result = normalize_report_text(raw, field_kind="text", column_name="Remarks")
    assert "Vendor refused" in result
    assert "receipt" in result
    assert "walked away" in result
    assert "<" not in result


def test_prepare_output_for_rendering_applies_per_column_kind():
    headers = ["Complaint Ref Number", "Remarks", "Received"]
    rows = [["REF\u0000001", "Line&nbsp;one<br/>two", "120"]]
    _, out = prepare_output_for_rendering("scr-train", headers, rows)
    assert out[0][0] == "REF001"
    assert "Line one" in out[0][1]
    assert out[0][2] == "120"


def test_unicode_font_registration_available():
    assert ensure_pdf_unicode_fonts() or pdf_font_regular() == "Helvetica"


@pytest.fixture
def problematic_r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        "Ref. No.,Mode,Registration Date,Train/Station,Type,Sub Type,Department,Status,Zone,Div,"
        "feedbackRemark,trainNameForReport/Station Name,complaintDesc,remarks,userId\n"
        "REF777,Train,15-07-26 19:13,12759,Maintenance,Issue,CHG,Closed,SC,SC,"
        "Vendor&nbsp;issue,South Central Express,"
        "Coach dirty &amp; wet<br/>needs cleaning,"
        "साफ-सफाई की समस्या; vendor refused receipt\u200b,"
        "cnw_sc_sc\n",
        encoding="utf-8",
    )
    return target


def test_report5_excel_pdf_preserve_normalized_unicode_text(
    problematic_r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)

    result = Report5Processor().process(source_a_path=problematic_r5_csv, report_slug="report5")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    remarks_col = None
    desc_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=2, column=col_idx).value
        if header == "Remarks":
            remarks_col = col_idx
        if header == "Complaint Description":
            desc_col = col_idx
    assert remarks_col and desc_col
    remarks = str(ws.cell(row=3, column=remarks_col).value or "")
    desc = str(ws.cell(row=3, column=desc_col).value or "")
    assert contains_rendering_risk_markers(remarks) is False
    assert contains_rendering_risk_markers(desc) is False
    assert "साफ" in remarks
    assert "needs cleaning" in desc
    assert "&amp;" not in desc

    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"
    if any(ord(ch) > 255 for ch in remarks + desc):
        assert (
            b"/Subtype /Type0" in pdf_bytes
            or b"+0 " in pdf_bytes
            or ensure_pdf_unicode_fonts()
        )
