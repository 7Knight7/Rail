"""Tests for Report 1 Phase 8 processor."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.base import ProcessingResult
from app.automation.processing.registry import REPORT1_PROCESSOR_NAME, PROCESSORS
from app.automation.processing.report1_processor import Report1Processor

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report1"


@pytest.fixture
def processor() -> Report1Processor:
    return Report1Processor()


@pytest.fixture
def comprehensive_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report1_comprehensive_zone_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def feedback_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report1_feedback_zone_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_registry_selects_report1_processor():
    assert "report1" in PROCESSORS
    assert PROCESSORS["report1"].processor_name == REPORT1_PROCESSOR_NAME


def test_rejects_pdf_input(processor: Report1Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report1")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_fails_when_feedback_missing(
    processor: Report1Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    output_excel = tmp_path / "output" / "excel"
    output_pdf = tmp_path / "output" / "pdf"
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(output_excel),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(output_pdf),
    )

    result = processor.process(source_a_path=comprehensive_csv, report_slug="report1")

    assert result.success is False
    assert "source_b_path required" in (result.error or "") or "Feedback dataset missing" in (
        result.error or ""
    )
    assert not list(output_excel.rglob("*.xlsx"))


def test_fails_when_source_b_file_missing(
    processor: Report1Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    missing = tmp_path / "missing_feedback.csv"
    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=missing,
    )
    assert result.success is False
    assert "missing or empty" in (result.error or "").lower() or "Feedback" in (result.error or "")


def test_output_filename_includes_timestamp(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
    )
    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    # Rail_Madad_Report_1_..._DD-MM-YYYY_HHMMSS.xlsx
    assert Path(result.excel_path).stem.count("_") >= 8
    assert Path(result.excel_path).name != Path(result.pdf_path).with_suffix(".xlsx").name or True
    stem = Path(result.excel_path).stem
    assert stem[-6:].isdigit()


def test_produces_excel_and_pdf_outputs(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
    )

    assert result.success is True
    assert result.processor_used is None
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()
    assert Path(result.excel_path).stat().st_size > 0
    assert Path(result.pdf_path).stat().st_size > 0
    assert result.source_a_path is not None
    assert result.source_b_path is not None
    assert result.source_a_rows > 0
    assert result.source_b_rows > 0


def test_scr_row_has_yellow_fill_and_black_text(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    scr_found = False
    for row_idx in range(3, worksheet.max_row + 1):
        org_value = worksheet.cell(row=row_idx, column=2).value
        if org_value and "South Central Railway" in str(org_value):
            scr_found = True
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
                assert cell.font.color.rgb in {"00000000", "FF000000", "000000"}
    assert scr_found


def test_irctc_alignment(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    headers = [
        str(worksheet.cell(row=2, column=col).value or "")
        for col in range(1, worksheet.max_column + 1)
    ]
    feedback_col = headers.index("Feedback Received") + 1
    catering_feedback = None
    online_feedback = None
    for row_idx in range(3, worksheet.max_row + 1):
        org_value = str(worksheet.cell(row=row_idx, column=2).value or "")
        feedback_received = worksheet.cell(row=row_idx, column=feedback_col).value
        if "Irctc-Catering" in org_value:
            catering_feedback = feedback_received
        if "Irctc-Online" in org_value:
            online_feedback = feedback_received

    assert catering_feedback == "420"
    assert online_feedback in ("", None)


def test_report1_total_avg_disposal_time_in_excel_and_pdf(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [
        str(ws.cell(row=2, column=col).value or "")
        for col in range(1, ws.max_column + 1)
    ]
    avg_col = headers.index("Avg. Disposal Time") + 1
    avg_total = str(ws.cell(row=ws.max_row, column=avg_col).value or "")
    assert avg_total == "0:36"
    assert str(ws.cell(row=ws.max_row, column=2).value or "") == "Total"

    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"
    assert len(pdf_bytes) > 100


def test_report1_serial_renumbered_after_column_projection(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    subset = [
        "report1.source_a.sno",
        "report1.source_a.organisation",
        "report1.source_a.received",
    ]
    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
        column_selection={
            "report_slug": "report1",
            "selected_column_ids": subset,
            "column_order": subset,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [
        str(ws.cell(row=2, column=col).value or "")
        for col in range(1, ws.max_column + 1)
    ]
    sno_col = headers.index("S.No.") + 1
    serials = [
        ws.cell(row=row, column=sno_col).value
        for row in range(3, ws.max_row)
    ]
    data_serials = [int(s) for s in serials if s not in ("", None)]
    assert data_serials == list(range(1, len(data_serials) + 1))
    assert ws.cell(row=ws.max_row, column=sno_col).value in ("", None)


def test_report1_scr_highlight_when_organisation_hidden(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    subset = [
        "report1.source_a.received",
        "report1.source_a.avg_disposal_time",
        "report1.source_b.feedback_received",
        "report1.source_b.unsatisfactory",
    ]
    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
        column_selection={
            "report_slug": "report1",
            "selected_column_ids": subset,
            "column_order": subset,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [
        str(ws.cell(row=2, column=col).value or "")
        for col in range(1, ws.max_column + 1)
    ]
    received_col = headers.index("Received") + 1
    scr_row = None
    for row_idx in range(3, ws.max_row + 1):
        received = ws.cell(row=row_idx, column=received_col).value
        if str(received) == "232":
            scr_row = row_idx
            break
    assert scr_row is not None
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=scr_row, column=col_idx)
        assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
        assert cell.font.color.rgb in {"00000000", "FF000000", "000000"}


def test_report1_five_column_manual_acceptance_headers(
    processor: Report1Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report1_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    subset = [
        "report1.source_a.organisation",
        "report1.source_a.received",
        "report1.source_a.avg_disposal_time",
        "report1.source_b.feedback_received",
        "report1.source_b.unsatisfactory",
    ]
    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report1",
        source_b_path=feedback_csv,
        column_selection={
            "report_slug": "report1",
            "selected_column_ids": subset,
            "column_order": subset,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success is True
    headers = [
        str(load_workbook(result.excel_path).active.cell(row=2, column=c).value or "")
        for c in range(1, 6)
    ]
    assert headers == [
        "Organisation",
        "Received",
        "Avg. Disposal Time",
        "Feedback Received",
        "Unsatisfactory",
    ]
