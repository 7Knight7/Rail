"""Tests for SCR Report 5/6 output column filter system."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.column_config import (
    output_column_catalog,
    project_scr_for_output,
    sanitize_projection_keys,
    validate_projection_selection,
)
from app.automation.processing.report5_processor import Report5Processor
from app.automation.processing.report6_processor import Report6Processor
from app.automation.processing.scr_output_columns import (
    REPORT5_DEFAULT_SCR_IDS,
    REPORT5_SCR_IDS,
    REPORT6_DEFAULT_SCR_IDS,
    REPORT6_SCR_IDS,
    migrate_scr_to_namespaced_ids,
    scr_default_ids,
)
from app.automation.scr_field_map import canonicalize_scr_rows

FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def test_report5_and_report6_allowlists_are_independent():
    assert REPORT5_SCR_IDS.isdisjoint(REPORT6_SCR_IDS)
    assert "scr-train.created_on" in REPORT5_SCR_IDS
    assert "scr-station.created_on" in REPORT6_SCR_IDS
    assert "scr-train.created_on" not in REPORT6_SCR_IDS


def test_default_selections_match_spec():
    assert len(REPORT5_DEFAULT_SCR_IDS) == 13
    assert "scr-train.dept_code" not in REPORT5_DEFAULT_SCR_IDS
    assert len(REPORT6_DEFAULT_SCR_IDS) == 11
    assert "scr-station.created_on" not in REPORT6_DEFAULT_SCR_IDS
    assert "scr-station.train_name_for_report" not in REPORT6_DEFAULT_SCR_IDS


def test_catalog_exposes_full_allowlist_not_only_defaults():
    r5 = output_column_catalog("scr-train")
    r6 = output_column_catalog("scr-station")
    assert len(r5) > len(REPORT5_DEFAULT_SCR_IDS)
    assert len(r6) > len(REPORT6_DEFAULT_SCR_IDS)
    assert any(c["id"] == "scr-train.dept_code" for c in r5)
    assert any(c["group_title"] == "Identifiers" for c in r5)


def test_flexible_subset_validation_report5():
    subset = [
        "scr-train.complaint_ref_no",
        "scr-train.created_on",
        "scr-train.train_station",
        "scr-train.feedback_remark",
        "scr-train.complaint_desc",
    ]
    validate_projection_selection("scr-train", subset)


def test_flexible_subset_validation_report6():
    subset = [
        "scr-station.complaint_ref_no",
        "scr-station.train_station",
        "scr-station.comp_type_name",
        "scr-station.remarks",
    ]
    validate_projection_selection("scr-station", subset)


def test_legacy_keys_migrate_to_namespaced():
    migrated = migrate_scr_to_namespaced_ids(
        "scr-train",
        ["complaintRefNo", "createdOn", "scr-train.feedback_remark"],
    )
    assert migrated == [
        "scr-train.complaint_ref_no",
        "scr-train.created_on",
        "scr-train.feedback_remark",
    ]


def test_clear_all_blocks_validation():
    with pytest.raises(ValueError, match="at least one"):
        validate_projection_selection("scr-train", [])


def test_cross_report_ids_rejected():
    with pytest.raises(ValueError, match="approved allowlist"):
        validate_projection_selection(
            "scr-train",
            ["scr-station.complaint_ref_no"],
        )


def test_project_scr_rows_preserves_order_and_regenerates_sno():
    rows = canonicalize_scr_rows(
        [
            {
                "Ref. No.": "R1",
                "Mode": "Train",
                "Registration Date": "15-07-26",
                "Train/Station": "12760",
                "Type": "Maintenance",
                "Sub Type": "Issue",
                "Dept": "CHG",
                "Status": "Closed",
                "Zone": "SC",
                "Div": "SC",
                "feedbackRemark": "Remark",
                "trainNameForReport/Station Name": "Express",
                "complaintDesc": "Desc",
                "remarks": "Rem",
                "userId": "u1",
            }
        ]
    )
    keys = [
        "scr-train.complaint_ref_no",
        "scr-train.sno",
        "scr-train.feedback_remark",
    ]
    headers, out_rows, labels, resolved, _source = project_scr_for_output(
        "scr-train",
        rows,
        selected_keys=keys,
    )
    assert headers == ["Complaint Ref Number", "S.No.", "Feedback Remark"]
    assert out_rows[0] == ["R1", "1", "Remark"]
    assert resolved == keys


def test_project_without_sno_omits_serial():
    rows = canonicalize_scr_rows(
        [
            {
                "Ref. No.": "R1",
                "Mode": "Train",
                "Train/Station": "12760",
                "Type": "Maintenance",
                "Sub Type": "Issue",
                "Dept": "CHG",
                "Status": "Closed",
                "Zone": "SC",
                "Div": "SC",
                "feedbackRemark": "",
                "trainNameForReport/Station Name": "Express",
                "complaintDesc": "Desc",
                "remarks": "Rem",
                "userId": "u1",
            }
        ]
    )
    keys = ["scr-train.complaint_ref_no", "scr-train.user_id"]
    headers, out_rows, *_ = project_scr_for_output("scr-train", rows, selected_keys=keys)
    assert headers == ["Complaint Ref Number", "User ID"]
    assert "S.No." not in headers
    assert out_rows[0] == ["R1", "u1"]


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _headers_from_excel(path: str) -> list[str]:
    ws = load_workbook(path).active
    return [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "scr-train"
    extracted.mkdir(parents=True)
    target = extracted / "scr-train_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "scr-station"
    extracted.mkdir(parents=True)
    target = extracted / "scr-station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report5_manual_subset_reaches_excel(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    subset = [
        "scr-train.complaint_ref_no",
        "scr-train.created_on",
        "scr-train.train_station",
        "scr-train.feedback_remark",
        "scr-train.complaint_desc",
    ]
    result = Report5Processor().process(
        source_a_path=r5_csv,
        report_slug="scr-train",
        column_selection={
            "report_slug": "scr-train",
            "selected_column_ids": subset,
            "column_order": subset,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert len(headers) == 5
    assert headers == [
        "Complaint Ref Number",
        "Created On",
        "Train/Station",
        "Feedback Remark",
        "Complaint Description",
    ]


def test_report6_manual_subset_reaches_excel(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    subset = [
        "scr-station.complaint_ref_no",
        "scr-station.train_station",
        "scr-station.comp_type_name",
        "scr-station.remarks",
    ]
    result = Report6Processor().process(
        source_a_path=r6_csv,
        report_slug="scr-station",
        column_selection={
            "report_slug": "scr-station",
            "selected_column_ids": subset,
            "column_order": subset,
            "configuration_source": "manual_snapshot",
        },
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert len(headers) == 4
    assert result.selected_column_ids == subset


def test_sanitize_select_all_only_current_report():
    all_r5 = sanitize_projection_keys(list(REPORT5_SCR_IDS), "scr-train")
    assert len(all_r5) == len(REPORT5_SCR_IDS)
    assert all(key.startswith("scr-train.") for key in all_r5)


def test_defaults_per_slug():
    assert scr_default_ids("scr-train") == REPORT5_DEFAULT_SCR_IDS
    assert scr_default_ids("scr-station") == REPORT6_DEFAULT_SCR_IDS
