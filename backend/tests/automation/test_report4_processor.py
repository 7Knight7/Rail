"""Tests for Report 4 Phase 8 processor (Cause-wise Top 10 Trains - 7 Types)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.registry import PROCESSORS
from app.automation.processing.report4_processor import Report4Processor, TOP_N, OUTPUT_HEADERS
from app.automation.report4_filters import COMPLAINT_TYPES_ORDERED, get_type_configs

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report4"


@pytest.fixture
def processor() -> Report4Processor:
    return Report4Processor()


@pytest.fixture
def sample_type_datasets() -> dict[str, list[dict[str, str]]]:
    """Create sample datasets for all 7 complaint types."""
    datasets = {}
    for idx, type_name in enumerate(COMPLAINT_TYPES_ORDERED):
        base_received = 100 - (idx * 10)
        rows = []
        for row_idx in range(15):
            zone = "South Central Railway" if row_idx == 2 else f"Zone {row_idx}"
            rows.append({
                "S.No.": str(row_idx + 1),
                "Train No.": f"1234{idx}{row_idx}",
                "Train Name": f"{type_name} Train {row_idx + 1}",
                "Owning Zone": zone,
                "Owning Division": f"Division {row_idx}",
                "Received": str(base_received - row_idx * 5),
            })
        datasets[type_name] = rows
    return datasets


def test_registry_selects_report4_processor():
    assert "report4" in PROCESSORS
    assert PROCESSORS["report4"].processor_name == "report4_causewise_top10_processor"


def test_rejects_pdf_input(processor: Report4Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report4")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_produces_excel_and_pdf_outputs(
    processor: Report4Processor,
    sample_type_datasets: dict[str, list[dict[str, str]]],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    dummy_csv = tmp_path / "extracted" / "report4" / "dummy.csv"
    dummy_csv.parent.mkdir(parents=True, exist_ok=True)
    dummy_csv.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy_csv,
        report_slug="report4",
        type_datasets=sample_type_datasets,
    )

    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()


def test_seven_types_in_exact_order(
    processor: Report4Processor,
    sample_type_datasets: dict[str, list[dict[str, str]]],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    dummy_csv = tmp_path / "extracted" / "report4" / "dummy.csv"
    dummy_csv.parent.mkdir(parents=True, exist_ok=True)
    dummy_csv.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy_csv,
        report_slug="report4",
        type_datasets=sample_type_datasets,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    section_titles_found = []
    for row_idx in range(1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_idx, column=1).value
        if cell_value and "Rail Madad 10 trains having maximum" in str(cell_value):
            section_titles_found.append(cell_value)

    assert len(section_titles_found) == 7

    type_configs = get_type_configs()
    for idx, config in enumerate(type_configs):
        assert config.section_title in section_titles_found[idx]


def test_top_10_per_type(
    processor: Report4Processor,
    sample_type_datasets: dict[str, list[dict[str, str]]],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    dummy_csv = tmp_path / "extracted" / "report4" / "dummy.csv"
    dummy_csv.parent.mkdir(parents=True, exist_ok=True)
    dummy_csv.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy_csv,
        report_slug="report4",
        type_datasets=sample_type_datasets,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    in_section = False
    section_count = 0
    rows_in_current_section = 0

    for row_idx in range(1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_idx, column=1).value
        
        if cell_value and "Rail Madad 10 trains having maximum" in str(cell_value):
            if in_section:
                assert rows_in_current_section <= TOP_N
            in_section = True
            section_count += 1
            rows_in_current_section = 0
        elif in_section and cell_value == "S.No.":
            pass
        elif in_section and cell_value and str(cell_value).strip().isdigit():
            rows_in_current_section += 1

    if in_section:
        assert rows_in_current_section <= TOP_N

    assert section_count == 7


def test_scr_rows_highlighted_per_section(
    processor: Report4Processor,
    sample_type_datasets: dict[str, list[dict[str, str]]],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    dummy_csv = tmp_path / "extracted" / "report4" / "dummy.csv"
    dummy_csv.parent.mkdir(parents=True, exist_ok=True)
    dummy_csv.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy_csv,
        report_slug="report4",
        type_datasets=sample_type_datasets,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    scr_rows_found = 0
    for row_idx in range(1, worksheet.max_row + 1):
        owning_zone = worksheet.cell(row=row_idx, column=3).value
        if owning_zone and "south central railway" in str(owning_zone).lower():
            scr_rows_found += 1
            for col_idx in range(1, len(OUTPUT_HEADERS) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
                assert cell.font.color.rgb in {"00000000", "FF000000", "000000"}

    assert scr_rows_found >= 7


def test_train_no_preserved_as_text(
    processor: Report4Processor,
    sample_type_datasets: dict[str, list[dict[str, str]]],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    dummy_csv = tmp_path / "extracted" / "report4" / "dummy.csv"
    dummy_csv.parent.mkdir(parents=True, exist_ok=True)
    dummy_csv.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy_csv,
        report_slug="report4",
        type_datasets=sample_type_datasets,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    train_no_col = 5
    text_format_found = False
    for row_idx in range(3, min(worksheet.max_row + 1, 20)):
        cell = worksheet.cell(row=row_idx, column=train_no_col)
        if cell.value and str(cell.value).strip().isdigit():
            if cell.number_format == "@":
                text_format_found = True
                break

    assert text_format_found


def test_complaint_types_ordering():
    """Verify the complaint types are in the exact specified order."""
    expected_order = [
        "Security",
        "Coach Cleanliness",
        "Bedroll",
        "Water Availability",
        "Electrical Equipment",
        "Catering and Vending Services",
        "Coach Maintenance",
    ]
    assert COMPLAINT_TYPES_ORDERED == expected_order
