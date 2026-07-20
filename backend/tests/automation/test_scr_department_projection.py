"""Tests for Department removal from Reports 5/6 final output projection."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.column_config import output_column_catalog
from app.automation.processing.output_columns import (
    REPORT5_OUTPUT_COLUMNS,
    REPORT5_SELECTED_KEYS,
    REPORT5_VISIBLE_LABELS,
    REPORT6_OUTPUT_COLUMNS,
    REPORT6_SELECTED_KEYS,
    REPORT6_VISIBLE_LABELS,
)
from app.automation.processing.report5_processor import Report5Processor
from app.automation.processing.report6_processor import Report6Processor
from app.automation.scr_field_map import canonicalize_scr_rows
from app.features.reports.config_store import load_report_config, save_report_config

FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _headers_from_excel(path: str) -> list[str]:
    ws = load_workbook(path).active
    return [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]


def test_report5_selected_keys_exclude_department_and_status():
    assert len(REPORT5_SELECTED_KEYS) == 13
    assert "scr-train.dept_code" not in REPORT5_SELECTED_KEYS
    assert "scr-train.status" not in REPORT5_SELECTED_KEYS
    assert len(REPORT5_VISIBLE_LABELS) == 13
    assert "Department" not in REPORT5_VISIBLE_LABELS
    assert "Status" not in REPORT5_VISIBLE_LABELS


def test_report6_selected_keys_exclude_department_and_status():
    assert len(REPORT6_SELECTED_KEYS) == 11
    assert "scr-station.dept_code" not in REPORT6_SELECTED_KEYS
    assert "scr-station.status" not in REPORT6_SELECTED_KEYS
    assert len(REPORT6_VISIBLE_LABELS) == 11
    assert "Department" not in REPORT6_VISIBLE_LABELS
    assert "Status" not in REPORT6_VISIBLE_LABELS


def test_output_column_catalog_includes_optional_fields():
    r5_ids = {entry["id"] for entry in output_column_catalog("scr-train")}
    r6_ids = {entry["id"] for entry in output_column_catalog("scr-station")}
    assert "scr-train.dept_code" in r5_ids
    assert "scr-station.dept_code" in r6_ids
    assert "scr-train.dept_code" not in REPORT5_SELECTED_KEYS


def test_saved_config_with_department_migrates_to_13_cols(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    config_dir = tmp_path / "report-configs"
    monkeypatch.setattr("app.features.reports.config_store.CONFIG_DIR", config_dir)

    old_keys = [column.key for column in REPORT5_OUTPUT_COLUMNS]
    assert "department" in old_keys
    save_report_config("scr-train", {"column_order": old_keys, "selected_column_ids": old_keys})

    migrated = load_report_config("scr-train")
    assert migrated is not None
    assert "department" not in migrated["column_order"]
    assert all(str(k).startswith("scr-train.") for k in migrated["column_order"])


def test_saved_config_with_department_migrates_to_11_cols(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    config_dir = tmp_path / "report-configs"
    monkeypatch.setattr("app.features.reports.config_store.CONFIG_DIR", config_dir)

    old_keys = [column.key for column in REPORT6_OUTPUT_COLUMNS]
    assert "department" in old_keys
    save_report_config("scr-station", {"column_order": old_keys, "selected_column_ids": old_keys})

    migrated = load_report_config("scr-station")
    assert migrated is not None
    assert "department" not in migrated["column_order"]
    assert all(str(k).startswith("scr-station.") for k in migrated["column_order"])


def test_internal_rows_retain_department_after_canonicalize():
    raw = [
        {
            "Ref. No.": "REF001",
            "Mode": "Train",
            "Registration Date": "15-07-26",
            "Train/Station": "12760",
            "Type": "Maintenance",
            "Sub Type": "Issue",
            "Department": "CHG",
            "Status": "Closed",
            "Zone": "SC",
            "Div": "SC",
            "feedbackRemark": "",
            "trainNameForReport/Station Name": "Express",
            "complaintDesc": "Desc",
            "remarks": "Rem",
            "userId": "u1",
        }
    ]
    rows = canonicalize_scr_rows(raw)
    assert rows[0].get("deptCode") == "CHG"


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report5_excel_pdf_exclude_department(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT5_VISIBLE_LABELS
    assert "Department" not in headers
    assert result.selected_column_ids is not None
    assert "scr-train.dept_code" not in (result.selected_column_ids or [])


def test_report6_excel_pdf_exclude_department(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    result = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT6_VISIBLE_LABELS
    assert "Department" not in headers
    assert result.selected_column_ids is not None
    assert "scr-station.dept_code" not in (result.selected_column_ids or [])
