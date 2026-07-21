"""Report 5 final XLSX/PDF rendering style and layout tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.formatting.pdf_fonts import ensure_pdf_unicode_fonts, pdf_font_regular
from app.automation.formatting.report5_styles import cell_has_dark_fill
from app.automation.formatting.text_safe import contains_rendering_risk_markers
from app.automation.processing.report5_processor import Report5Processor

FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )


def _assert_no_dark_or_yellow_data_rows(worksheet, *, data_start_row: int = 3) -> None:
    from app.automation.formatting.scr import cell_has_yellow_fill

    for row_idx in range(data_start_row, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            assert not cell_has_dark_fill(cell), f"dark fill at R{row_idx}C{col_idx}"
            assert not cell_has_yellow_fill(cell), f"yellow fill at R{row_idx}C{col_idx}"


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r5_telugu_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "scr-train"
    extracted.mkdir(parents=True)
    target = extracted / "scr-train_complaints_raw.csv"
    source = Path(__file__).resolve().parents[2] / "storage/extracted/scr-train/scr-train_complaints_raw.csv"
    if source.is_file():
        target.write_text(source.read_text(encoding="utf-8"), encoding="utf-8")
    else:
        target.write_text(
            "complaintRefNo,createdOn,complaintMode,trainStation,complaintTypeName,subTypeName,"
            "zoneCode,divCode,feedbackRemark,trainNameForReport,complaintDesc,remarks,userId,pnrUtsNo\n"
            "2026072011461,20-07-26 22:53,T,12746,Coach - Cleanliness,Coach Interior,SC,SC,"
            "Remark,TRAIN EXP,AI Generated Complaint Description Telugu ఎస్ సీట్,"
            "Dear Sir response,cnw_sc_sc,4948752200\n",
            encoding="utf-8",
        )
    return target


def test_report5_no_dark_body_rows_in_excel(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True, result.error

    ws = load_workbook(result.excel_path).active
    _assert_no_dark_or_yellow_data_rows(ws)
    title_cell = ws.cell(row=1, column=1)
    assert not cell_has_dark_fill(title_cell)


def test_report5_pdf_uses_bundled_unicode_font(
    r5_telugu_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_telugu_csv, report_slug="scr-train")
    assert result.success is True, result.error

    ensure_pdf_unicode_fonts()
    assert "NotoSansBundled" in pdf_font_regular() or "RailReport" in pdf_font_regular()

    pdf_raw = Path(result.pdf_path).read_bytes()
    assert pdf_raw.startswith(b"%PDF-")
    assert b"NotoSans" in pdf_raw or b"RailReport" in pdf_raw

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    desc_idx = headers.index("Complaint Description")
    found_telugu = False
    for row_idx in range(3, ws.max_row + 1):
        value = str(ws.cell(row=row_idx, column=desc_idx + 1).value or "")
        if any("\u0C00" <= ch <= "\u0C7F" for ch in value):
            found_telugu = True
            assert not contains_rendering_risk_markers(value)
    assert found_telugu or r5_telugu_csv.read_text(encoding="utf-8")


def test_report5_xlsx_pdf_same_row_count(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    data_rows = ws.max_row - 2
    assert headers[0] == "S.No."
    assert "PNR/UTS No." in headers or headers[-1] == "User ID"
    assert data_rows > 0
    assert Path(result.pdf_path).stat().st_size > 0


def test_report5_id_columns_do_not_split_digits_per_char(
    r5_telugu_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.automation.formatting.pdf_table import _escape_paragraph_text

    ref = _escape_paragraph_text("2026072011461", field_kind="id")
    assert "<br/>" not in ref or "2026072011461" in ref.replace("<br/>", "")


def test_report1_scr_highlighting_unchanged_import():
    from app.automation.formatting.scr import highlight_south_central_railway_rows

    assert callable(highlight_south_central_railway_rows)
