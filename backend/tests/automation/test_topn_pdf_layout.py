"""Focused PDF layout tests for Report 3 (train-no) and Report 4 (types)."""

from __future__ import annotations

from pathlib import Path

import pytest
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape

from app.automation.formatting.topn_pdf import (
    TOPN_BODY_FONT_PT,
    TOPN_HEADER_FONT_PT,
    TOPN_SAFE_MARGIN_PT,
    TOPN_TITLE_FONT_PT,
    allocate_topn_column_widths,
    build_topn_fitted_table,
    choose_topn_landscape_layout,
)
from app.automation.processing.report3_processor import Report3Processor, TOP_N
from app.automation.processing.report4_processor import Report4Processor
from app.automation.processing.topn_output_columns import topn_default_ids, topn_labels
from app.automation.report4_filters import COMPLAINT_TYPES_ORDERED, get_type_configs

FIXTURES_R3 = Path(__file__).resolve().parent.parent / "fixtures" / "report3"


def _default_headers(slug: str) -> list[str]:
    return topn_labels(topn_default_ids(slug), slug)


def test_topn_layout_prefers_a3_and_fits_width():
    headers = _default_headers("train-no")
    pagesize, col_widths, margin = choose_topn_landscape_layout(headers)
    assert pagesize == landscape(A3)
    assert margin == TOPN_SAFE_MARGIN_PT
    usable = pagesize[0] - (2 * margin)
    assert abs(sum(col_widths) - usable) < 0.6
    train_idx = headers.index("Train Name")
    rating_idx = headers.index("Average Rating")
    received_idx = headers.index("Received")
    assert col_widths[train_idx] > col_widths[rating_idx] > col_widths[received_idx]


def test_topn_fitted_table_wraps_within_printable_width():
    headers = _default_headers("types")
    row = []
    samples = {
        "S.No.": "1",
        "S.No": "1",
        "Train Name": "PURI-ANVT NEELACHAL [SUPERFAST] EXTRA LONG NAME",
        "Owning Zone": "Northeast Frontier Railway",
        "Owning Division": "TINSUKIA DIVISION",
        "Train No.": "12875",
        "Train No": "12875",
        "Received": "91",
        "% Share": "1.13",
        "Closed": "69",
        "% Closed": "75.82",
        "Pending": "22",
        "Average Rating": "Satisfactory",
    }
    for header in headers:
        row.append(samples.get(header, "x"))
    table_data = [headers, row]
    style = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
    ]
    table, pagesize, margin, col_widths = build_topn_fitted_table(table_data, style)
    usable = pagesize[0] - (2 * margin)
    wrapped_w, _ = table.wrap(usable, pagesize[1])
    assert wrapped_w <= usable + 1.0
    assert TOPN_BODY_FONT_PT >= 8.5
    assert TOPN_HEADER_FONT_PT >= 9.0
    assert TOPN_TITLE_FONT_PT >= 15.0
    assert abs(sum(col_widths) - usable) < 0.6


def test_report3_pdf_layout_valid_and_has_20_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    processor = Report3Processor()
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    extracted = tmp_path / "extracted" / "train-no"
    extracted.mkdir(parents=True, exist_ok=True)
    src = extracted / "trainwise_raw.csv"
    src.write_text((FIXTURES_R3 / "trainwise_raw.csv").read_text(encoding="utf-8"), encoding="utf-8")

    result = processor.process(source_a_path=src, report_slug="train-no")
    assert result.success is True
    pdf_path = Path(result.pdf_path)
    payload = pdf_path.read_bytes()
    assert payload.startswith(b"%PDF")
    assert result.processed_row_count == TOP_N

    from openpyxl import load_workbook

    wb = load_workbook(result.excel_path)
    ws = wb.active
    excel_rows = ws.max_row - 2
    assert excel_rows == TOP_N

    headers = list(result.visible_columns or _default_headers("train-no"))
    table_data = [headers]
    for r in range(3, 3 + TOP_N):
        table_data.append([str(ws.cell(r, c).value or "") for c in range(1, len(headers) + 1)])
    table, pagesize, margin, _ = build_topn_fitted_table(
        table_data,
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
    )
    usable = pagesize[0] - (2 * margin)
    assert table.wrap(usable, pagesize[1])[0] <= usable + 1.0
    assert pagesize == landscape(A3)


def test_report4_pdf_seven_sections_page_breaks(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    processor = Report4Processor()
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    datasets = {}
    for idx, type_name in enumerate(COMPLAINT_TYPES_ORDERED):
        rows = []
        for row_idx in range(12):
            rows.append(
                {
                    "S.No.": str(row_idx + 1),
                    "Train No.": f"12{idx}{row_idx:02d}",
                    "Train Name": f"{type_name} Train {row_idx + 1} LONG",
                    "Owning Zone": "South Central Railway" if row_idx == 1 else f"Zone {row_idx}",
                    "Owning Division": f"Division {row_idx}",
                    "Received": str(100 - row_idx),
                    "Closed": str(80 - row_idx),
                    "Pending": str(20),
                    "% Share": "1.00",
                    "% Closed": "80.00",
                    "Average Rating": "Satisfactory",
                }
            )
        datasets[type_name] = rows

    dummy = tmp_path / "extracted" / "types" / "dummy.csv"
    dummy.parent.mkdir(parents=True, exist_ok=True)
    dummy.write_text("dummy", encoding="utf-8")

    result = processor.process(
        source_a_path=dummy,
        report_slug="types",
        type_datasets=datasets,
    )
    assert result.success is True
    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes.startswith(b"%PDF")

    page_count = len(__import__("re").findall(rb"/Type\s*/Page(?!s)\b", pdf_bytes))
    assert page_count >= 7

    configs = get_type_configs()
    assert len(configs) == 7
    from openpyxl import load_workbook

    wb = load_workbook(result.excel_path)
    ws = wb.active
    found = [
        ws.cell(r, 1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value and "Rail Madad 10 trains having maximum" in str(ws.cell(r, 1).value)
    ]
    assert len(found) == 7
    for idx, cfg in enumerate(configs):
        assert cfg.section_title == found[idx]


def test_allocate_widths_selected_column_subset():
    headers = ["Train Name", "Received", "Average Rating"]
    widths = allocate_topn_column_widths(headers, 1000.0)
    assert abs(sum(widths) - 1000.0) < 0.01
    assert widths[0] > widths[2] > widths[1]
