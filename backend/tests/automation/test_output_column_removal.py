"""Tests for removed visible columns on Reports 1, 2, 5 and 6."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.output_columns import (
    REPORT1_VISIBLE_LABELS,
    REPORT2_VISIBLE_LABELS,
    REPORT5_VISIBLE_LABELS,
    REPORT6_VISIBLE_LABELS,
)
from app.automation.processing.report1_processor import Report1Processor
from app.automation.processing.report5_processor import Report5Processor
from app.automation.processing.report6_processor import Report6Processor
from app.automation.processing.report2_processor import Report2Processor
from app.automation.scr_field_map import canonicalize_scr_rows

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"
FIXTURES_R2 = Path(__file__).resolve().parent.parent / "fixtures" / "report2"
FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _headers_from_excel(path: str) -> list[str]:
    ws = load_workbook(path).active
    return [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]


def test_report1_visible_column_count_and_removed_fields():
    assert len(REPORT1_VISIBLE_LABELS) == 12
    assert "Closing Balance" not in REPORT1_VISIBLE_LABELS


def test_report2_pct_balance_maps_to_pct_disposal_removed():
    assert len(REPORT2_VISIBLE_LABELS) == 12
    assert "% Disposal" not in REPORT2_VISIBLE_LABELS
    assert "% Balance" not in REPORT2_VISIBLE_LABELS


def test_report5_visible_column_count_and_status_removed():
    assert len(REPORT5_VISIBLE_LABELS) == 13
    assert "Status" not in REPORT5_VISIBLE_LABELS


def test_report6_visible_column_count_and_status_removed():
    assert len(REPORT6_VISIBLE_LABELS) == 11
    assert "Status" not in REPORT6_VISIBLE_LABELS


def test_scr_internal_rows_retain_status():
    raw = [
        {
            "Ref. No.": "REF001",
            "Mode": "Train",
            "Registration Date": "15-07-26",
            "Train/Station": "12760",
            "Type": "Maintenance",
            "Sub Type": "Issue",
            "Department": "CHG",
            "Status": "Closed",
            "Zone": "SC",
            "Div": "SC",
            "feedbackRemark": "",
            "trainNameForReport/Station Name": "Express",
            "complaintDesc": "Desc",
            "remarks": "Rem",
            "userId": "u1",
        }
    ]
    rows = canonicalize_scr_rows(raw)
    assert rows[0].get("status") == "Closed"


@pytest.fixture
def r1_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "comprehensive.csv"
    feedback = extracted / "feedback.csv"
    comprehensive.write_text((FIXTURES_R1 / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"))
    feedback.write_text((FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"))
    return comprehensive, feedback


@pytest.fixture
def r2_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "comprehensive.csv"
    feedback = extracted / "feedback.csv"
    comprehensive.write_text((FIXTURES_R2 / "division_comprehensive_raw.csv").read_text(encoding="utf-8"))
    feedback.write_text((FIXTURES_R2 / "division_feedback_raw.csv").read_text(encoding="utf-8"))
    return comprehensive, feedback


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text((FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"))
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text((FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"))
    return target


def test_report1_excel_pdf_exclude_closing_balance(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT1_VISIBLE_LABELS
    assert "Closing Balance" not in headers


def test_report2_excel_pdf_exclude_pct_disposal(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT2_VISIBLE_LABELS
    assert "% Disposal" not in headers


def test_report5_excel_pdf_exclude_status(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT5_VISIBLE_LABELS
    assert "Status" not in headers


def test_report6_excel_pdf_exclude_status(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    result = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == REPORT6_VISIBLE_LABELS
    assert "Status" not in headers
