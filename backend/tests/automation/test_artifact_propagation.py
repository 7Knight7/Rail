"""Tests for run-scoped artifact propagation (no stale filesystem fallback)."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.automation.handlers.base import BaseReportHandler
from app.automation.processing.report5_processor import OUTPUT_HEADERS, Report5Processor
from app.automation.run_registry import list_run_artifacts, register_artifact, validate_artifact_file
from app.infrastructure.database.models import AutomationArtifactModel, AutomationRunModel


class _ConcreteHandler(BaseReportHandler):
    async def execute(self, page, session, report):
        raise NotImplementedError


def test_build_success_result_has_no_latest_pdf_fallback_url():
    result = _ConcreteHandler().build_success_result(
        slug="report1",
        pdf_path="/tmp/report1.pdf",
    )
    assert result.pdf_download_url is None
    assert result.pdf_path == "/tmp/report1.pdf"


@pytest.mark.asyncio
async def test_register_artifact_rejects_stale_file(
    test_session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.automation.config import config

    pdf_root = tmp_path / "pdf" / "report1"
    pdf_root.mkdir(parents=True)
    monkeypatch.setattr(config, "output_pdf_dir", str(tmp_path / "pdf"))
    monkeypatch.setattr(config, "output_excel_dir", str(tmp_path / "excel"))
    monkeypatch.setattr(config, "extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(config, "pdf_archive_dir", str(tmp_path / "archive"))

    pdf_path = pdf_root / "stale.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 stale")
    stale_ts = (datetime.now(UTC) - timedelta(hours=2)).timestamp()
    import os

    os.utime(pdf_path, (stale_ts, stale_ts))

    run_id = str(uuid4())
    started = datetime.now(UTC)
    test_session.add(
        AutomationRunModel(
            id=run_id,
            profile_id="test-profile",
            status="running",
            trigger_type="cdp_in_process",
            success_count=0,
            failure_count=0,
            started_at=started,
        )
    )
    await test_session.commit()

    art = await register_artifact(
        test_session,
        run_id=run_id,
        report_slug="report1",
        report_name="report1",
        file_type="pdf",
        file_path=pdf_path,
    )
    assert art is not None
    assert art.status == "missing"
    assert art.file_size_bytes == 0


@pytest.mark.asyncio
async def test_list_run_artifacts_newest_first(test_session):
    run_id = str(uuid4())
    test_session.add(
        AutomationRunModel(
            id=run_id,
            profile_id="test-profile",
            status="completed",
            trigger_type="cdp_in_process",
            success_count=1,
            failure_count=0,
        )
    )
    older_id = str(uuid4())
    newer_id = str(uuid4())
    now = datetime.now(UTC)
    test_session.add_all(
        [
            AutomationArtifactModel(
                id=older_id,
                run_id=run_id,
                artifact_type="pdf",
                file_path="storage/output/pdf/report1/old.pdf",
                file_size_bytes=10,
                report_slug="report1",
                report_name="report1",
                status="ready",
                created_at=now - timedelta(minutes=5),
            ),
            AutomationArtifactModel(
                id=newer_id,
                run_id=run_id,
                artifact_type="pdf",
                file_path="storage/output/pdf/report1/new.pdf",
                file_size_bytes=10,
                report_slug="report1",
                report_name="report1",
                status="ready",
                created_at=now,
            ),
        ]
    )
    await test_session.commit()

    artifacts = await list_run_artifacts(test_session, run_id)
    pdf_ids = [a.id for a in artifacts if a.artifact_type == "pdf"]
    assert pdf_ids[0] == newer_id


def test_report5_template_trims_extra_columns(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    template_path = (
        Path(__file__).resolve().parents[2]
        / "app"
        / "infrastructure"
        / "seed"
        / "sample_workbooks"
        / "scr_train_original.xlsx"
    )
    if not template_path.exists():
        pytest.skip("template workbook missing")

    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    csv_path = extracted / "data.csv"
    csv_path.write_text(
        "Ref. No.,Mode,Registration Date,Train/Station,Type,Sub Type,Department,Status,"
        "Zone,Div,feedbackRemark,trainNameForReport/Station Name,complaintDesc,remarks,userId\n"
        "REF001,Train,15-07-26,Train A,Type1,Sub1,Dept1,Pending,SC,HYB,,Train A,Desc,Rem,uid\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    processor = Report5Processor()
    monkeypatch.setattr(processor, "_find_template", lambda: template_path)

    result = processor.process(source_a_path=csv_path, report_slug="report5")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, len(OUTPUT_HEADERS) + 1)]
    assert headers == OUTPUT_HEADERS
    assert ws.max_column == len(OUTPUT_HEADERS)
    assert "Mode" not in headers
