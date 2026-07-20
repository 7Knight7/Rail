"""Tests for Report 1/2/5/6 output column updates."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.output_columns import (
    REMOVED_OUTPUT_LABELS,
    REPORT5_OUTPUT_COLUMNS,
    REPORT5_VISIBLE_LABELS,
    REPORT6_OUTPUT_COLUMNS,
    REPORT6_VISIBLE_LABELS,
    SOURCE_B_DATA_COLUMNS,
    assert_column_order_contains,
    assert_exact_visible_columns,
    default_visible_column_keys,
    migrate_selected_column_keys,
    output_labels,
)
from app.automation.processing.report1_processor import HIDDEN_COLUMNS, Report1Processor
from app.automation.processing.report2_processor import (
    HIDDEN_COLUMNS as R2_HIDDEN,
    Report2Processor,
)
from app.automation.processing.report3_processor import OUTPUT_HEADERS as R3_HEADERS
from app.automation.processing.report4_processor import OUTPUT_HEADERS as R4_HEADERS
from app.automation.processing.report5_processor import OUTPUT_HEADERS as R5_HEADERS, Report5Processor
from app.automation.processing.report6_processor import OUTPUT_HEADERS as R6_HEADERS, Report6Processor

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"
FIXTURES_R2 = Path(__file__).resolve().parent.parent / "fixtures" / "report2"
FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _headers_from_excel(path: str) -> list[str]:
    ws = load_workbook(path).active
    return [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]


def _visible_headers(all_headers: list[str], hidden: set[int]) -> list[str]:
    return [header for idx, header in enumerate(all_headers, start=1) if idx not in hidden]


@pytest.fixture
def r1_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "report1_comprehensive_zone_raw.csv"
    feedback = extracted / "report1_feedback_zone_raw.csv"
    comprehensive.write_text(
        (FIXTURES_R1 / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    feedback.write_text(
        (FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comprehensive, feedback


@pytest.fixture
def r2_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "report2_division_comprehensive_raw.csv"
    feedback = extracted / "report2_division_feedback_raw.csv"
    comprehensive.write_text(
        (FIXTURES_R2 / "division_comprehensive_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    feedback.write_text(
        (FIXTURES_R2 / "division_feedback_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comprehensive, feedback


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report1_exact_visible_columns(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert len(headers) == 12
    assert "Organisation" in headers
    assert "Feedback Received" in headers
    assert "Closing Balance" not in headers
    for label in REMOVED_OUTPUT_LABELS:
        assert label not in headers
    assert result.output_columns == headers
    assert result.visible_columns == headers


def test_report1_subset_projection(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    subset = [
        "report1.source_a.organisation",
        "report1.source_a.received",
        "report1.source_a.closing_balance",
        "report1.source_b.feedback_received",
    ]
    ctx_token = None
    from app.automation.run_context import reset_run_context, set_run_context
    from app.automation.run_context import RunContext
    from app.automation.timing import RunTiming

    ctx = RunContext(
        run_id="subset-test",
        timing=RunTiming(run_id="subset-test"),
        manual_config={"report_slug": "report1", "column_order": subset},
    )
    ctx_token = set_run_context(ctx)
    try:
        result = Report1Processor().process(
            source_a_path=comprehensive,
            report_slug="report1",
            source_b_path=feedback,
        )
    finally:
        reset_run_context(ctx_token)
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert headers == ["Organisation", "Received", "Closing Balance", "Feedback Received"]


def test_report1_default_excludes_unselected_source_a_columns(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert "Closing Balance" not in headers
    assert "Opening Balance" not in headers
    assert_column_order_contains(["Closed", "Avg. Disposal Time"], headers)
    assert result.visible_columns == headers


def test_report1_totals_row(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    total_row = [str(ws.cell(row=ws.max_row, column=c).value or "") for c in range(1, ws.max_column + 1)]
    headers = _headers_from_excel(result.excel_path)
    org_idx = headers.index("Organisation")
    assert total_row[org_idx] == "Total"
    total_visible = [
        str(ws.cell(row=ws.max_row, column=headers.index(h) + 1).value or "")
        for h in headers
    ]
    assert total_visible[headers.index("Received")] == "9321"
    avg_disposal = total_visible[headers.index("Avg. Disposal Time")]
    assert avg_disposal == "0:36"


def test_report2_exact_visible_columns(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert len(headers) == 12
    assert "Division" in headers
    assert "Feedback Received" in headers
    assert "% Disposal" not in headers
    assert result.output_columns == headers
    assert result.visible_columns == headers


def test_report2_visible_columns_include_disposal(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    headers = _headers_from_excel(result.excel_path)
    assert "% Disposal" not in headers
    assert "Division" in headers
    assert "Closing Balance" not in headers
    assert_column_order_contains(["Closed", "Avg. Disposal Time"], headers)


def test_report2_totals_row(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    total_row = [str(ws.cell(row=ws.max_row, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert total_row[1] == "Total"
    assert total_row[0] in {"", "None"}
    headers = _headers_from_excel(result.excel_path)
    total_visible = [
        str(ws.cell(row=ws.max_row, column=headers.index(h) + 1).value or "")
        for h in headers
    ]
    assert total_visible[headers.index("Received")] == "9750"


def test_report5_columns_and_data(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    assert R5_HEADERS == REPORT5_VISIBLE_LABELS
    assert len(R5_HEADERS) == 13
    headers = _headers_from_excel(result.excel_path)
    assert_exact_visible_columns(REPORT5_VISIBLE_LABELS, headers, report_label="Report 5")
    assert "Department" not in headers
    assert "Mode" not in headers
    assert "Ref. No." not in headers
    for label in REMOVED_OUTPUT_LABELS:
        if label in {"Mode", "Complaint Mode", "Complaint Date"}:
            assert label not in headers
    ws = load_workbook(result.excel_path).active
    first_data = [str(ws.cell(row=3, column=c).value or "") for c in range(1, len(headers) + 1)]
    assert first_data[1] == "REF001"
    assert first_data[2] == "15-07-26 21:15"
    assert first_data[6] == "SC"
    assert first_data[9] == "Rajdhani Express"
    assert first_data[10] == "Coach was dirty"
    assert "Status" not in headers
    assert result.output_columns == R5_HEADERS
    assert result.visible_columns == R5_HEADERS


def test_report6_columns_and_data(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    result = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert result.success is True
    assert R6_HEADERS == REPORT6_VISIBLE_LABELS
    assert len(R6_HEADERS) == 11
    headers = _headers_from_excel(result.excel_path)
    assert_exact_visible_columns(REPORT6_VISIBLE_LABELS, headers, report_label="Report 6")
    assert "Department" not in headers
    assert "Complaint Date" not in headers
    assert "Mode" not in headers
    assert "Train Name For Report" not in headers
    ws = load_workbook(result.excel_path).active
    first_data = [str(ws.cell(row=3, column=c).value or "") for c in range(1, len(headers) + 1)]
    assert first_data[1] == "REF101"
    assert first_data[5] == "SC"
    assert first_data[8] == "Garbage on platform"
    assert "Status" not in headers


def test_excel_pdf_metadata_column_parity_all_reports(
    r1_paths: tuple[Path, Path],
    r2_paths: tuple[Path, Path],
    r5_csv: Path,
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    r1 = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert r1.success
    r1_headers = _headers_from_excel(r1.excel_path)
    assert r1.output_columns == r1_headers
    assert r1.visible_columns == r1_headers

    comp2, fb2 = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    r2 = Report2Processor().process(
        source_a_path=comp2,
        report_slug="report2",
        source_b_path=fb2,
    )
    assert r2.success
    r2_headers = _headers_from_excel(r2.excel_path)
    assert r2.output_columns == r2_headers
    assert r2.visible_columns == r2_headers

    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    r5 = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert r5.success
    r5_headers = _headers_from_excel(r5.excel_path)
    assert r5.output_columns == r5_headers
    assert r5.visible_columns == r5_headers

    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    r6 = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert r6.success
    r6_headers = _headers_from_excel(r6.excel_path)
    assert r6.output_columns == r6_headers
    assert r6.visible_columns == r6_headers


def test_excel_pdf_column_parity_r5(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    excel_headers = _headers_from_excel(result.excel_path)
    assert result.output_columns == excel_headers
    assert result.visible_columns == excel_headers
    assert result.output_columns == R5_HEADERS
    assert Path(result.pdf_path).stat().st_size > 0


def test_migrate_selected_column_keys_drops_mode_and_renames_legacy():
    migrated = migrate_selected_column_keys(
        ["refNo", "mode", "zone", "div", "complaintRefNo", "type"]
    )
    assert migrated == ["complaintRefNo", "zoneCode", "divCode", "complaintTypeName"]
    assert default_visible_column_keys(REPORT5_OUTPUT_COLUMNS) == [
        column.key for column in REPORT5_OUTPUT_COLUMNS
    ]


def test_reports_3_and_4_default_eleven_columns():
    assert len(R3_HEADERS) == 11
    assert R3_HEADERS[0] == "S.No."
    assert "Train Name" in R3_HEADERS
    assert "Average Rating" in R3_HEADERS
    assert len(R4_HEADERS) == 11
    assert R4_HEADERS[0] == "S.No."
    assert "Train Name" in R4_HEADERS
    assert "Average Rating" in R4_HEADERS
