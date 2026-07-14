"""Tests for Report 2 Phase 8 processor (Division Wise Top 25 dual-source)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.registry import PROCESSORS
from app.automation.processing.report2_processor import Report2Processor, TOP_N, HIDDEN_COLUMNS

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report2"


@pytest.fixture
def processor() -> Report2Processor:
    return Report2Processor()


@pytest.fixture
def comprehensive_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report2_division_comprehensive_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "division_comprehensive_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def feedback_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report2_division_feedback_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "division_feedback_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(
        "app.automation.processing.report2_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report2_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report2_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )


def test_registry_selects_report2_processor():
    assert "report2" in PROCESSORS
    assert PROCESSORS["report2"].processor_name == "report2_division_wise_processor"


def test_rejects_pdf_input(processor: Report2Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report2")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_fails_when_feedback_missing(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(source_a_path=comprehensive_csv, report_slug="report2")

    assert result.success is False
    # Error message now indicates explicit source_b_path is required (no fallback)
    assert "requires explicit source_b_path" in (result.error or "") or "Feedback" in (result.error or "")
    assert not list((tmp_path / "output" / "excel").rglob("*.xlsx"))


def test_produces_excel_and_pdf_outputs_with_feedback(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )

    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()
    assert result.source_a_rows is not None
    assert result.source_b_rows is not None
    assert result.source_b_rows > 0


def test_top_25_selection(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    data_rows = 0
    for row_idx in range(3, worksheet.max_row + 1):
        org_value = worksheet.cell(row=row_idx, column=2).value
        if org_value and "total" not in str(org_value).lower():
            data_rows += 1

    assert data_rows <= TOP_N


def test_columns_hidden(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    from openpyxl.utils import get_column_letter

    for col_idx in HIDDEN_COLUMNS:
        col_letter = get_column_letter(col_idx)
        if col_letter in worksheet.column_dimensions:
            assert worksheet.column_dimensions[col_letter].hidden is True


def test_scr_row_has_yellow_fill_and_black_text(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    scr_found = False
    for row_idx in range(3, worksheet.max_row + 1):
        org_value = worksheet.cell(row=row_idx, column=2).value
        if org_value and "south central railway" in str(org_value).lower():
            scr_found = True
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
                assert cell.font.color.rgb in {"00000000", "FF000000", "000000"}
    assert scr_found


def test_report_title_contains_bottom_25(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    title_cell = worksheet.cell(row=1, column=1).value
    assert "Bottom 25" in title_cell or "Report No 2" in title_cell


def test_descending_order_preserved(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify that rows are in descending order by Received column."""
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    received_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col_idx).value == "Received":
            received_col = col_idx
            break

    assert received_col is not None

    values = []
    for row_idx in range(3, worksheet.max_row):
        val = worksheet.cell(row=row_idx, column=received_col).value
        if val and str(val).strip().isdigit():
            values.append(int(val))

    for i in range(len(values) - 1):
        assert values[i] >= values[i + 1], f"Row {i+3} not in descending order"


def test_fails_without_explicit_source_b_path(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify processor fails when source_b_path is None (no fallback)."""
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=None,  # Explicitly None
    )

    assert result.success is False
    assert "requires explicit source_b_path" in (result.error or "")


def test_fails_when_source_b_file_missing(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify processor fails when source_b_path points to nonexistent file."""
    _patch_outputs(monkeypatch, tmp_path)

    nonexistent_path = tmp_path / "does_not_exist.csv"

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=nonexistent_path,
    )

    assert result.success is False
    assert "Source B file missing" in (result.error or "")


def test_output_contains_feedback_columns(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify final Excel output includes all Feedback columns."""
    from app.automation.processing.report2_processor import SOURCE_B_DATA_COLUMNS

    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    headers = [worksheet.cell(row=2, column=col).value for col in range(1, worksheet.max_column + 1)]

    for col in SOURCE_B_DATA_COLUMNS:
        assert col in headers, f"Missing feedback column: {col}"


def test_output_includes_modification_times(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify ProcessingResult includes source and output modification times."""
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True
    assert result.source_a_mtime is not None
    assert result.source_b_mtime is not None
    assert result.output_mtime is not None
    assert result.run_timestamp is not None


def test_output_filename_includes_timestamp(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify output filename includes run timestamp to prevent stale reuse."""
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    excel_name = Path(result.excel_path).name
    pdf_name = Path(result.pdf_path).name

    import re
    # Pattern: Rail_Madad_Report_2_Division_Wise_Bottom_25_DD-MM-YYYY_HHMMSS.xlsx
    timestamp_pattern = r"_\d{6}\.(xlsx|pdf)$"
    assert re.search(timestamp_pattern, excel_name), f"Excel filename missing timestamp: {excel_name}"
    assert re.search(timestamp_pattern, pdf_name), f"PDF filename missing timestamp: {pdf_name}"


def test_serial_numbers_are_sequential(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify S.No. column is regenerated as 1..N after merge."""
    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    sno_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col_idx).value == "S.No.":
            sno_col = col_idx
            break

    assert sno_col is not None

    serial_numbers = []
    for row_idx in range(3, worksheet.max_row):
        val = worksheet.cell(row=row_idx, column=sno_col).value
        if val and str(val).strip().isdigit():
            serial_numbers.append(int(val))

    expected = list(range(1, len(serial_numbers) + 1))
    assert serial_numbers == expected, f"S.No. not sequential: {serial_numbers}"


# -----------------------------------------------------------------------------
# Base-name matching tests
# -----------------------------------------------------------------------------


def test_extract_base_division_removes_railway_zone_suffix():
    """Verify _extract_base_division strips parenthetical railway zone suffix."""
    processor = Report2Processor()

    # Railway zone suffixes (Source A format)
    assert processor._extract_base_division("DELHI DIVISION (Northern Railway)") == "delhi division"
    assert processor._extract_base_division("LUCKNOW DIVISION (North Eastern Railway)") == "lucknow division"
    assert processor._extract_base_division("SECUNDERABAD DIVISION (South Central Railway)") == "secunderabad division"


def test_extract_base_division_removes_station_code_suffix():
    """Verify _extract_base_division strips parenthetical station code suffix."""
    processor = Report2Processor()

    # Station code suffixes (Source B format)
    assert processor._extract_base_division("DELHI DIVISION (DLI)") == "delhi division"
    assert processor._extract_base_division("LUCKNOW DIVISION (LJN)") == "lucknow division"
    assert processor._extract_base_division("MUMBAI DIVISION (CSTM)") == "mumbai division"


def test_extract_base_division_normalizes_divn_to_division():
    """Verify DIVN abbreviation is normalized to DIVISION."""
    processor = Report2Processor()

    assert processor._extract_base_division("JALANDHAR DIVN (Northern Railway)") == "jalandhar division"
    assert processor._extract_base_division("MUMBAI DIVN (CR)") == "mumbai division"


def test_extract_base_division_normalizes_ampersand():
    """Verify & is normalized to AND."""
    processor = Report2Processor()

    assert processor._extract_base_division("LUCKNOW & VARANASI (NER)") == "lucknow and varanasi"
    assert processor._extract_base_division("TEST&DIVISION") == "test and division"


def test_base_name_matching_different_suffixes(
    processor: Report2Processor,
    comprehensive_csv: Path,
    feedback_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify divisions match when Source A has railway zone and Source B has station code."""
    from app.automation.processing.report2_processor import SOURCE_B_DATA_COLUMNS

    _patch_outputs(monkeypatch, tmp_path)

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=feedback_csv,
    )
    assert result.success is True

    # Load output and verify feedback columns have data
    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    # Find Feedback Received column
    feedback_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col_idx).value == "Feedback Received":
            feedback_col = col_idx
            break

    assert feedback_col is not None, "Feedback Received column not found"

    # Check that at least some data rows have feedback values
    feedback_values = []
    for row_idx in range(3, min(worksheet.max_row, 10)):  # Check first several data rows
        val = worksheet.cell(row=row_idx, column=feedback_col).value
        if val and str(val).strip():
            feedback_values.append(val)

    assert len(feedback_values) > 0, "No Feedback Received values found - merge likely failed"


def test_shuffled_source_b_matches_correctly(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify Source B rows in random order still match correctly by name."""
    _patch_outputs(monkeypatch, tmp_path)

    # Create shuffled feedback CSV
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)
    shuffled_feedback = extracted / "shuffled_feedback.csv"

    # Shuffled order - not matching Source A order
    shuffled_feedback.write_text(
        """S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,Unsatisfactory,% Unsatisfactory
1,MUMBAI DIVISION (CSTM),320,47.06,240,65,15,4.69
2,DELHI DIVISION (DLI),420,49.41,320,85,15,3.57
3,CHENNAI DIVISION (MAS),290,46.77,215,60,15,5.17
4,LUCKNOW DIVISION (LJN),350,48.61,260,75,15,4.29
5,KOLKATA DIVISION (HWH),260,44.83,190,55,15,5.77
6,Total,1640,47.54,1185,350,105,6.40
""",
        encoding="utf-8",
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=shuffled_feedback,
    )
    assert result.success is True

    # Verify Delhi Division got the correct feedback (420, not Mumbai's 320)
    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    # Find Organisation and Feedback Received columns
    org_col = None
    feedback_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=2, column=col_idx).value
        if header == "Organisation":
            org_col = col_idx
        elif header == "Feedback Received":
            feedback_col = col_idx

    assert org_col is not None
    assert feedback_col is not None

    # Find Delhi Division row and verify feedback value
    for row_idx in range(3, worksheet.max_row + 1):
        org = worksheet.cell(row=row_idx, column=org_col).value
        if org and "DELHI" in str(org).upper():
            feedback = worksheet.cell(row=row_idx, column=feedback_col).value
            assert str(feedback) == "420", f"Delhi got wrong feedback: {feedback}"
            break


def test_unmatched_source_a_leaves_feedback_blank(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify unmatched Source A divisions have blank feedback, not shifted data."""
    _patch_outputs(monkeypatch, tmp_path)

    # Create feedback CSV missing some divisions
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)
    partial_feedback = extracted / "partial_feedback.csv"

    # Only includes Delhi and Mumbai - missing Lucknow, Chennai, etc.
    partial_feedback.write_text(
        """S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,Unsatisfactory,% Unsatisfactory
1,DELHI DIVISION (DLI),420,49.41,320,85,15,3.57
2,MUMBAI DIVISION (CSTM),320,47.06,240,65,15,4.69
3,Total,740,48.24,560,150,30,4.05
""",
        encoding="utf-8",
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=partial_feedback,
    )
    assert result.success is True  # Should still succeed with partial matches

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    # Find Organisation and Feedback Received columns
    org_col = None
    feedback_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=2, column=col_idx).value
        if header == "Organisation":
            org_col = col_idx
        elif header == "Feedback Received":
            feedback_col = col_idx

    # Verify Lucknow Division has blank feedback (not Delhi's or Mumbai's)
    for row_idx in range(3, worksheet.max_row + 1):
        org = worksheet.cell(row=row_idx, column=org_col).value
        if org and "LUCKNOW" in str(org).upper():
            feedback = worksheet.cell(row=row_idx, column=feedback_col).value
            assert feedback is None or str(feedback).strip() == "", \
                f"Lucknow should have blank feedback, got: {feedback}"
            break


def test_fails_when_no_divisions_match(
    processor: Report2Processor,
    comprehensive_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify processing fails when Source A and Source B have no matching divisions."""
    _patch_outputs(monkeypatch, tmp_path)

    # Create feedback CSV with completely different divisions
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)
    no_match_feedback = extracted / "no_match_feedback.csv"

    no_match_feedback.write_text(
        """S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,Unsatisfactory,% Unsatisfactory
1,UNKNOWN DIVISION (XYZ),100,50.00,70,20,10,10.00
2,FAKE DIVISION (ABC),200,40.00,140,40,20,10.00
3,Total,300,45.00,210,60,30,10.00
""",
        encoding="utf-8",
    )

    result = processor.process(
        source_a_path=comprehensive_csv,
        report_slug="report2",
        source_b_path=no_match_feedback,
    )

    assert result.success is False
    assert "REPORT2_MERGE_FAILED" in (result.error or "")
    assert "No divisions matched" in (result.error or "")


def test_fails_when_all_feedback_blank(
    processor: Report2Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify processing fails when all feedback columns are blank after merge."""
    _patch_outputs(monkeypatch, tmp_path)

    # Create comprehensive CSV with matching divisions
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)

    comprehensive = extracted / "comprehensive.csv"
    comprehensive.write_text(
        """S.No.,Organisation,Opening Balance,Received,% Share,Closed,Closing Balance
1,TEST DIVISION (Railway),10,100,50.00,90,20
2,Total,,100,,90,
""",
        encoding="utf-8",
    )

    # Create feedback CSV with matching division but empty feedback values
    blank_feedback = extracted / "blank_feedback.csv"
    blank_feedback.write_text(
        """S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,Unsatisfactory,% Unsatisfactory
1,TEST DIVISION (TST),,,,,,
2,Total,,,,,,
""",
        encoding="utf-8",
    )

    result = processor.process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=blank_feedback,
    )

    assert result.success is False
    assert "REPORT2_MERGE_FAILED" in (result.error or "")
    assert "blank" in (result.error or "").lower()


def test_divn_normalizes_to_match_division(
    processor: Report2Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """Verify DIVN in one source matches DIVISION in the other."""
    _patch_outputs(monkeypatch, tmp_path)

    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True, exist_ok=True)

    # Source A uses DIVN
    comprehensive = extracted / "comprehensive_divn.csv"
    comprehensive.write_text(
        """S.No.,Organisation,Opening Balance,Received,% Share,Closed,Closing Balance
1,JALANDHAR DIVN (Northern Railway),10,100,50.00,90,20
2,Total,,100,,90,
""",
        encoding="utf-8",
    )

    # Source B uses DIVISION
    feedback = extracted / "feedback_division.csv"
    feedback.write_text(
        """S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,Unsatisfactory,% Unsatisfactory
1,JALANDHAR DIVISION (JUC),50,50.00,35,10,5,10.00
2,Total,50,50.00,35,10,5,10.00
""",
        encoding="utf-8",
    )

    result = processor.process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )

    assert result.success is True

    # Verify feedback was merged
    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    feedback_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col_idx).value == "Feedback Received":
            feedback_col = col_idx
            break

    # Row 3 should have the feedback value
    feedback_val = worksheet.cell(row=3, column=feedback_col).value
    assert str(feedback_val) == "50", f"DIVN/DIVISION mismatch - got: {feedback_val}"
