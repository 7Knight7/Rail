"""Tests for Report 5 Phase 8 processor (SCR Train Mode Unsatisfactory Feedback)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.registry import PROCESSORS
from app.automation.processing.output_columns import REPORT5_VISIBLE_LABELS
from app.automation.processing.report5_processor import Report5Processor

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report5"


@pytest.fixture
def processor() -> Report5Processor:
    return Report5Processor()


@pytest.fixture
def train_complaints_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_registry_selects_report5_processor():
    assert "report5" in PROCESSORS
    assert PROCESSORS["report5"].processor_name == "report5_scr_train_processor"


def test_rejects_pdf_input(processor: Report5Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report5")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_produces_excel_and_pdf_outputs(
    processor: Report5Processor,
    train_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=train_complaints_csv, report_slug="report5")

    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()


def test_filters_train_mode_only(
    processor: Report5Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True, exist_ok=True)
    
    mixed_csv = extracted / "mixed_complaints.csv"
    mixed_csv.write_text(
        "Ref. No.,Mode,Registration Date,Train/Station,Type,Sub Type,Department,Status,Zone,Div,"
        "feedbackRemark,trainNameForReport/Station Name,complaintDesc,remarks,userId\n"
        "REF001,Train,15-07-26,Train A,Type1,SubType1,Dept1,Pending,SC,HYB,,Train A,Desc1,Rem1,u1\n"
        "REF002,Station,15-07-26,Station A,Type2,SubType2,Dept2,Pending,SC,HYB,,Station A,Desc2,Rem2,u2\n"
        "REF003,Train,15-07-26,Train B,Type3,SubType3,Dept3,Pending,SC,HYB,,Train B,Desc3,Rem3,u3\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=mixed_csv, report_slug="report5")

    assert result.success is True
    assert result.processed_row_count == 2


def test_fails_if_no_train_rows(
    processor: Report5Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True, exist_ok=True)
    
    station_only_csv = extracted / "station_only.csv"
    station_only_csv.write_text(
        "Ref. No.,Mode,Registration Date,Train/Station,Type,Sub Type,Department,Status,Zone,Div,"
        "feedbackRemark,trainNameForReport/Station Name,complaintDesc,remarks,userId\n"
        "REF001,Station,15-07-26,Station A,Type1,SubType1,Dept1,Pending,SC,HYB,,Stn A,Desc,Rem,u1\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=station_only_csv, report_slug="report5")

    assert result.success is False
    assert "Train" in (result.error or "")


def test_scr_row_not_highlighted(
    processor: Report5Processor,
    train_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.automation.formatting.scr import cell_has_yellow_fill

    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=train_complaints_csv, report_slug="report5")
    assert result.success is True

    worksheet = load_workbook(result.excel_path).active
    for row_idx in range(3, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            assert not cell_has_yellow_fill(cell)


def test_output_headers_match_spec(
    processor: Report5Processor,
    train_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.Report5Processor._find_template",
        lambda self: None,
    )

    result = processor.process(source_a_path=train_complaints_csv, report_slug="report5")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    for col_idx, expected in enumerate(REPORT5_VISIBLE_LABELS, start=1):
        actual = worksheet.cell(row=2, column=col_idx).value
        assert actual == expected


def test_report_title_contains_train(
    processor: Report5Processor,
    train_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.Report5Processor._find_template",
        lambda self: None,
    )

    result = processor.process(source_a_path=train_complaints_csv, report_slug="report5")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    title = worksheet.cell(row=1, column=1).value
    
    assert "Train" in title or "Report No 5" in title


def test_report5_pdf_survives_tall_text_cell(
    processor: Report5Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    import csv

    long_text = "Word " * 500
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True, exist_ok=True)
    csv_path = extracted / "report5_complaints_raw.csv"
    with (FIXTURES_DIR / "train_complaints_raw.csv").open(encoding="utf-8", newline="") as src:
        rows = list(csv.DictReader(src))
    rows[0]["complaintDesc"] = long_text
    rows[0]["remarks"] = long_text
    rows[0]["feedbackRemark"] = long_text
    with csv_path.open("w", encoding="utf-8", newline="") as dst:
        writer = csv.DictWriter(dst, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report5_processor.Report5Processor._find_template",
        lambda self: None,
    )

    result = processor.process(source_a_path=csv_path, report_slug="report5")
    assert result.success is True, result.error
    pdf_path = Path(result.pdf_path or "")
    assert pdf_path.is_file()
    assert pdf_path.read_bytes()[:5] == b"%PDF-"
    assert result.processed_row_count == len(rows)
    assert len(result.output_columns or []) == 13
