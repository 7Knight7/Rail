"""Tests that manual column selection reaches Report 1/2 generation output."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.report1_processor import Report1Processor
from app.automation.processing.report2_processor import Report2Processor

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"
FIXTURES_R2 = Path(__file__).resolve().parent.parent / "fixtures" / "report2"


@pytest.fixture
def r1_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comp = extracted / "comprehensive.csv"
    fb = extracted / "feedback.csv"
    comp.write_text(
        (FIXTURES_R1 / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    fb.write_text(
        (FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comp, fb


@pytest.fixture
def r2_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "division"
    extracted.mkdir(parents=True)
    comp = extracted / "comprehensive.csv"
    fb = extracted / "feedback.csv"
    comp.write_text(
        (FIXTURES_R2 / "division_comprehensive_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    fb.write_text(
        (FIXTURES_R2 / "division_feedback_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comp, fb


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def test_report1_manual_snapshot_without_run_context(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    subset = [
        "report1.source_a.organisation",
        "report1.source_a.received",
        "report1.source_a.avg_disposal_time",
        "report1.source_b.feedback_received",
    ]
    column_selection = {
        "report_slug": "report1",
        "selected_column_ids": subset,
        "column_order": subset,
        "configuration_source": "manual_snapshot",
    }
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
        column_selection=column_selection,
    )
    assert result.success is True
    assert result.configuration_source == "manual_snapshot"
    assert result.selected_column_ids == subset
    headers = [
        str(load_workbook(result.excel_path).active.cell(row=2, column=c).value or "")
        for c in range(1, 5)
    ]
    assert headers == [
        "Organisation",
        "Received",
        "Avg. Disposal Time",
        "Feedback Received",
    ]


def test_report2_manual_snapshot_without_run_context(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    subset = [
        "division.source_a.division",
        "division.source_a.received",
        "division.source_a.percent_disposal",
        "division.source_b.unsatisfactory",
    ]
    column_selection = {
        "report_slug": "division",
        "selected_column_ids": subset,
        "column_order": subset,
        "configuration_source": "manual_snapshot",
    }
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="division",
        source_b_path=feedback,
        column_selection=column_selection,
    )
    assert result.success is True
    assert result.configuration_source == "manual_snapshot"
    headers = [
        str(load_workbook(result.excel_path).active.cell(row=2, column=c).value or "")
        for c in range(1, 5)
    ]
    assert headers == ["Division", "Received", "% Disposal", "Unsatisfactory"]


def test_report2_five_column_manual_acceptance_headers(
    r2_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    subset = [
        "division.source_a.division",
        "division.source_a.received",
        "division.source_a.percent_disposal",
        "division.source_b.feedback_received",
        "division.source_b.unsatisfactory",
    ]
    column_selection = {
        "report_slug": "division",
        "selected_column_ids": subset,
        "column_order": subset,
        "configuration_source": "manual_snapshot",
    }
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="division",
        source_b_path=feedback,
        column_selection=column_selection,
    )
    assert result.success is True
    assert result.configuration_source == "manual_snapshot"
    headers = [
        str(load_workbook(result.excel_path).active.cell(row=2, column=c).value or "")
        for c in range(1, 6)
    ]
    assert headers == [
        "Division",
        "Received",
        "% Disposal",
        "Feedback Received",
        "Unsatisfactory",
    ]
