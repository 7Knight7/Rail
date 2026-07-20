"""Tests for previous-day report_date, sequential S.No., and Report 1/2 PDF fit."""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A2, A3, A4, landscape

from app.automation.formatting.pdf_table import build_fitted_table, choose_landscape_layout
from app.automation.formatting.serial import apply_serial_number, renumber_data_rows
from app.automation.processing.report1_processor import Report1Processor
from app.automation.processing.report2_processor import Report2Processor
from app.automation.utils import artifact_filename_timestamp, previous_day_report_date

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"


def _expected_report_date(now: datetime | None = None) -> str:
    moment = now or datetime.now()
    return (moment - timedelta(days=1)).strftime("%d-%m-%Y")


def _monkeypatch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _pdf_mediabox_size(pdf_path: Path) -> tuple[float, float]:
    """Parse the first MediaBox width/height from a simple ReportLab PDF."""
    raw = pdf_path.read_bytes().decode("latin-1", errors="ignore")
    match = re.search(r"/MediaBox\s*\[\s*([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s*\]", raw)
    assert match, "PDF MediaBox not found"
    x0, y0, x1, y1 = (float(match.group(i)) for i in range(1, 5))
    return x1 - x0, y1 - y0


def test_previous_day_report_date_is_yesterday():
    fixed = datetime(2026, 7, 13, 15, 30, 0)
    assert previous_day_report_date(now=fixed) == "12-07-2026"
    assert previous_day_report_date(now=fixed) != fixed.strftime("%d-%m-%Y")


def test_artifact_filename_timestamp_uses_previous_day_date():
    fixed = datetime(2026, 7, 13, 21, 5, 9)
    stamp = artifact_filename_timestamp(now=fixed)
    assert stamp.startswith("2026-07-12_")
    assert stamp.endswith("21-05-09")


def test_renumber_data_rows_is_sequential_and_skips_total():
    headers = ["S.No.", "Organisation", "Received"]
    rows = [
        ["20", "A", "100"],
        ["54", "B", "90"],
        ["41", "Total", "190"],
    ]
    out = renumber_data_rows(headers, rows)
    assert [r[0] for r in out[:-1]] == ["1", "2"]
    assert out[-1][0] == ""
    assert out[0][1] == "A"
    assert out[1][1] == "B"


def test_apply_serial_number_preserves_identifiers():
    headers = ["S.No.", "Train No.", "Ref. No."]
    values = ["99", "12760", "RM-1"]
    assert apply_serial_number(headers, values, 3) == ["3", "12760", "RM-1"]
    assert apply_serial_number(headers, values, None) == ["", "12760", "RM-1"]


def test_wide_table_falls_back_to_a3_without_overflow():
    headers = [f"Col{i}" for i in range(20)]
    row = [f"value-{i}-xxxxxxxx" for i in range(20)]
    table_data = [headers, row]
    pagesize, col_widths, _font, margin = choose_landscape_layout(table_data)
    assert pagesize in {landscape(A4), landscape(A3), landscape(A2)}
    usable = pagesize[0] - (2 * margin)
    assert sum(col_widths) <= usable + 0.5
    table, chosen, used_margin = build_fitted_table(
        table_data,
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
    )
    wrapped_w, _ = table.wrap(chosen[0] - 2 * used_margin, chosen[1])
    assert wrapped_w <= chosen[0] - 2 * used_margin + 1.0


@pytest.fixture
def r1_inputs(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "report1_comprehensive_zone_raw.csv"
    feedback = extracted / "report1_feedback_zone_raw.csv"
    comprehensive.write_text(
        "S.No.,Organisation,Opening Balance,Received,% Share,Closed,Closing Balance,"
        "% Disposal,Avg. Disposal Time,Avg. Rating,Avg. Pendency Time,Forwarded,Avg. FRT\n"
        "20,Northern Railway,1,10,1,9,1,90,0:10,Satisfactory,0:05,10,0:01\n"
        "54,South Central Railway,1,9,1,8,1,88,0:10,Satisfactory,0:05,9,0:01\n"
        "41,Irctc-Catering,1,8,1,7,1,87,0:10,Satisfactory,0:05,8,0:01\n"
        "42,Irctc-Online,1,7,1,6,1,86,0:10,Satisfactory,0:05,7,0:01\n"
        ",Total,,34,,30,,,,,,,\n",
        encoding="utf-8",
    )
    feedback.write_text(
        (FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comprehensive, feedback


def test_report1_title_and_filename_use_previous_day(
    r1_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    expected = _expected_report_date()
    today = datetime.now().strftime("%d-%m-%Y")

    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    assert expected in Path(result.excel_path).name
    assert expected in Path(result.pdf_path).name
    if expected != today:
        assert today not in Path(result.pdf_path).name

    workbook = load_workbook(result.excel_path)
    title = str(workbook.active.cell(row=1, column=1).value or "")
    assert expected in title
    if expected != today:
        assert today not in title


def test_report1_pdf_fits_width_and_shows_edge_columns(
    r1_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)

    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True

    width, height = _pdf_mediabox_size(Path(result.pdf_path))
    assert width > height  # landscape
    assert width >= landscape(A4)[0] - 1.0
    assert abs(width - landscape(A4)[0]) < 1.0 or abs(width - landscape(A3)[0]) < 1.0 or abs(width - landscape(A2)[0]) < 1.0

    workbook = load_workbook(result.excel_path)
    ws = workbook.active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    visible_indices = list(range(1, len(headers) + 1))
    visible_headers = headers
    assert visible_headers[0] == "S.No."
    assert visible_headers[-1] == "% Unsatisfactory"

    first_vals = [str(ws.cell(row=r, column=1).value or "") for r in range(3, ws.max_row)]
    last_col = len(headers)
    last_vals = [str(ws.cell(row=r, column=last_col).value or "") for r in range(3, ws.max_row)]
    assert any(v.isdigit() for v in first_vals)
    assert any(v.strip() for v in last_vals)

    data_rows = []
    for row_idx in range(3, ws.max_row + 1):
        data_rows.append([str(ws.cell(row=row_idx, column=i).value or "") for i in visible_indices])
    table_data = [visible_headers, *data_rows]
    table, pagesize, margin = build_fitted_table(
        table_data,
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
    )
    usable = pagesize[0] - 2 * margin
    wrapped_w, _ = table.wrap(usable, pagesize[1])
    assert wrapped_w <= usable + 1.0
    assert pagesize[0] > pagesize[1]


def test_report1_sno_regenerated_sequentially(
    r1_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)

    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    workbook = load_workbook(result.excel_path)
    ws = workbook.active
    serials: list[str] = []
    for row_idx in range(3, ws.max_row + 1):
        org = str(ws.cell(row=row_idx, column=2).value or "")
        sno = str(ws.cell(row=row_idx, column=1).value or "")
        if "total" in org.lower():
            assert sno in {"", "None"} or not sno.isdigit()
            continue
        serials.append(sno)
    assert serials == [str(i) for i in range(1, len(serials) + 1)]
    assert "20" not in serials
    assert "54" not in serials


@pytest.fixture
def r2_inputs(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report2"
    extracted.mkdir(parents=True)
    target = extracted / "report2_division_comprehensive_raw.csv"
    feedback = extracted / "report2_division_feedback_raw.csv"
    lines = [
        "S.No.,Organisation,Opening Balance,Received,% Share,Closed,Closing Balance,"
        "% Disposal,Avg. Disposal Time,Avg. Rating,Avg. Pendency Time,Forwarded,Avg. FRT"
    ]
    feedback_lines = [
        "S.No.,Organisation,Feedback Received,% Feedback,Excellent,Satisfactory,"
        "Unsatisfactory,% Unsatisfactory"
    ]
    portal_snos = [20, 54, 41, 42, 33, 18, 77, 12, 9, 100]
    for i in range(30):
        sno = portal_snos[i % len(portal_snos)] + i
        received = 1000 - i
        org = "South Central Railway Guntakal Division" if i == 2 else f"Division {i+1}"
        lines.append(
            f"{sno},{org},1,{received},1,1,1,90,0:10,Satisfactory,0:05,1,0:01"
        )
        feedback_lines.append(f"{i+1},{org},{received // 2},40,10,5,2,4.0")
    lines.append(",Total,,15000,,30,,,,,,,")
    feedback_lines.append(",Total,5000,,,,,,")
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")
    feedback.write_text("\n".join(feedback_lines) + "\n", encoding="utf-8")
    return target, feedback


def test_report2_sno_is_one_through_n_after_top25(
    r2_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    ws = workbook.active
    serials: list[int] = []
    orgs: list[str] = []
    for row_idx in range(3, ws.max_row + 1):
        org = str(ws.cell(row=row_idx, column=2).value or "")
        sno = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if "total" in org.lower():
            assert sno == ""
            continue
        serials.append(int(sno))
        orgs.append(org)

    assert serials == list(range(1, len(serials) + 1))
    assert serials[0] == 1
    assert serials[-1] == len(serials)
    assert len(serials) == 25
    assert orgs[0] == "Division 1"
    assert "South Central Railway" in orgs[2]


def test_report2_title_and_filename_use_previous_day(
    r2_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    expected = _expected_report_date()
    today = datetime.now().strftime("%d-%m-%Y")

    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    assert expected in Path(result.excel_path).name
    assert expected in Path(result.pdf_path).name
    if expected != today:
        assert today not in Path(result.pdf_path).name

    workbook = load_workbook(result.excel_path)
    title = str(workbook.active.cell(row=1, column=1).value or "")
    assert expected in title
    if expected != today:
        assert today not in title


@pytest.mark.parametrize(
    ("module", "processor_cls", "slug", "fixture_builder"),
    [
        (
            "app.automation.processing.report1_processor",
            Report1Processor,
            "report1",
            "r1",
        ),
        (
            "app.automation.processing.report2_processor",
            Report2Processor,
            "report2",
            "r2",
        ),
    ],
)
def test_processors_never_use_today_in_final_names(
    module: str,
    processor_cls: type,
    slug: str,
    fixture_builder: str,
    request: pytest.FixtureRequest,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _monkeypatch_outputs(monkeypatch, module, tmp_path)
    expected = _expected_report_date()
    today = datetime.now().strftime("%d-%m-%Y")
    if fixture_builder == "r1":
        comprehensive, feedback = request.getfixturevalue("r1_inputs")
        result = processor_cls().process(
            source_a_path=comprehensive,
            report_slug=slug,
            source_b_path=feedback,
        )
    else:
        comprehensive, feedback = request.getfixturevalue("r2_inputs")
        result = processor_cls().process(
            source_a_path=comprehensive,
            report_slug=slug,
            source_b_path=feedback,
        )
    assert result.success is True
    for path in (result.excel_path, result.pdf_path):
        name = Path(path).name
        assert expected in name
        if expected != today:
            assert today not in name


def test_report1_pdf_totals_match_excel(
    r1_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    closed_col = headers.index("Closed") + 1
    closed_total = str(ws.cell(row=ws.max_row, column=closed_col).value or "")
    assert closed_total == "30"
    assert str(ws.cell(row=ws.max_row, column=2).value or "") == "Total"
    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"
    assert len(pdf_bytes) > 100


def test_report2_pdf_totals_match_excel(
    r2_inputs: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r2_inputs
    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report2_processor", tmp_path)
    result = Report2Processor().process(
        source_a_path=comprehensive,
        report_slug="report2",
        source_b_path=feedback,
    )
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    received_col = headers.index("Received") + 1
    received_total = str(ws.cell(row=ws.max_row, column=received_col).value or "")
    assert received_total == "24700"
    assert "% Disposal" not in headers
    assert "Total" in str(ws.cell(row=ws.max_row, column=2).value or "")
    pdf_bytes = Path(result.pdf_path).read_bytes()
    assert pdf_bytes[:5] == b"%PDF-"
    assert len(pdf_bytes) > 100


FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def _count_pdf_pages(pdf_path: Path) -> int:
    raw = pdf_path.read_bytes().decode("latin-1", errors="ignore")
    return raw.count("/Type /Page") - raw.count("/Type /Pages")


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report5_pdf_fits_width_and_shows_edge_columns(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.automation.processing.report5_processor import Report5Processor
    from app.automation.formatting.pdf_table import build_wrapped_fitted_table

    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True

    width, height = _pdf_mediabox_size(Path(result.pdf_path))
    assert width > height

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert headers[0] == "S.No."
    assert headers[-1] == "User ID"
    assert len(headers) == 13
    assert "Status" not in headers

    data_rows = []
    for row_idx in range(3, ws.max_row + 1):
        data_rows.append([str(ws.cell(row=row_idx, column=i).value or "") for i in range(1, len(headers) + 1)])
    table_data = [headers, *data_rows]
    table, pagesize, margin = build_wrapped_fitted_table(
        table_data,
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
    )
    usable = pagesize[0] - 2 * margin
    wrapped_w, _ = table.wrap(usable, pagesize[1])
    assert wrapped_w <= usable + 1.0


def test_report6_pdf_fits_width_and_shows_edge_columns(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from app.automation.processing.report6_processor import Report6Processor
    from app.automation.formatting.pdf_table import build_wrapped_fitted_table

    _monkeypatch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    result = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    headers = [str(ws.cell(row=2, column=c).value or "") for c in range(1, ws.max_column + 1)]
    assert headers[0] == "S.No."
    assert headers[-1] == "User ID"
    assert len(headers) == 11
    assert "Status" not in headers
    assert "Created On" not in headers

    data_rows = []
    for row_idx in range(3, ws.max_row + 1):
        data_rows.append([str(ws.cell(row=row_idx, column=i).value or "") for i in range(1, len(headers) + 1)])
    table, pagesize, margin = build_wrapped_fitted_table(
        [headers, *data_rows],
        [("GRID", (0, 0), (-1, -1), 0.5, colors.black)],
    )
    usable = pagesize[0] - 2 * margin
    wrapped_w, _ = table.wrap(usable, pagesize[1])
    assert wrapped_w <= usable + 1.0
    assert Path(result.pdf_path).read_bytes()[:5] == b"%PDF-"
