"""Live acceptance checklist for Report 3 and Report 4 column filters (fixture-based)."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.report3_processor import Report3Processor
from app.automation.processing.report4_processor import Report4Processor
from app.automation.report4_filters import COMPLAINT_TYPES_ORDERED, get_type_configs

FIXTURES_R3 = Path(__file__).resolve().parent.parent / "fixtures" / "report3"
FIXTURES_R4 = Path(__file__).resolve().parent.parent / "fixtures" / "report4"

FOUR_COLUMN_SELECTION = {
    "selected_column_ids": [
        "train-no.train_name",
        "train-no.train_no",
        "train-no.received",
        "train-no.pending",
    ],
    "column_order": [
        "train-no.train_name",
        "train-no.train_no",
        "train-no.received",
        "train-no.pending",
    ],
    "configuration_source": "manual_snapshot",
}

FIVE_COLUMN_TYPES = {
    "selected_column_ids": [
        "types.owning_zone",
        "types.owning_division",
        "types.closed",
        "types.percent_closed",
        "types.average_rating",
    ],
    "column_order": [
        "types.owning_zone",
        "types.owning_division",
        "types.closed",
        "types.percent_closed",
        "types.average_rating",
    ],
    "configuration_source": "manual_snapshot",
}


def _write_r4_index(base: Path) -> Path:
    index_path = base / "types_combined_index.csv"
    rows = []
    for type_config in get_type_configs():
        slug = type_config.name.lower().replace(" ", "_").replace("&", "and")
        csv_path = base / f"report4_{slug}_raw.csv"
        csv_path.write_text(
            (FIXTURES_R4 / "security_raw.csv").read_text(encoding="utf-8"),
            encoding="utf-8",
        )
        rows.append(
            {
                "type_name": type_config.name,
                "csv_path": str(csv_path),
                "row_count": "10",
                "status": "success",
                "error": "",
            }
        )
    with index_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["type_name", "csv_path", "row_count", "status", "error"],
        )
        writer.writeheader()
        writer.writerows(rows)
    return index_path


@pytest.fixture
def train_csv(tmp_path: Path) -> Path:
    target = tmp_path / "train-no.csv"
    target.write_text(
        (FIXTURES_R3 / "trainwise_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report3_four_column_acceptance(train_csv: Path, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )
    expected_headers = ["Train Name", "Train No.", "Received", "Pending"]
    headers, rows, _, keys, _ = Report3Processor.build_projected_table(
        train_csv,
        column_selection=FOUR_COLUMN_SELECTION,
    )
    assert headers == expected_headers
    assert len(rows) == 20
    assert keys == FOUR_COLUMN_SELECTION["selected_column_ids"]

    result = Report3Processor().process(
        source_a_path=train_csv,
        report_slug="train-no",
        column_selection=FOUR_COLUMN_SELECTION,
    )
    assert result.success is True, result.error
    assert result.output_columns == expected_headers
    ws = load_workbook(result.excel_path).active
    excel_headers = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert excel_headers == expected_headers
    assert Path(result.pdf_path).read_bytes()[:4] == b"%PDF"


def test_report4_five_column_acceptance(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report4_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )
    index_path = _write_r4_index(tmp_path)
    expected_headers = [
        "Owning Zone",
        "Owning Division",
        "Closed",
        "% Closed",
        "Average Rating",
    ]
    sections, headers, keys, _ = Report4Processor.build_projected_sections(
        index_path,
        column_selection=FIVE_COLUMN_TYPES,
    )
    assert headers == expected_headers
    assert keys == FIVE_COLUMN_TYPES["selected_column_ids"]
    assert len(sections) == len(COMPLAINT_TYPES_ORDERED)
    for section in sections:
        assert section.headers == expected_headers

    result = Report4Processor().process(
        source_a_path=index_path,
        report_slug="types",
        column_selection=FIVE_COLUMN_TYPES,
    )
    assert result.success is True, result.error
    assert result.output_columns == expected_headers
