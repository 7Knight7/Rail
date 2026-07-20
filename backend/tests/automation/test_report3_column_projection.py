"""Report 3 column projection tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.report3_processor import Report3Processor

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report3"

SELECTED_FOUR = [
    "train-no.train_name",
    "train-no.train_no",
    "train-no.received",
    "train-no.pending",
]


@pytest.fixture
def train_csv(tmp_path: Path) -> Path:
    target = tmp_path / "train-no.csv"
    target.write_text(
        (FIXTURES_DIR / "trainwise_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_default_projection_has_eleven_columns(train_csv: Path, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )
    result = Report3Processor().process(source_a_path=train_csv, report_slug="train-no")
    assert result.success is True, result.error
    assert len(result.output_columns or []) == 11
    assert result.processed_row_count == 20


def test_selected_four_columns_in_excel_and_pdf(train_csv: Path, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "pdf"),
    )
    selection = {
        "selected_column_ids": SELECTED_FOUR,
        "column_order": SELECTED_FOUR,
        "configuration_source": "manual_snapshot",
    }
    result = Report3Processor().process(
        source_a_path=train_csv,
        report_slug="train-no",
        column_selection=selection,
    )
    assert result.success is True, result.error
    assert result.output_columns == ["Train Name", "Train No.", "Received", "Pending"]
    ws = load_workbook(result.excel_path).active
    headers = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert headers == result.output_columns
    assert ws.max_row == 22  # title + header + 20 data rows
    assert Path(result.pdf_path).read_bytes()[:4] == b"%PDF"


def test_top20_order_preserved_with_projection(train_csv: Path):
    processor = Report3Processor()
    rows, _ = processor._read_csv(train_csv)
    data_rows, _ = processor._split_total_row(rows)
    top20 = data_rows[:20]
    canonical = Report3Processor.build_canonical_rows(top20)
    _, projected, _, keys, _ = Report3Processor.build_projected_table(
        train_csv,
        selected_keys=["train-no.train_name"],
    )
    assert len(projected) == 20
    assert projected[0][0] == canonical[0]["trainName"]
    assert projected[-1][0] == canonical[19]["trainName"]
