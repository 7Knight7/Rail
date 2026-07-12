"""Tests for Report 3 Phase 8 processor (Top 20 Trains)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.registry import PROCESSORS
from app.automation.processing.report3_processor import Report3Processor, TOP_N

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report3"


@pytest.fixture
def processor() -> Report3Processor:
    return Report3Processor()


@pytest.fixture
def trainwise_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report3"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report3_trainwise_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "trainwise_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_registry_selects_report3_processor():
    assert "report3" in PROCESSORS
    assert PROCESSORS["report3"].processor_name == "report3_top20_trains_processor"


def test_rejects_pdf_input(processor: Report3Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report3")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_produces_excel_and_pdf_outputs(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")

    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()
    assert Path(result.excel_path).stat().st_size > 0
    assert Path(result.pdf_path).stat().st_size > 0


def test_top_20_selection(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    data_rows = 0
    for row_idx in range(3, worksheet.max_row + 1):
        train_no = worksheet.cell(row=row_idx, column=2).value
        if train_no and "total" not in str(train_no).lower():
            data_rows += 1

    assert data_rows <= TOP_N
    assert data_rows == TOP_N


def test_train_no_preserved_as_text(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    for row_idx in range(3, min(worksheet.max_row + 1, 8)):
        train_no_cell = worksheet.cell(row=row_idx, column=2)
        assert train_no_cell.number_format == "@"


def test_scr_row_has_yellow_fill_and_black_text(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    scr_found = False
    for row_idx in range(3, worksheet.max_row + 1):
        owning_zone = worksheet.cell(row=row_idx, column=4).value
        if owning_zone and "south central railway" in str(owning_zone).lower():
            scr_found = True
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
                assert cell.font.color.rgb in {"00000000", "FF000000", "000000"}
    assert scr_found


def test_report_title_contains_top_20(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    title_cell = worksheet.cell(row=1, column=1).value
    assert "20 Trains" in title_cell or "Report No 3" in title_cell


def test_output_columns_match_spec(
    processor: Report3Processor,
    trainwise_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report3_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=trainwise_csv, report_slug="report3")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    expected_headers = ["S.No.", "Train No.", "Train Name", "Owning Zone", "Owning Division", "Received"]
    for col_idx, expected in enumerate(expected_headers, start=1):
        actual = worksheet.cell(row=2, column=col_idx).value
        assert actual == expected
