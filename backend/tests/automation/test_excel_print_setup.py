"""Tests for Excel print setup on report workbooks."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.report1_processor import Report1Processor
from app.automation.processing.report5_processor import Report5Processor

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"
FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


@pytest.fixture
def r1_paths(tmp_path: Path) -> tuple[Path, Path]:
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "report1_comprehensive_zone_raw.csv"
    feedback = extracted / "report1_feedback_zone_raw.csv"
    comprehensive.write_text(
        (FIXTURES_R1 / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    feedback.write_text(
        (FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return comprehensive, feedback


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report1_excel_print_setup(
    r1_paths: tuple[Path, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    comprehensive, feedback = r1_paths
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_area is not None
    assert ws.print_area.endswith(str(ws.max_row))
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True


def test_report5_excel_print_setup(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    ws = load_workbook(result.excel_path).active
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.print_area is not None
    assert "$A$1" in ws.print_area
