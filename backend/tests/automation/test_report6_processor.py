"""Tests for Report 6 SCR Phase 8 processor (Station Mode Unsatisfactory Feedback)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.processing.registry import PROCESSORS
from app.automation.processing.report6_processor import Report6Processor, OUTPUT_HEADERS

FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


@pytest.fixture
def processor() -> Report6Processor:
    return Report6Processor()


@pytest.fixture
def station_complaints_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True, exist_ok=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_DIR / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_registry_selects_report6_processor():
    assert "report6_station" in PROCESSORS
    assert PROCESSORS["report6_station"].processor_name == "report6_scr_station_processor"


def test_rejects_pdf_input(processor: Report6Processor, tmp_path: Path):
    pdf_path = tmp_path / "input.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fake")

    result = processor.process(source_a_path=pdf_path, report_slug="report6_station")

    assert result.success is False
    assert "PDF" in (result.error or "")


def test_produces_excel_and_pdf_outputs(
    processor: Report6Processor,
    station_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=station_complaints_csv, report_slug="report6_station")

    assert result.success is True
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()


def test_filters_station_mode_only(
    processor: Report6Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True, exist_ok=True)
    
    mixed_csv = extracted / "mixed_complaints.csv"
    mixed_csv.write_text(
        "Ref. No.,Complaint Date,Train/Station,Mode,Type,Sub Type,Department,Status\n"
        "REF001,2026-07-10,Train A,Train,Type1,SubType1,Dept1,Pending\n"
        "REF002,2026-07-10,Station A,Station,Type2,SubType2,Dept2,Pending\n"
        "REF003,2026-07-10,Station B,Station,Type3,SubType3,Dept3,Pending\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=mixed_csv, report_slug="report6_station")

    assert result.success is True
    assert result.processed_row_count == 2


def test_empty_header_only_csv_success(
    processor: Report6Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True, exist_ok=True)
    empty_csv = extracted / "empty.csv"
    empty_csv.write_text("Ref. No.,Mode\n", encoding="utf-8")

    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(processor, "_find_template", lambda: None)

    result = processor.process(source_a_path=empty_csv, report_slug="report6_station")

    assert result.success is True
    assert result.processed_row_count == 0
    assert result.excel_path is not None
    assert result.pdf_path is not None
    assert Path(result.excel_path).exists()
    assert Path(result.pdf_path).exists()


def test_fails_if_no_station_rows(
    processor: Report6Processor,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True, exist_ok=True)
    
    train_only_csv = extracted / "train_only.csv"
    train_only_csv.write_text(
        "Ref. No.,Complaint Date,Train/Station,Mode,Type,Sub Type,Department,Status\n"
        "REF001,2026-07-10,Train A,Train,Type1,SubType1,Dept1,Pending\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=train_only_csv, report_slug="report6_station")

    assert result.success is False
    assert "Station" in (result.error or "")


def test_scr_row_highlighted(
    processor: Report6Processor,
    station_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )

    result = processor.process(source_a_path=station_complaints_csv, report_slug="report6_station")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    
    scr_found = False
    for row_idx in range(3, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value and "south central railway" in str(cell.value).lower():
                scr_found = True
                assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00", "FFFFFF00"}
                break
    
    assert scr_found


def test_output_headers_match_spec(
    processor: Report6Processor,
    station_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.Report6Processor._find_template",
        lambda self: None,
    )

    result = processor.process(source_a_path=station_complaints_csv, report_slug="report6_station")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active

    for col_idx, expected in enumerate(OUTPUT_HEADERS, start=1):
        actual = worksheet.cell(row=2, column=col_idx).value
        assert actual == expected


def test_report_title_contains_station(
    processor: Report6Processor,
    station_complaints_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.extracted_data_dir",
        str(tmp_path / "extracted"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_excel_dir",
        str(tmp_path / "output" / "excel"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.config.output_pdf_dir",
        str(tmp_path / "output" / "pdf"),
    )
    monkeypatch.setattr(
        "app.automation.processing.report6_processor.Report6Processor._find_template",
        lambda self: None,
    )

    result = processor.process(source_a_path=station_complaints_csv, report_slug="report6_station")
    assert result.success is True

    workbook = load_workbook(result.excel_path)
    worksheet = workbook.active
    title = worksheet.cell(row=1, column=1).value
    
    assert "Station" in title or "Report No 6" in title


def test_train_station_data_separation():
    """Verify report5 and report6 processors have different expected modes."""
    from app.automation.processing.report5_processor import Report5Processor
    from app.automation.processing.report6_processor import Report6Processor
    
    r5 = Report5Processor()
    r6 = Report6Processor()
    
    assert r5.expected_mode == "Train"
    assert r6.expected_mode == "Station"
    assert r5.expected_mode != r6.expected_mode
