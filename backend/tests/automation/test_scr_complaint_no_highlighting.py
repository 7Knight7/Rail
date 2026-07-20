"""Tests: Reports 5/6 must not apply SCR yellow row highlighting."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.automation.formatting.scr import cell_has_yellow_fill
from app.automation.processing.report1_processor import Report1Processor
from app.automation.processing.report5_processor import Report5Processor
from app.automation.processing.report6_processor import Report6Processor

FIXTURES_R1 = Path(__file__).resolve().parent.parent / "fixtures" / "report1"
FIXTURES_R5 = Path(__file__).resolve().parent.parent / "fixtures" / "report5"
FIXTURES_R6 = Path(__file__).resolve().parent.parent / "fixtures" / "report6"


def _patch_outputs(monkeypatch: pytest.MonkeyPatch, module: str, tmp_path: Path) -> None:
    monkeypatch.setattr(f"{module}.config.extracted_data_dir", str(tmp_path / "extracted"))
    monkeypatch.setattr(f"{module}.config.output_excel_dir", str(tmp_path / "output" / "excel"))
    monkeypatch.setattr(f"{module}.config.output_pdf_dir", str(tmp_path / "output" / "pdf"))


def _assert_no_yellow_data_rows(worksheet, *, data_start_row: int = 3) -> None:
    for row_idx in range(data_start_row, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            assert not cell_has_yellow_fill(cell), (
                f"Unexpected yellow fill at row {row_idx}, col {col_idx} "
                f"(value={cell.value!r})"
            )


@pytest.fixture
def r5_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    target = extracted / "report5_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R5 / "train_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


@pytest.fixture
def r6_csv(tmp_path: Path) -> Path:
    extracted = tmp_path / "extracted" / "report6_station"
    extracted.mkdir(parents=True)
    target = extracted / "report6_station_complaints_raw.csv"
    target.write_text(
        (FIXTURES_R6 / "station_complaints_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    return target


def test_report5_no_highlight_when_remarks_contain_scr(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report5"
    extracted.mkdir(parents=True)
    r5_csv = extracted / "report5_complaints_raw.csv"
    r5_csv.write_text(
        "Ref. No.,Mode,Registration Date,Train/Station,Type,Sub Type,Department,Status,Zone,Div,"
        "feedbackRemark,trainNameForReport/Station Name,complaintDesc,remarks,userId\n"
        "REF999,Train,15-07-26 19:13,12759,Maintenance,Issue,CHG,Closed,SC,SC,"
        "Feedback text,South Central Railway Express,Coach issue,"
        "Complaint at South Central Railway station,sc_user_sc\n",
        encoding="utf-8",
    )
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    remarks_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=2, column=col_idx).value == "Remarks":
            remarks_col = col_idx
            break
    assert remarks_col is not None
    assert "South Central Railway" in str(ws.cell(row=3, column=remarks_col).value or "")
    _assert_no_yellow_data_rows(ws)


def test_report6_no_highlight_when_zone_and_div_are_sc(
    r6_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report6_processor", tmp_path)
    monkeypatch.setattr(Report6Processor, "_find_template", lambda self: None)
    result = Report6Processor().process(source_a_path=r6_csv, report_slug="report6_station")
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    zone_col = div_col = user_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=2, column=col_idx).value
        if header == "Zone Code":
            zone_col = col_idx
        elif header == "Div Code":
            div_col = col_idx
        elif header == "User ID":
            user_col = col_idx
    assert zone_col and div_col and user_col
    assert any(str(ws.cell(row=r, column=zone_col).value or "") == "SC" for r in range(3, ws.max_row + 1))
    assert any(str(ws.cell(row=r, column=div_col).value or "") == "SC" for r in range(3, ws.max_row + 1))
    assert any("sc" in str(ws.cell(row=r, column=user_col).value or "").lower() for r in range(3, ws.max_row + 1))
    _assert_no_yellow_data_rows(ws)


def test_report5_pdf_has_no_yellow_data_rows(
    r5_csv: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _patch_outputs(monkeypatch, "app.automation.processing.report5_processor", tmp_path)
    monkeypatch.setattr(Report5Processor, "_find_template", lambda self: None)
    result = Report5Processor().process(source_a_path=r5_csv, report_slug="report5")
    assert result.success is True
    pdf_text = Path(result.pdf_path).read_bytes().decode("latin-1", errors="ignore")
    assert "/DeviceRGB" in pdf_text or pdf_text.startswith("%PDF")
    assert " 1 1 0 rg" not in pdf_text  # reportlab yellow stroke/fill tuple unlikely in our table


def test_report1_scr_row_highlighting_unchanged(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    extracted = tmp_path / "extracted" / "report1"
    extracted.mkdir(parents=True)
    comprehensive = extracted / "comprehensive.csv"
    feedback = extracted / "feedback.csv"
    comprehensive.write_text(
        (FIXTURES_R1 / "comprehensive_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    feedback.write_text(
        (FIXTURES_R1 / "feedback_zone_raw.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    _patch_outputs(monkeypatch, "app.automation.processing.report1_processor", tmp_path)
    result = Report1Processor().process(
        source_a_path=comprehensive,
        report_slug="report1",
        source_b_path=feedback,
    )
    assert result.success is True

    ws = load_workbook(result.excel_path).active
    scr_yellow = False
    for row_idx in range(3, ws.max_row):
        org = ws.cell(row=row_idx, column=2).value
        if org and "South Central Railway" in str(org):
            if cell_has_yellow_fill(ws.cell(row=row_idx, column=1)):
                scr_yellow = True
                break
    assert scr_yellow
