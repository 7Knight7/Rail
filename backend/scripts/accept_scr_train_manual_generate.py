"""Live acceptance helper for scr-train process-only manual generation."""

from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

ACCEPTANCE_FILTER = [
    "scr-train.complaint_ref_no",
    "scr-train.created_on",
    "scr-train.user_id",
    "scr-train.comp_type_name",
    "scr-train.complaint_desc",
]


async def main() -> int:
    from app.features.reports.scr_train_manual import (
        _run_scr_train_process_only_worker,
        preflight_scr_train_manual,
    )
    from app.features.reports.schemas import ManualGenerateRequest
    from app.features.reports.service import ManualReportService, build_config_snapshot
    from app.automation.run_registry import create_cdp_run
    from app.infrastructure.database.models import AutomationRunModel
    from app.infrastructure.database.session import SessionLocal
    from openpyxl import load_workbook

    body = ManualGenerateRequest(
        selected_column_ids=ACCEPTANCE_FILTER,
        column_order=ACCEPTANCE_FILTER,
        requested_formats=["xlsx", "pdf"],
    )
    snapshot = build_config_snapshot(body, report_slug="scr-train")
    snapshot["report_slug"] = "scr-train"
    snapshot["generation_mode"] = "process_only"

    source = await preflight_scr_train_manual(snapshot)

    async with SessionLocal() as db:
        run = await create_cdp_run(
            db,
            user_id="live-acceptance",
            trigger_type="manual_report",
            manual_config=snapshot,
        )
        run_id = run.id

    await _run_scr_train_process_only_worker(
        run_id,
        user_id="live-acceptance",
        manual_config=snapshot,
    )

    async with SessionLocal() as db:
        status = await ManualReportService().get_run_status(
            db, run_id, expected_slug="scr-train"
        )
        run = await db.get(AutomationRunModel, run_id)
        report = json.loads(run.result_json or "{}")["result"]["reports"][0]

    excel_path = report.get("excel_path")
    pdf_path = report.get("pdf_path")
    excel_headers: list[str] = []
    excel_rows = 0
    if excel_path:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            rows = list(wb.active.iter_rows(values_only=True))
            excel_headers = [str(c).strip() if c else "" for c in rows[1]]
            excel_rows = len(rows) - 2
        finally:
            wb.close()

    result = {
        "run_id": run_id,
        "source_path": str(source),
        "status": status.status,
        "processing_success": status.processing_success,
        "error": status.error,
        "visible_columns": status.visible_columns,
        "processed_row_count": status.processed_row_count,
        "excel_artifact_id": status.excel_artifact_id,
        "pdf_artifact_id": status.pdf_artifact_id,
        "excel_download_url": status.excel_download_url,
        "pdf_download_url": status.pdf_download_url,
        "pdf_preview_url": status.pdf_preview_url,
        "excel_path": excel_path,
        "pdf_path": pdf_path,
        "excel_headers": excel_headers,
        "excel_data_rows": excel_rows,
        "pdf_starts_with_pdf": Path(pdf_path).read_bytes()[:5] == b"%PDF-"
        if pdf_path
        else False,
        "selected_column_ids": ACCEPTANCE_FILTER,
    }
    print(json.dumps(result, indent=2))
    ok = (
        status.status == "Completed"
        and status.processing_success
        and status.excel_artifact_id
        and status.pdf_artifact_id
        and excel_headers == status.visible_columns
        and result["pdf_starts_with_pdf"]
    )
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
