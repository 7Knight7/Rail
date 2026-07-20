"""Verify sample Excel headers for Reports 1, 2, 5, 6."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1] / "storage" / "output" / "samples" / "excel"


def latest_xlsx(folder: str) -> Path:
    d = BASE / folder
    files = sorted(d.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not files:
        raise FileNotFoundError(f"No xlsx in {d}")
    return files[-1]


def headers(path: Path, header_row: int = 2) -> list[str]:
    ws = load_workbook(path).active
    return [str(ws.cell(row=header_row, column=c).value or "") for c in range(1, ws.max_column + 1)]


def main() -> None:
    checks = {
        "report1": ([], ["Closing Balance"]),
        "report2": ([], ["% Disposal", "% Balance", "Closing Balance"]),
        "scr-train": (
            [
                "Complaint Ref Number",
                "Created On",
                "Comp Type Name",
                "Sub Type Name",
                "Zone Code",
                "Div Code",
                "Feedback Remark",
                "Train Name For Report",
                "Complaint Description",
                "User ID",
            ],
            ["Mode", "Complaint Mode", "Ref. No.", "Status", "Department"],
        ),
        "scr-station": (
            [
                "Complaint Ref Number",
                "Comp Type Name",
                "Sub Type Name",
                "Zone Code",
                "Div Code",
                "Feedback Remark",
                "Complaint Description",
                "User ID",
            ],
            ["Mode", "Complaint Mode", "Complaint Date", "Train Name For Report", "Status", "Department"],
        ),
    }
    for folder, (must_have, must_not) in checks.items():
        path = latest_xlsx(folder)
        hdrs = headers(path)
        print(f"\n{folder}: {path.name} ({len(hdrs)} cols)")
        print(hdrs)
        for label in must_have:
            assert label in hdrs, f"{folder}: missing {label!r}"
        for label in must_not:
            assert label not in hdrs, f"{folder}: stale {label!r} present"
        print("OK")


if __name__ == "__main__":
    main()
